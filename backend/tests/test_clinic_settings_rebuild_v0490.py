from __future__ import annotations

import io
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

from app.schemas.clinic_setup_v2 import ClinicDoctorCreateV2, ClinicServiceCreateV2
from app.schemas.historical_import import HistoricalImportDocument
from app.services.clinic_setup_import import (
    build_clinic_setup_template,
    preview_clinic_setup_workbook,
)
from app.services.historical_import import (
    _assign_patient_phone_if_available,
    _fill_missing_patient_facts,
    _normalize_allocation,
    _normalize_appointment,
    _normalize_package,
    _normalize_patient,
    _normalize_payment,
    _patient_identity,
    build_historical_import_template,
)

ROOT = Path(__file__).resolve().parents[2]
BACKEND = ROOT / "backend"
FRONTEND = ROOT / "frontend"


def _service():
    return SimpleNamespace(id=uuid4(), name="Full Body Laser", duration_minutes=60, price_minor=250000)


def test_setup_contract_uses_price_not_price_minor() -> None:
    fields = ClinicServiceCreateV2.model_fields
    assert "price" in fields
    assert "price_minor" not in fields
    payload = ClinicServiceCreateV2(name="Full Body Laser", duration_minutes=60, price="2500")
    assert str(payload.price) == "2500"


def test_doctor_contract_has_no_active_field() -> None:
    fields = ClinicDoctorCreateV2.model_fields
    assert "active" not in fields
    assert "is_active" not in fields
    assert fields["doctor_type"].default == "regular"


def test_history_document_contract_is_small_and_standalone() -> None:
    assert set(HistoricalImportDocument.model_fields) == {"name", "format", "content_base64"}


def test_patient_phone_identity_is_stable_and_name_is_not_identity() -> None:
    first = _patient_identity({"phone": "010 1234 5678", "full_name": "Sara Ahmed"})
    second = _patient_identity({"phone": "+20 1012345678", "full_name": "Different Name"})
    assert first[0] == second[0]
    assert first[0] and first[0].startswith("phone:")
    payload, code, _message = _normalize_patient({"full_name": "Sara Ahmed"})
    assert payload is None
    assert code == "patient_identity_missing"


def test_append_patient_merge_fills_only_missing_facts() -> None:
    patient = SimpleNamespace(
        first_name="Sara",
        last_name="Ahmed",
        phone="01012345678",
        phone_normalized="201012345678",
        birth_date=None,
        source="instagram",
    )
    _fill_missing_patient_facts(
        patient,
        {
            "full_name": "Different Name",
            "phone": None,
            "phone_normalized": None,
            "birth_date": "1995-04-12",
            "source": "referral",
        },
    )
    assert patient.first_name == "Sara"
    assert patient.last_name == "Ahmed"
    assert patient.phone == "01012345678"
    assert patient.birth_date == date(1995, 4, 12)
    assert patient.source == "instagram"


def test_append_patient_merge_does_not_steal_phone_from_another_patient() -> None:
    workspace_id = uuid4()
    patient = SimpleNamespace(
        id=uuid4(),
        first_name="Sara",
        last_name="Ahmed",
        phone=None,
        phone_normalized=None,
        birth_date=None,
        source="other",
    )
    phone_owner = SimpleNamespace(id=uuid4())

    class FakeDB:
        def scalar(self, _statement):
            return phone_owner

    _fill_missing_patient_facts(
        patient,
        {
            "phone": "01010000696",
            "phone_normalized": "+201010000696",
            "source": "whatsapp",
        },
        db=FakeDB(),
        workspace_id=workspace_id,
    )

    assert patient.phone is None
    assert patient.phone_normalized is None
    assert patient.source == "whatsapp"


def test_patient_phone_assignment_allows_same_canonical_owner() -> None:
    patient = SimpleNamespace(id=uuid4(), phone=None, phone_normalized=None)

    class FakeDB:
        def scalar(self, _statement):
            return SimpleNamespace(id=patient.id)

    assigned = _assign_patient_phone_if_available(
        FakeDB(),
        workspace_id=uuid4(),
        patient=patient,
        phone="01012345678",
        phone_normalized="+201012345678",
    )

    assert assigned is True
    assert patient.phone == "01012345678"
    assert patient.phone_normalized == "+201012345678"


def test_append_patient_payload_changes_do_not_force_replace() -> None:
    source = (BACKEND / "app/services/historical_import.py").read_text(encoding="utf-8")
    append_block = source[source.index('if batch.mode == "append":'):source.index('else:', source.index('if batch.mode == "append":'))]
    assert 'row.entity_type in {"payment_allocation", "patient"}' in append_block
    assert 'link.payload_hash != row.payload_hash' in append_block
    assert 'Appointment/payment/package facts remain strict' in source

def test_patient_id_can_identify_patient_without_phone_or_email() -> None:
    payload, code, _message = _normalize_patient({"patient_id": "P-1001", "full_name": "Sara"})
    assert code is None
    assert payload is not None
    assert payload["identity"] == "id:P-1001"
    assert "email" not in payload


def test_appointment_accepts_separate_date_and_time_and_has_no_end_at() -> None:
    service = _service()
    payload, code, _message = _normalize_appointment(
        {
            "patient_phone": "01012345678",
            "service_name": "Full Body Laser",
            "date": "2026-08-15",
            "start_time": "18:30",
        },
        services_by_id={service.id: service},
        services_by_name={service.name.casefold(): service},
    )
    assert code is None
    assert payload is not None
    assert payload["date"] == "2026-08-15"
    assert payload["start_time"].startswith("18:30")
    assert "end_at" not in payload
    assert payload["source_record_id"].startswith("appointment:")


def test_appointment_requires_patient_and_resolvable_service_only() -> None:
    service = _service()
    payload, code, _ = _normalize_appointment(
        {"service_name": service.name, "date": "2026-08-15", "start_time": "18:30"},
        services_by_id={service.id: service},
        services_by_name={service.name.casefold(): service},
    )
    assert payload is None and code == "appointment_patient_identity_missing"
    payload, code, _ = _normalize_appointment(
        {"patient_phone": "01012345678", "service_name": "Unknown", "date": "2026-08-15", "start_time": "18:30"},
        services_by_id={service.id: service},
        services_by_name={service.name.casefold(): service},
    )
    assert payload is None and code == "appointment_service_unknown"


def test_payment_uses_signed_amount_without_transaction_type() -> None:
    payload, code, _ = _normalize_payment(
        {"patient_phone": "01012345678", "amount": "-750.50", "paid_at": "2026-08-15"},
        source_file="payments.xlsx",
        sheet="payments",
        row_number=17,
    )
    assert code is None
    assert payload is not None
    assert payload["amount_minor"] == -75050
    assert "transaction_type" not in payload
    assert payload["transaction_id"] is None
    assert payload["source_record_id"].startswith("payment:")


def test_generated_transaction_source_id_is_stable_for_same_source_row() -> None:
    args = ({"patient_id": "P1", "amount": "1000", "paid_at": "2026-08-15"},)
    a, _, _ = _normalize_payment(*args, source_file="payments.csv", sheet="payments", row_number=10)
    b, _, _ = _normalize_payment(*args, source_file="payments.csv", sheet="payments", row_number=10)
    c, _, _ = _normalize_payment(*args, source_file="payments.csv", sheet="payments", row_number=11)
    assert a and b and c
    assert a["source_record_id"] == b["source_record_id"]
    assert a["source_record_id"] != c["source_record_id"]


def test_reference_transaction_id_is_optional_and_only_preserves_explicit_fact() -> None:
    payload, code, _ = _normalize_payment(
        {"patient_id": "P1", "amount": "-100", "paid_at": "2026-08-15"},
        source_file="payments.csv",
        sheet="payments",
        row_number=3,
    )
    assert code is None and payload
    assert payload["reference_transaction_id"] is None


def test_allocations_require_explicit_ids() -> None:
    payload, code, _ = _normalize_allocation({"amount": 100})
    assert payload is None and code == "allocation_reference_missing"


def test_package_supports_opening_remaining_without_total() -> None:
    service = _service()
    payload, code, _ = _normalize_package(
        {
            "patient_phone": "01012345678",
            "service_name": service.name,
            "package_name": "Full Body 6",
            "sessions_remaining": 4,
            "price": 10000,
        },
        services_by_id={service.id: service},
        services_by_name={service.name.casefold(): service},
    )
    assert code is None and payload
    assert payload["sessions_remaining"] == 4
    assert payload["sessions_total"] is None
    assert payload["price_minor"] == 1_000_000  # internal implementation detail only


def test_package_negative_remaining_is_rejected() -> None:
    service = _service()
    payload, code, _ = _normalize_package(
        {"patient_id": "P1", "service_name": service.name, "sessions_remaining": -1},
        services_by_id={service.id: service},
        services_by_name={service.name.casefold(): service},
    )
    assert payload is None and code == "package_remaining_invalid"


def test_official_template_is_fixed_contract() -> None:
    wb = load_workbook(io.BytesIO(build_historical_import_template()), read_only=True)
    try:
        assert wb.sheetnames == ["README", "patients", "appointments", "payments", "payment_allocations", "packages"]
        headers = {}
        for name in wb.sheetnames[1:]:
            headers[name] = [cell.value for cell in next(wb[name].iter_rows(min_row=1, max_row=1))]
        flattened = {str(value) for values in headers.values() for value in values}
        assert "email" not in flattened
        assert "gender" not in flattened
        assert "currency" not in flattened
        assert "end_at" not in flattened
        assert "transaction_type" not in flattened
        assert "price_minor" not in flattened
        assert "active" not in flattened
        assert "price" in headers["packages"]
        assert "doctor_id" in headers["appointments"]
        assert "sessions_remaining" in headers["packages"]
    finally:
        wb.close()


def test_old_universal_onboarding_routes_are_gone() -> None:
    onboarding = (BACKEND / "app/api/routes/onboarding.py").read_text(encoding="utf-8")
    clinic = (BACKEND / "app/api/routes/clinic.py").read_text(encoding="utf-8")
    assert "/integration/sessions" not in onboarding
    assert "/integration/import/preview" not in clinic
    assert "/integration/import/apply" not in clinic


def test_history_apply_is_background_and_pollable() -> None:
    source = (BACKEND / "app/api/routes/clinic_setup_v2.py").read_text(encoding="utf-8")
    assert "status_code=status.HTTP_202_ACCEPTED" in source
    assert "background_tasks.add_task" in source
    assert '@router.get("/history/batches/{batch_id}"' in source


def test_replace_never_deletes_patients_or_workspace_wide_appointments() -> None:
    source = (BACKEND / "app/services/historical_import.py").read_text(encoding="utf-8")
    block = source[source.index("def _safe_remove_previous_imports"):source.index("def apply_historical_import")]
    assert "delete(Patient)" not in block
    assert "AgentAction.appointment_id.in_(appointment_ids)" in block
    assert ".values(appointment_id=None)" in block
    assert "workspace_id=None" not in block
    assert "Appointment.id.in_(appointment_ids)" in block


def test_historical_doctors_are_passive_until_admin_configures_availability() -> None:
    source = (BACKEND / "app/services/historical_import.py").read_text(encoding="utf-8")
    doctor_block = source[source.index("def _create_visiting_doctor"):source.index("def _unassigned_historical_doctor")]
    assert "doctor_type=\"visiting\"" in doctor_block
    assert "booking_enabled=False" in doctor_block


def test_visiting_doctors_use_dated_windows_in_booking_engine() -> None:
    source = (BACKEND / "app/services/booking.py").read_text(encoding="utf-8")
    assert "DoctorAvailabilityWindow" in source
    assert 'if doctor.doctor_type == "visiting"' in source
    assert "_window_intersections" in source


def test_primary_branch_is_used_as_single_branch_boundary() -> None:
    source = (BACKEND / "app/agents/tools/clinic_tools.py").read_text(encoding="utf-8")
    adapter = (BACKEND / "app/integrations/clinic/tia_database.py").read_text(encoding="utf-8")
    assert "primary_branch_id" in source
    assert "primary_branch_id" in adapter


def test_migration_removes_legacy_session_tables_and_creates_history_staging() -> None:
    source = (BACKEND / "alembic/versions/0051_clinic_setup_v2.py").read_text(encoding="utf-8")
    assert "DROP TABLE IF EXISTS clinic_integration_onboarding_events" in source
    assert "DROP TABLE IF EXISTS clinic_integration_onboarding_sessions" in source
    assert "DROP COLUMN IF EXISTS onboarding_session_id" in source
    assert '"clinic_historical_import_batches"' in source
    assert '"clinic_historical_import_rows"' in source
    assert '"clinic_historical_import_links"' in source
    assert "primary_branch_id" in source
    assert "doctor_availability_windows" in source


def test_migration_tolerates_legacy_payment_reference_constraint_drift() -> None:
    source = (BACKEND / "alembic/versions/0051_clinic_setup_v2.py").read_text(encoding="utf-8")
    assert '"ck_payment_transactions_payment_transaction_reference_valid"' in source
    assert '"payment_transaction_reference_valid"' in source
    assert "DROP CONSTRAINT IF EXISTS" in source
    assert (
        "ADD CONSTRAINT ck_payment_transactions_payment_transaction_reference_valid"
        in source
    )


def test_frontend_no_longer_uses_mapping_wizard_contract() -> None:
    page = (FRONTEND / "src/app/(dashboard)/setup/integration/page.tsx").read_text(encoding="utf-8")
    uploader = (FRONTEND / "src/app/(dashboard)/setup/integration/history-uploader.tsx").read_text(encoding="utf-8")
    setup_page = (FRONTEND / "src/app/(dashboard)/setup/page.tsx").read_text(encoding="utf-8")
    combined = page + uploader + setup_page
    assert "IntegrationWizard" not in combined
    assert "price_minor" not in combined
    assert "doctor.is_active" not in combined
    assert "replace_previous_imports" in combined
    assert "append" in combined
    assert "clinic-history-template" in combined


def test_history_import_is_bounded_and_setup_mutations_are_audited() -> None:
    history = (BACKEND / "app/services/historical_import.py").read_text(encoding="utf-8")
    route = (BACKEND / "app/api/routes/clinic_setup_v2.py").read_text(encoding="utf-8")
    frontend_actions = (FRONTEND / "src/app/(dashboard)/setup/integration/actions.ts").read_text(encoding="utf-8")
    assert "MAX_DOCUMENT_BYTES = 25 * 1024 * 1024" in history
    assert "MAX_UPLOAD_BYTES = 60 * 1024 * 1024" in history
    assert "MAX_IMPORT_ROWS = 250_000" in history
    assert "processed_rows > MAX_IMPORT_ROWS" in history
    assert "files.length > 10" in frontend_actions
    for action in (
        "clinic.profile_updated",
        "clinic.service_created",
        "clinic.doctor_created",
        "clinic.hours_updated",
        "clinic.booking_policy_updated",
    ):
        assert action in route
    assert "clinic.history_imported" in history


def test_history_ui_resumes_durable_preview_import_and_failed_batches() -> None:
    page = (FRONTEND / "src/app/(dashboard)/setup/integration/page.tsx").read_text(encoding="utf-8")
    uploader = (FRONTEND / "src/app/(dashboard)/setup/integration/history-uploader.tsx").read_text(encoding="utf-8")
    assert '["importing", "preview_ready", "failed"].includes(batch.status)' in page
    assert "batch && !preview" in uploader
    assert "إعادة محاولة الاستيراد" in uploader
    assert "readHistoricalImportBatchAction" in uploader


def test_setup_excel_template_is_fixed_and_uploadable_without_mapping() -> None:
    wb = load_workbook(io.BytesIO(build_clinic_setup_template()), read_only=True)
    try:
        assert wb.sheetnames == [
            "README", "clinic_profile", "services", "doctors", "doctor_services",
            "clinic_hours", "doctor_hours", "visiting_windows", "booking_policy",
        ]
        services_headers = [cell.value for cell in next(wb["services"].iter_rows(min_row=1, max_row=1))]
        doctor_headers = [cell.value for cell in next(wb["doctors"].iter_rows(min_row=1, max_row=1))]
        clinic_hours_headers = [cell.value for cell in next(wb["clinic_hours"].iter_rows(min_row=1, max_row=1))]
        doctor_hours_headers = [cell.value for cell in next(wb["doctor_hours"].iter_rows(min_row=1, max_row=1))]
        assert services_headers == ["name", "category", "duration_minutes", "price"]
        assert doctor_headers == ["full_name", "doctor_type", "specialization"]
        assert clinic_hours_headers == ["day", "start_time", "end_time"]
        assert doctor_hours_headers == ["doctor_name", "day", "start_time", "end_time"]
        assert "weekday" not in clinic_hours_headers
        assert "weekday" not in doctor_hours_headers
        assert "price_minor" not in services_headers
        assert "active" not in doctor_headers
    finally:
        wb.close()


def test_setup_ui_uses_blank_review_draft_and_clear_next_step() -> None:
    page = (FRONTEND / "src/app/(dashboard)/setup/page.tsx").read_text(encoding="utf-8")
    importer = (FRONTEND / "src/app/(dashboard)/setup/setup-importer.tsx").read_text(encoding="utf-8")
    actions = (FRONTEND / "src/app/(dashboard)/setup/actions.ts").read_text(encoding="utf-8")
    route = (BACKEND / "app/api/routes/clinic_setup_v2.py").read_text(encoding="utf-8")
    assert "ClinicSetupImporter" in page
    assert 'accept=".xlsx"' in importer
    assert "الخانات تبدأ فاضية" in importer
    assert "قراءة الملف وتعبئة الخانات" in importer
    assert "تحميل البيانات المحفوظة" in importer
    assert "التالي: البيانات التاريخية" in importer
    assert "emptyDraft" in importer
    assert '"/clinic/setup-v2/preview"' in actions
    assert '"/clinic/setup-v2/apply-draft"' in actions
    assert '@router.post("/setup-v2/preview"' in route
    assert '@router.post("/setup-v2/apply-draft"' in route
    assert '@router.get("/setup-v2/template")' in route



def test_setup_preview_preserves_missing_cells_as_blank_without_weekday() -> None:
    import base64

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "doctor_hours"
    ws.append(["doctor_name", "day", "start_time", "end_time"])
    ws.append(["Dr. Mariam", "الثلاثاء", "10:00", None])
    services = wb.create_sheet("services")
    services.append(["name", "category", "duration_minutes", "price"])
    services.append(["PRP Face", None, 45, None])
    output = io.BytesIO()
    wb.save(output)

    result = preview_clinic_setup_workbook(
        filename="setup.xlsx",
        content_base64=base64.b64encode(output.getvalue()).decode("ascii"),
    )
    assert result.draft.doctor_hours[0]["day"] == "الثلاثاء"
    assert "weekday" not in result.draft.doctor_hours[0]
    assert result.draft.doctor_hours[0]["end_time"] is None
    assert result.draft.services[0]["category"] is None
    assert result.draft.services[0]["price"] is None


def test_setup_template_booking_policy_values_are_blank() -> None:
    wb = load_workbook(io.BytesIO(build_clinic_setup_template()), data_only=True)
    try:
        rows = list(wb["booking_policy"].iter_rows(min_row=2, values_only=True))
        assert rows
        assert all(value in (None, "") for _setting, value in rows)
    finally:
        wb.close()

def test_setup_excel_optional_columns_are_tolerant_and_preserve_required_runtime_fields() -> None:
    import base64

    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from app.database.base import Base
    from app.models.booking_settings import BookingSettings
    from app.models.branch import Branch
    from app.models.doctor import Doctor
    from app.models.doctor_branch import DoctorBranch
    from app.models.doctor_service import DoctorService
    from app.models.service import Service
    from app.models.staff import Staff
    from app.models.user import User
    from app.models.working_hours import (
        BranchWorkingHour,
        DoctorAvailabilityWindow,
        DoctorWorkingHour,
    )
    from app.models.workspace import Workspace
    from app.services.clinic_setup_import import import_clinic_setup_workbook

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(
        engine,
        tables=[
            Workspace.__table__, Branch.__table__, BookingSettings.__table__, Service.__table__,
            User.__table__, Staff.__table__, Doctor.__table__, DoctorBranch.__table__,
            DoctorService.__table__, BranchWorkingHour.__table__, DoctorWorkingHour.__table__,
            DoctorAvailabilityWindow.__table__,
        ],
    )

    workbook = load_workbook(io.BytesIO(build_clinic_setup_template()))
    workbook["clinic_profile"]["B2"] = "Flexible Clinic"
    workbook["services"].append(["Hydrafacial", None, 60, 1800])
    workbook["doctors"].append(["Dr. Mariam Adel", "regular", None])
    workbook["doctor_services"].append(["Dr. Mariam Adel", "Hydrafacial"])
    workbook["clinic_hours"].append(["Monday", "10:00", "22:00"])
    workbook["doctor_hours"].append(["Dr. Mariam Adel", "الإثنين", "12:00", "20:00"])
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    workspace = SimpleNamespace(id=uuid4(), name="Test Workspace", primary_branch_id=None, timezone="Africa/Cairo")
    with Session(engine) as db:
        result = import_clinic_setup_workbook(
            db,
            workspace=workspace,
            filename="setup.xlsx",
            content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
        )
        db.commit()
        service = db.scalar(select(Service).where(Service.name == "Hydrafacial"))
        assert service is not None
        assert service.price_minor == 180000
        doctor = db.scalar(select(Doctor))
        assert doctor is not None and doctor.doctor_type == "regular"
        assert db.scalar(select(BranchWorkingHour).where(BranchWorkingHour.weekday == 0)) is not None
        assert db.scalar(select(DoctorWorkingHour).where(DoctorWorkingHour.weekday == 0)) is not None
        assert result.skipped_counts["services"] == 0
        assert result.skipped_counts["doctors"] == 0


def test_setup_apply_does_not_invent_defaults_for_incomplete_new_rows() -> None:
    import base64

    from openpyxl import Workbook
    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from app.database.base import Base
    from app.models.booking_settings import BookingSettings
    from app.models.branch import Branch
    from app.models.doctor import Doctor
    from app.models.doctor_branch import DoctorBranch
    from app.models.doctor_service import DoctorService
    from app.models.service import Service
    from app.models.staff import Staff
    from app.models.user import User
    from app.models.working_hours import (
        BranchWorkingHour,
        DoctorAvailabilityWindow,
        DoctorWorkingHour,
    )
    from app.models.workspace import Workspace
    from app.services.clinic_setup_import import import_clinic_setup_workbook

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(
        engine,
        tables=[
            Workspace.__table__, Branch.__table__, BookingSettings.__table__, Service.__table__,
            User.__table__, Staff.__table__, Doctor.__table__, DoctorBranch.__table__,
            DoctorService.__table__, BranchWorkingHour.__table__, DoctorWorkingHour.__table__,
            DoctorAvailabilityWindow.__table__,
        ],
    )
    workbook = Workbook()
    profile = workbook.active
    profile.title = "clinic_profile"
    profile.append(["field", "value"])
    profile.append(["clinic_name", "Draft Clinic"])
    services = workbook.create_sheet("services")
    services.append(["name", "category", "duration_minutes", "price"])
    services.append(["Missing Price", None, 45, None])
    doctors = workbook.create_sheet("doctors")
    doctors.append(["full_name", "doctor_type", "specialization"])
    doctors.append(["Dr. Missing Type", None, None])
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    workspace = SimpleNamespace(id=uuid4(), name="Workspace Name", primary_branch_id=None, timezone="Africa/Cairo")
    with Session(engine) as db:
        result = import_clinic_setup_workbook(
            db,
            workspace=workspace,
            filename="setup.xlsx",
            content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
        )
        db.commit()
        assert db.scalar(select(Service).where(Service.name == "Missing Price")) is None
        assert db.scalar(select(Doctor)) is None
        assert result.skipped_counts["services"] == 1
        assert result.skipped_counts["doctors"] == 1
        assert any(issue.sheet == "services" and "price" in issue.message for issue in result.issues)
        assert any(issue.sheet == "doctors" and "doctor_type" in issue.message for issue in result.issues)


def test_setup_ui_has_one_draft_save_action_and_one_next_step_location() -> None:
    page = (FRONTEND / "src/app/(dashboard)/setup/page.tsx").read_text(encoding="utf-8")
    importer = (FRONTEND / "src/app/(dashboard)/setup/setup-importer.tsx").read_text(encoding="utf-8")
    actions = (FRONTEND / "src/app/(dashboard)/setup/actions.ts").read_text(encoding="utf-8")
    assert "حفظ إعدادات العيادة" in importer
    assert "applyClinicSetupDraftAction" in actions
    assert "saveClinicProfileV2" not in actions
    assert "saveClinicHoursV2" not in actions
    assert "saveRegularDoctorHoursV2" not in actions
    assert "saveVisitingDoctorWindowsV2" not in actions
    assert "saveBookingPolicyV2" not in actions
    assert "التالي: البيانات التاريخية" not in page
    assert "dirty" in importer
    assert "تعديلات غير محفوظة" in importer


def test_historical_patient_contract_has_no_gender_and_defaults_to_female() -> None:
    payload, code, _ = _normalize_patient({
        "patient_id": "P1",
        "full_name": "Sara Ahmed",
        "gender": "male",
    })
    assert code is None and payload is not None
    assert "gender" not in payload
    source = (BACKEND / "app/services/historical_import.py").read_text(encoding="utf-8")
    patient_block = source[source.index("def _patient_for_identity"):source.index("def _doctor_catalog")]
    assert 'gender="female"' in patient_block


def test_setup_excel_import_applies_valid_rows_and_keeps_invalid_rows_editable() -> None:
    import base64

    from sqlalchemy import create_engine, select
    from sqlalchemy.orm import Session

    from app.database.base import Base
    from app.models.booking_settings import BookingSettings
    from app.models.branch import Branch
    from app.models.doctor import Doctor
    from app.models.doctor_branch import DoctorBranch
    from app.models.doctor_service import DoctorService
    from app.models.service import Service
    from app.models.staff import Staff
    from app.models.user import User
    from app.models.working_hours import (
        BranchWorkingHour,
        DoctorAvailabilityWindow,
        DoctorWorkingHour,
    )
    from app.models.workspace import Workspace
    from app.services.clinic_setup_import import import_clinic_setup_workbook

    engine = create_engine("sqlite+pysqlite:///:memory:")
    Base.metadata.create_all(
        engine,
        tables=[
            Workspace.__table__, Branch.__table__, BookingSettings.__table__, Service.__table__,
            User.__table__, Staff.__table__, Doctor.__table__, DoctorBranch.__table__,
            DoctorService.__table__, BranchWorkingHour.__table__, DoctorWorkingHour.__table__,
            DoctorAvailabilityWindow.__table__,
        ],
    )

    workbook = load_workbook(io.BytesIO(build_clinic_setup_template()))
    workbook["clinic_profile"]["B2"] = "Tia Test Clinic"
    services = workbook["services"]
    services.append(["PRP Face", "Skin", 45, 2000])
    services.append(["Broken Service", "Skin", None, 1200])
    doctors = workbook["doctors"]
    doctors.append(["Dr. Salma Nabil", "regular", "Dermatology"])
    doctor_services = workbook["doctor_services"]
    doctor_services.append(["Dr. Salma Nabil", "PRP Face"])
    clinic_hours = workbook["clinic_hours"]
    clinic_hours.append(["Tuesday", "10:00", "22:00"])
    doctor_hours = workbook["doctor_hours"]
    doctor_hours.append(["Dr. Salma Nabil", "الثلاثاء", "12:00", "20:00"])
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()

    workspace = SimpleNamespace(id=uuid4(), name="Test Workspace", primary_branch_id=None, timezone="Africa/Cairo")
    with Session(engine) as db:
        result = import_clinic_setup_workbook(
            db,
            workspace=workspace,
            filename="setup.xlsx",
            content_base64=base64.b64encode(stream.getvalue()).decode("ascii"),
        )
        db.commit()

        assert result.imported_counts["clinic_profile"] == 1
        assert result.imported_counts["services"] == 1
        assert result.skipped_counts["services"] == 1
        assert any(issue.sheet == "services" and issue.message for issue in result.issues)
        assert db.scalar(select(Service).where(Service.name == "PRP Face")).price_minor == 200000
        salma = db.execute(
            select(Doctor, Staff)
            .join(Staff, (Staff.workspace_id == Doctor.workspace_id) & (Staff.id == Doctor.staff_id))
            .where(Staff.first_name == "Dr.", Staff.last_name == "Salma Nabil")
        ).first()
        assert salma is not None
        assert db.scalar(select(BranchWorkingHour).where(BranchWorkingHour.weekday == 1)) is not None
        assert db.scalar(select(DoctorWorkingHour).where(DoctorWorkingHour.weekday == 1)) is not None


def test_payment_reference_constraint_repair_is_expression_based() -> None:
    source = (BACKEND / "alembic/versions/0052_payment_reference_constraint_repair.py").read_text(encoding="utf-8")
    assert 'down_revision: str | Sequence[str] | None = "0051_clinic_setup_v2"' in source
    assert "pg_get_constraintdef(con.oid) ILIKE '%reference_transaction_id%'" in source
    assert "ALTER TABLE alembic_version" in source
    assert "ALTER COLUMN version_num TYPE VARCHAR(255)" in source
    assert source.index("_widen_alembic_version_column()") < source.index("_drop_reference_checks()", source.index("def upgrade"))
    assert "ck_payment_transactions_reference_valid_v2" in source
    assert "transaction_type = 'refund' OR reference_transaction_id IS NULL" in source
    assert "ck_payment_transactions_ck_payment_transactions" not in source


def test_payment_model_matches_repaired_reference_constraint() -> None:
    source = (BACKEND / "app/models/payment_transaction.py").read_text(encoding="utf-8")
    assert 'name="reference_valid_v2"' in source
    assert "transaction_type = 'refund' OR reference_transaction_id IS NULL" in source
