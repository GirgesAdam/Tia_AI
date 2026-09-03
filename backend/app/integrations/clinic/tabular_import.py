from __future__ import annotations

import base64
import csv
import hashlib
import io
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime, time
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from app.core.doctor_names import normalize_doctor_display_name
from app.integrations.clinic.structural_transform import (
    StructuralTransformError,
    apply_structural_transforms,
)
from app.schemas.clinic_import import (
    AppointmentSheetMapping,
    AppointmentSource,
    BillingContext,
    ClinicImportCapabilities,
    ClinicImportDocument,
    ClinicImportIssue,
    ClinicImportMapping,
    ClinicImportOverrides,
    ClinicImportPreviewResponse,
    DoctorSheetMapping,
    NormalizedAppointmentImport,
    NormalizedBranchImport,
    NormalizedDoctorImport,
    NormalizedPackageImport,
    NormalizedPackageUsageImport,
    NormalizedPaymentAllocationImport,
    NormalizedPaymentImport,
    NormalizedServiceImport,
    NormalizedWorkingHourImport,
    PaymentMethod,
    PaymentStatus,
)
from app.schemas.crm import normalize_patient_identity_phone

MAX_DOCUMENT_COUNT = 100
MAX_DOCUMENT_BYTES = 10 * 1024 * 1024
MAX_TOTAL_DOCUMENT_BYTES = 50 * 1024 * 1024
MAX_TOTAL_ROWS = 100_000

UNKNOWN_BRANCH_EXTERNAL_ID = "__tia_unknown_branch__"
UNKNOWN_DOCTOR_EXTERNAL_ID = "__tia_unknown_doctor__"


class ClinicImportError(ValueError):
    pass


@dataclass(frozen=True)
class TabularWorkbook:
    sheets: dict[str, list[dict[str, Any]]]
    document_names: tuple[str, ...]


def _appointment_mappings(mapping: ClinicImportMapping) -> list[AppointmentSheetMapping]:
    """Return every configured appointment source in stable import order."""

    return [
        item for item in [mapping.appointments, *mapping.appointment_sources] if item is not None
    ]


def _decode_document(document: ClinicImportDocument) -> bytes:
    try:
        payload = base64.b64decode(document.content_base64, validate=True)
    except Exception as exc:
        raise ClinicImportError(f"{document.name}: content_base64 is invalid.") from exc
    if not payload:
        raise ClinicImportError(f"{document.name}: document is empty.")
    if len(payload) > MAX_DOCUMENT_BYTES:
        raise ClinicImportError(
            f"{document.name}: document exceeds {MAX_DOCUMENT_BYTES // (1024 * 1024)} MB limit."
        )
    return payload


def _clean_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _csv_rows(payload: bytes, *, sheet_name: str) -> list[dict[str, Any]]:
    text: str
    for encoding in ("utf-8-sig", "utf-8"):
        try:
            text = payload.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ClinicImportError(f"{sheet_name}: CSV must be UTF-8 encoded.")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ClinicImportError(f"{sheet_name}: CSV has no header row.")
    headers = [str(name or "").strip() for name in reader.fieldnames]
    if any(not header for header in headers):
        raise ClinicImportError(f"{sheet_name}: CSV contains an empty header name.")
    if len(set(headers)) != len(headers):
        raise ClinicImportError(f"{sheet_name}: CSV contains duplicate header names.")
    rows: list[dict[str, Any]] = []
    for raw in reader:
        row = {str(key).strip(): _clean_cell(value) for key, value in raw.items() if key is not None}
        if any(value is not None for value in row.values()):
            rows.append(row)
    return rows


def _xlsx_sheets(payload: bytes, *, document_name: str) -> dict[str, list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - deployment guard
        raise ClinicImportError(
            "XLSX import requires openpyxl. Install backend requirements."
        ) from exc
    try:
        workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise ClinicImportError(f"{document_name}: invalid XLSX workbook.") from exc

    result: dict[str, list[dict[str, Any]]] = {}
    for sheet in workbook.worksheets:
        values = sheet.iter_rows(values_only=True)
        try:
            first = next(values)
        except StopIteration:
            result[sheet.title] = []
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in first]
        while headers and not headers[-1]:
            headers.pop()
        if not headers:
            result[sheet.title] = []
            continue
        if any(not header for header in headers):
            raise ClinicImportError(
                f"{document_name}/{sheet.title}: workbook contains an empty header name."
            )
        if len(set(headers)) != len(headers):
            raise ClinicImportError(
                f"{document_name}/{sheet.title}: workbook contains duplicate header names."
            )
        rows: list[dict[str, Any]] = []
        for raw in values:
            raw = raw[: len(headers)]
            row = {headers[i]: _clean_cell(raw[i] if i < len(raw) else None) for i in range(len(headers))}
            if any(value is not None for value in row.values()):
                rows.append(row)
        result[sheet.title] = rows
    return result


def validate_tabular_document_batch(documents: list[ClinicImportDocument]) -> None:
    if not documents:
        raise ClinicImportError("At least one tabular document is required.")
    if len(documents) > MAX_DOCUMENT_COUNT:
        raise ClinicImportError(
            f"Upload exceeds the {MAX_DOCUMENT_COUNT} document limit."
        )
    total_bytes = 0
    for document in documents:
        payload = _decode_document(document)
        total_bytes += len(payload)
        if total_bytes > MAX_TOTAL_DOCUMENT_BYTES:
            raise ClinicImportError(
                "Uploaded documents exceed the "
                f"{MAX_TOTAL_DOCUMENT_BYTES // (1024 * 1024)} MB total-size limit."
            )


def load_tabular_documents(documents: list[ClinicImportDocument]) -> TabularWorkbook:
    if not documents:
        raise ClinicImportError("At least one tabular document is required.")
    if len(documents) > MAX_DOCUMENT_COUNT:
        raise ClinicImportError(
            f"Upload exceeds the {MAX_DOCUMENT_COUNT} document limit."
        )
    sheets: dict[str, list[dict[str, Any]]] = {}
    names: list[str] = []
    total_rows = 0
    total_bytes = 0
    for document in documents:
        payload = _decode_document(document)
        total_bytes += len(payload)
        if total_bytes > MAX_TOTAL_DOCUMENT_BYTES:
            raise ClinicImportError(
                "Uploaded documents exceed the "
                f"{MAX_TOTAL_DOCUMENT_BYTES // (1024 * 1024)} MB total-size limit."
            )
        names.append(document.name)
        if document.format == "csv":
            sheet_name = Path(document.name).stem.strip() or "Sheet1"
            loaded = {sheet_name: _csv_rows(payload, sheet_name=document.name)}
        else:
            loaded = _xlsx_sheets(payload, document_name=document.name)
        for sheet_name, rows in loaded.items():
            if sheet_name in sheets:
                raise ClinicImportError(
                    f"Duplicate sheet name {sheet_name!r} across uploaded documents."
                )
            total_rows += len(rows)
            if total_rows > MAX_TOTAL_ROWS:
                raise ClinicImportError(
                    f"Import exceeds the {MAX_TOTAL_ROWS} row preview limit."
                )
            sheets[sheet_name] = rows
    return TabularWorkbook(sheets=sheets, document_names=tuple(names))


def _required(row: dict[str, Any], column: str, path: str) -> str:
    if column not in row:
        raise ClinicImportError(f"{path}: mapped column {column!r} does not exist.")
    value = row.get(column)
    if value is None or not str(value).strip():
        raise ClinicImportError(f"{path}: required value {column!r} is empty.")
    return str(value).strip()


def _optional(row: dict[str, Any], column: str | None) -> str | None:
    if not column:
        return None
    if column not in row:
        raise ClinicImportError(f"Mapped column {column!r} does not exist.")
    value = row.get(column)
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def _raw_optional(row: dict[str, Any], column: str | None) -> Any | None:
    if not column:
        return None
    if column not in row:
        raise ClinicImportError(f"Mapped column {column!r} does not exist.")
    return row.get(column)


def _positive_int(value: str, path: str) -> int:
    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ClinicImportError(f"{path}: expected an integer, got {value!r}.") from exc
    if not 1 <= parsed <= 1440:
        raise ClinicImportError(f"{path}: duration must be between 1 and 1440 minutes.")
    return parsed


def _positive_count(value: str, path: str, *, max_value: int = 10000) -> int:
    try:
        parsed = int(str(value).strip())
    except ValueError as exc:
        raise ClinicImportError(f"{path}: expected a positive integer, got {value!r}.") from exc
    if not 1 <= parsed <= max_value:
        raise ClinicImportError(f"{path}: value must be between 1 and {max_value}.")
    return parsed


def _price_minor(value: str | float | int, path: str) -> int:
    try:
        amount = Decimal(str(value).replace(",", "").strip())
    except InvalidOperation as exc:
        raise ClinicImportError(f"{path}: invalid price {value!r}.") from exc
    if amount < 0:
        raise ClinicImportError(f"{path}: price cannot be negative.")
    return int((amount * Decimal("100")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _timezone(timezone_name: str | None, *, path: str) -> ZoneInfo:
    if not timezone_name:
        raise ClinicImportError(f"{path}: appointment default_timezone is required for local times.")
    try:
        return ZoneInfo(timezone_name)
    except Exception as exc:
        raise ClinicImportError(f"{path}: unknown timezone {timezone_name!r}.") from exc


def _normalize_ampm(text: str) -> str:
    replacements = {
        "مساءً": "PM",
        "مساء": "PM",
        "م": "PM",
        "صباحًا": "AM",
        "صباحا": "AM",
        "صباح": "AM",
        "ص": "AM",
    }
    result = text.strip()
    for source, target in replacements.items():
        result = re.sub(rf"(?<!\w){re.escape(source)}(?!\w)", target, result, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", result).strip()


def _date_value(value: Any, path: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    try:
        return date.fromisoformat(text)
    except ValueError:
        pass
    for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ClinicImportError(f"{path}: unsupported date value {value!r}.")


def _time_value(value: Any, path: str) -> time:
    if isinstance(value, datetime):
        return value.timetz().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    text = _normalize_ampm(str(value))
    try:
        return time.fromisoformat(text)
    except ValueError:
        pass
    for pattern in ("%I:%M %p", "%I %p", "%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text.upper(), pattern).time()
        except ValueError:
            continue
    raise ClinicImportError(f"{path}: unsupported time value {value!r}.")


def _datetime_value(value: Any, *, timezone_name: str | None, path: str) -> datetime:
    if isinstance(value, datetime):
        parsed = value
    else:
        text = _normalize_ampm(str(value))
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError:
            parsed = None
            for pattern in (
                "%d/%m/%Y %H:%M",
                "%d-%m-%Y %H:%M",
                "%d/%m/%Y %I:%M %p",
                "%d-%m-%Y %I:%M %p",
                "%d/%m/%y %H:%M",
                "%d/%m/%y %I:%M %p",
            ):
                try:
                    parsed = datetime.strptime(text.upper(), pattern)
                    break
                except ValueError:
                    continue
            if parsed is None:
                raise ClinicImportError(
                    f"{path}: unsupported datetime value {value!r}."
                ) from None
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        parsed = parsed.replace(tzinfo=_timezone(timezone_name, path=path))
    return parsed


def _weekday(value: str, path: str) -> int:
    text = str(value).strip().lower()
    names = {
        "mon": 0, "monday": 0, "الاثنين": 0, "الإثنين": 0,
        "tue": 1, "tuesday": 1, "الثلاثاء": 1,
        "wed": 2, "wednesday": 2, "الأربعاء": 2, "الاربعاء": 2,
        "thu": 3, "thursday": 3, "الخميس": 3,
        "fri": 4, "friday": 4, "الجمعة": 4,
        "sat": 5, "saturday": 5, "السبت": 5,
        "sun": 6, "sunday": 6, "الأحد": 6, "الاحد": 6,
    }
    if text in names:
        return names[text]
    try:
        number = int(text)
    except ValueError as exc:
        raise ClinicImportError(f"{path}: unknown weekday {value!r}.") from exc
    if 0 <= number <= 6:
        return number
    raise ClinicImportError(f"{path}: weekday must be 0..6 or a weekday name.")


def _time_text(value: str, path: str) -> str:
    parsed = _time_value(value, path)
    return parsed.replace(second=0, microsecond=0).isoformat(timespec="minutes")


def _rows(workbook: TabularWorkbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheets:
        raise ClinicImportError(f"Mapped sheet {sheet_name!r} was not found.")
    return workbook.sheets[sheet_name]


def _split_ids(value: str, delimiter: str) -> list[str]:
    return [item.strip() for item in value.split(delimiter) if item.strip()]


def _identity_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value)).casefold().strip()
    text = text.replace("_", " ")
    text = re.sub(r"[^\w\s]", " ", text, flags=re.UNICODE)
    text = re.sub(r"\s+", " ", text)
    return text


def _generated_source_key(kind: str, *parts: str) -> str:
    material = "\x1f".join(_identity_text(part) for part in parts if str(part).strip())
    digest = hashlib.sha256(material.encode("utf-8")).hexdigest()[:24]
    return f"auto:{kind}:{digest}"


def _entity_source_key(
    row: dict[str, Any],
    *,
    external_id_column: str | None,
    fallback_parts: tuple[str, ...],
    kind: str,
    path: str,
) -> tuple[str, bool]:
    raw = _optional(row, external_id_column)
    if raw:
        return raw, False
    if not any(str(part).strip() for part in fallback_parts):
        raise ClinicImportError(f"{path}: cannot derive a stable {kind} identity.")
    return _generated_source_key(kind, *fallback_parts), True


def _catalog_name_aliases(row: dict[str, Any], primary_name_column: str) -> list[str]:
    """Return conservative alternate display-name columns from one catalog row.

    Clinics often export Arabic and English names side-by-side while the mapping has one
    primary ``name`` field. Exact values from sibling *name* columns are safe aliases;
    addresses, specialties, categories, and free text are deliberately excluded.
    """

    values: list[str] = []
    for column, raw in row.items():
        if raw is None:
            continue
        header = _identity_text(column)
        if column == primary_name_column or "name" in header.split() or "اسم" in header.split():
            text = str(raw).strip()
            if text:
                values.append(text)
    return values


def _alias_map(
    items: list[Any],
    *,
    extra_aliases: dict[str, list[str]] | None = None,
) -> dict[str, str | None]:
    aliases: dict[str, str | None] = {}
    extra_aliases = extra_aliases or {}
    for item in items:
        raw_aliases = [item.external_id, item.name, *extra_aliases.get(item.external_id, [])]
        for raw in raw_aliases:
            key = _identity_text(raw)
            if not key:
                continue
            current = aliases.get(key)
            if current is None and key in aliases:
                continue
            if current is not None and current != item.external_id:
                aliases[key] = None
            else:
                aliases[key] = item.external_id
    return aliases


def _infer_branch_from_source_name(
    *,
    sheet_name: str,
    branches: list[NormalizedBranchImport],
    extra_aliases: dict[str, list[str]],
) -> str | None:
    """Resolve a per-source branch only from explicit source-name evidence.

    This is deliberately conservative: a branch must have an exact normalized alias phrase
    present in the sheet/file-derived name, and the match must identify exactly one branch.
    No fuzzy distance or model guess is used here.
    """

    source = _identity_text(sheet_name)
    if not source:
        return None
    source_tokens = source.split()
    candidates: set[str] = set()
    for branch in branches:
        for raw in [branch.external_id, branch.name, *extra_aliases.get(branch.external_id, [])]:
            alias = _identity_text(raw)
            if not alias:
                continue
            alias_tokens = alias.split()
            if len(alias_tokens) == 1 and len(alias_tokens[0]) < 3 and source != alias:
                continue
            width = len(alias_tokens)
            if any(source_tokens[i : i + width] == alias_tokens for i in range(len(source_tokens) - width + 1)):
                candidates.add(branch.external_id)
    if len(candidates) == 1:
        return next(iter(candidates))
    return None


def _reference_override_map(overrides: ClinicImportOverrides) -> dict[tuple[str, str], str]:
    result: dict[tuple[str, str], str] = {}
    for item in overrides.reference_mappings:
        key = (item.reference_kind, _identity_text(item.source_value))
        existing = result.get(key)
        if existing is not None and existing != item.target_external_id:
            raise ClinicImportError(
                f"Conflicting confirmed mappings for {item.reference_kind} value {item.source_value!r}."
            )
        result[key] = item.target_external_id
    return result


def _resolve_reference(
    value: str,
    aliases: dict[str, str | None],
    *,
    path: str,
    kind: str,
    reference_overrides: dict[tuple[str, str], str] | None = None,
) -> str:
    key = _identity_text(value)
    override = (reference_overrides or {}).get((kind, key))
    if override is not None:
        override_key = _identity_text(override)
        resolved_override = aliases.get(override_key)
        if override_key not in aliases or resolved_override is None:
            # Keep the confirmed target visible to semantic validation; direct API callers
            # cannot smuggle an unknown canonical target through a reference override.
            return override
        return resolved_override
    if key not in aliases:
        # Preserve an unresolved source reference so the later semantic validation
        # can emit a stable grouped unknown_*_reference issue instead of collapsing
        # the whole preview into a generic mapping error.
        return str(value).strip()
    resolved = aliases[key]
    if resolved is None:
        raise ClinicImportError(
            f"{path}: ambiguous {kind} reference {value!r}; add a stable key or disambiguating data."
        )
    return resolved


def _auto_appointment_lifecycle(value: str) -> str | None:
    """Normalize common source lifecycle labels without asking the administrator.

    Only high-confidence semantic aliases are included. Anything clinic-specific remains
    an aggregated warning and can be mapped explicitly.
    """

    text = _identity_text(value)
    scheduled = {
        "scheduled", "pending", "confirmed", "checked in", "in progress", "booked",
        "booking", "upcoming", "waiting", "open", "new", "مجدول", "مؤكد", "محجوز",
        "قادم", "انتظار",
    }
    completed = {
        "completed", "complete", "done", "attended", "finished", "closed",
        "تم", "مكتمل", "حضر", "حضر العميل", "منتهي",
    }
    cancelled = {
        "cancelled", "canceled", "cancel", "cancellation", "إلغاء", "الغاء", "ملغي",
        "ملغى", "rescheduled", "moved", "مؤجل", "تأجيل",
    }
    no_show = {
        "no show", "noshow", "ns", "absent", "did not attend", "dna",
        "لم يحضر", "غائب",
    }
    if text in scheduled:
        return "scheduled"
    if text in completed:
        return "completed"
    if text in cancelled:
        return "cancelled"
    if text in no_show:
        return "no_show"
    return None


def _normalize_lifecycle(value: str | None) -> str:
    text = (value or "unknown").strip().lower()
    if text in {"scheduled", "pending", "confirmed", "checked_in", "in_progress"}:
        return "scheduled"
    if text == "rescheduled":
        return "cancelled"
    if text in {"completed", "cancelled", "no_show", "unknown"}:
        return text
    raise ClinicImportError(f"Unsupported canonical appointment lifecycle {value!r}.")


def _mapped_value(raw: str, mapping: dict[str, Any]) -> Any | None:
    if raw in mapping:
        return mapping[raw]
    normalized = raw.strip().lower()
    if normalized in mapping:
        return mapping[normalized]
    for key, value in mapping.items():
        if str(key).strip().lower() == normalized:
            return value
    return None


def _normalize_payment_method(raw: str | None, mapping: dict[str, PaymentMethod], default: PaymentMethod) -> PaymentMethod:
    if not raw:
        return default
    mapped = _mapped_value(raw, mapping)
    if mapped is not None:
        return mapped
    text = _identity_text(raw)
    if text in {"cash", "كاش", "نقدي", "نقداً", "نقدا"}:
        return "cash"
    if any(token in text for token in ("visa", "mastercard", "master card", "card", "pos", "فيزا", "كارت", "بطاقة")):
        return "card"
    if any(token in text for token in ("instapay", "bank transfer", "bank_transfer", "تحويل بنكي", "تحويل بنك")):
        return "bank_transfer"
    if any(token in text for token in ("wallet", "vodafone cash", "orange cash", "etisalat cash", "محفظة")):
        return "wallet"
    return "other"



def _normalize_package_status(raw: str | None, mapping: dict[str, str], default: str) -> str:
    if not raw:
        return default
    mapped = _mapped_value(raw, mapping)
    if mapped is not None:
        return str(mapped)
    text = _identity_text(raw)
    if text in {"active", "valid", "open", "فعال", "ساري"}:
        return "active"
    if text in {"expired", "ended", "منتهي", "منتهية"}:
        return "expired"
    if text in {"cancelled", "canceled", "void", "ملغي", "ملغاة"}:
        return "cancelled"
    return default


def _normalize_billing_context(
    raw: str | None,
    mapping: dict[str, BillingContext],
    default: BillingContext,
) -> BillingContext:
    if not raw:
        return default
    mapped = _mapped_value(raw, mapping)
    if mapped is not None:
        return mapped
    text = _identity_text(raw)
    package_aliases = {
        "package prepaid",
        "package_prepaid",
        "prepaid package",
        "prepaid_package",
        "package covered",
        "package_covered",
        "bundle prepaid",
        "bundle_prepaid",
        "باقة مدفوعة",
        "باكدج مدفوع",
        "مدفوع ضمن باقة",
    }
    return "package_prepaid" if text in package_aliases else default


def _optional_datetime(
    raw: str | None, *, timezone_name: str | None, path: str
) -> datetime | None:
    if raw is None:
        return None
    value = _datetime_value(raw, timezone_name=timezone_name, path=path)
    return value

def _auto_payment_status(raw: str | None) -> PaymentStatus | None:
    text = _identity_text(raw)
    if not text:
        return None
    if text in {"paid", "paid in full", "fully paid", "settled", "مدفوع", "مدفوع بالكامل"}:
        return "paid"
    if text in {"unpaid", "not paid", "outstanding", "due", "غير مدفوع", "مستحق"}:
        return "unpaid"
    if text in {"partial", "partially paid", "partial payment", "part paid", "deposit", "عربون", "مدفوع جزئيا", "مدفوع جزئي"}:
        return "partial"
    if text in {"refunded", "refund", "reversed", "مرتجع", "مسترد"}:
        return "refunded"
    return None


def _normalize_appointment_source(
    raw: str | None,
    mapping: dict[str, AppointmentSource],
    default: AppointmentSource,
) -> AppointmentSource:
    if not raw:
        return default
    mapped = _mapped_value(raw, mapping)
    if mapped is not None:
        return mapped
    text = _identity_text(raw)
    aliases: dict[str, AppointmentSource] = {
        "whatsapp": "whatsapp",
        "whats app": "whatsapp",
        "wa": "whatsapp",
        "instagram": "instagram",
        "insta": "instagram",
        "ig": "instagram",
        "website": "website",
        "web": "website",
        "phone": "phone",
        "call": "phone",
        "facebook": "facebook",
        "fb": "facebook",
        "email": "email",
        "e mail": "email",
        "walkin": "walk_in",
        "walk in": "walk_in",
        "walk-in": "walk_in",
        "frontdesk": "staff",
        "front desk": "staff",
        "reception": "staff",
        "staff": "staff",
    }
    return aliases.get(text, "other")


def _is_historical_appointment(
    *,
    appointment_date: date,
    start_at: datetime | None,
    timezone_name: str | None,
    now: datetime | None = None,
) -> bool:
    """Return True only when the appointment is clearly in the past.

    Exact timestamps are compared in UTC. Date-only records are considered historical
    only when their date is strictly before today's date in the source timezone.
    This avoids marking today's date-only appointments completed prematurely.
    """

    current = now or datetime.now(UTC)
    if current.tzinfo is None:
        current = current.replace(tzinfo=UTC)
    if start_at is not None:
        candidate = start_at if start_at.tzinfo is not None else start_at.replace(tzinfo=UTC)
        return candidate.astimezone(UTC) < current.astimezone(UTC)

    try:
        zone = ZoneInfo(timezone_name or "UTC")
    except Exception:
        zone = ZoneInfo("UTC")
    return appointment_date < current.astimezone(zone).date()


def _appointment_temporal(
    row: dict[str, Any], mapping: AppointmentSheetMapping, *, path: str
) -> tuple[date, datetime | None, datetime | None, str]:
    timezone_name = mapping.default_timezone
    if mapping.start_at:
        raw_start = _raw_optional(row, mapping.start_at)
        if raw_start is None:
            raise ClinicImportError(f"{path}: required start datetime is empty.")
        start_at = _datetime_value(
            raw_start,
            timezone_name=timezone_name,
            path=f"{path}.{mapping.start_at}",
        )
        end_at = None
        if mapping.end_at:
            raw_end = _raw_optional(row, mapping.end_at)
            if raw_end is not None:
                end_at = _datetime_value(
                    raw_end,
                    timezone_name=timezone_name,
                    path=f"{path}.{mapping.end_at}",
                )
        elif mapping.end_time:
            raw_end_time = _raw_optional(row, mapping.end_time)
            if raw_end_time is not None:
                local_date = start_at.astimezone(_timezone(timezone_name, path=path)).date()
                end_at = datetime.combine(
                    local_date,
                    _time_value(raw_end_time, f"{path}.{mapping.end_time}"),
                    tzinfo=_timezone(timezone_name, path=path),
                )
        if end_at is not None and end_at <= start_at:
            raise ClinicImportError(f"{path}: appointment end must be after start.")
        return start_at.date(), start_at, end_at, "exact"

    assert mapping.appointment_date is not None
    raw_date = _raw_optional(row, mapping.appointment_date)
    if raw_date is None:
        raise ClinicImportError(f"{path}: appointment date is empty.")
    appointment_date = _date_value(raw_date, f"{path}.{mapping.appointment_date}")
    raw_time = _raw_optional(row, mapping.appointment_time)
    if raw_time is None:
        return appointment_date, None, None, "date_only"
    tz = _timezone(timezone_name, path=path)
    start_at = datetime.combine(
        appointment_date,
        _time_value(raw_time, f"{path}.{mapping.appointment_time}"),
        tzinfo=tz,
    )
    end_at = None
    if mapping.end_time:
        raw_end = _raw_optional(row, mapping.end_time)
        if raw_end is not None:
            end_at = datetime.combine(
                appointment_date,
                _time_value(raw_end, f"{path}.{mapping.end_time}"),
                tzinfo=tz,
            )
    elif mapping.end_at:
        raw_end = _raw_optional(row, mapping.end_at)
        if raw_end is not None:
            end_at = _datetime_value(raw_end, timezone_name=timezone_name, path=f"{path}.{mapping.end_at}")
    if end_at is not None and end_at <= start_at:
        raise ClinicImportError(f"{path}: appointment end must be after start.")
    return appointment_date, start_at, end_at, "exact"


def build_import_preview(
    workbook: TabularWorkbook,
    mapping: ClinicImportMapping,
    *,
    confirm_no_existing_appointments: bool = False,
    overrides: ClinicImportOverrides | None = None,
) -> ClinicImportPreviewResponse:
    overrides = overrides or ClinicImportOverrides()
    mapping = mapping.model_copy(deep=True)
    if overrides.appointment_status_map:
        for appointment_mapping in _appointment_mappings(mapping):
            appointment_mapping.status_map = {
                **dict(appointment_mapping.status_map),
                **dict(overrides.appointment_status_map),
            }
    for appointment_mapping in _appointment_mappings(mapping):
        if overrides.appointment_payment_status_map:
            appointment_mapping.payment_status_map = {
                **dict(appointment_mapping.payment_status_map),
                **dict(overrides.appointment_payment_status_map),
            }
        if overrides.appointment_payment_method_map:
            appointment_mapping.payment_method_map = {
                **dict(appointment_mapping.payment_method_map),
                **dict(overrides.appointment_payment_method_map),
            }
        if overrides.appointment_source_map:
            appointment_mapping.source_map = {
                **dict(appointment_mapping.source_map),
                **dict(overrides.appointment_source_map),
            }
    if overrides.confirm_no_existing_appointments is not None:
        confirm_no_existing_appointments = overrides.confirm_no_existing_appointments

    issues: list[ClinicImportIssue] = []
    services: list[NormalizedServiceImport] = []
    branches: list[NormalizedBranchImport] = []
    doctors: list[NormalizedDoctorImport] = []
    branch_hours: list[NormalizedWorkingHourImport] = []
    doctor_hours: list[NormalizedWorkingHourImport] = []
    appointments: list[NormalizedAppointmentImport] = []
    appointment_origin_sheets: list[str] = []
    appointment_origin_paths: list[str] = []
    payments: list[NormalizedPaymentImport] = []
    payment_allocations: list[NormalizedPaymentAllocationImport] = []
    packages: list[NormalizedPackageImport] = []
    package_usages: list[NormalizedPackageUsageImport] = []
    unlinked_patient_counts: dict[str, int] = {}
    service_extra_aliases: dict[str, list[str]] = {}
    branch_extra_aliases: dict[str, list[str]] = {}
    doctor_extra_aliases: dict[str, list[str]] = {}
    reference_overrides = _reference_override_map(overrides)
    unresolved_reference_groups: dict[tuple[str, str], dict[str, Any]] = {}
    unmapped_status_groups: dict[str, dict[str, Any]] = {}
    missing_exact_time_groups: dict[str, dict[str, Any]] = {}
    row_data_error_groups: dict[tuple[str, str], dict[str, Any]] = {}
    raw_sheet_summary = {name: len(rows) for name, rows in workbook.sheets.items()}
    transformed_sheet_summary: dict[str, int] = {}

    def record_unresolved_reference(*, kind: str, raw_value: str, path: str, sheet: str) -> None:
        key = (kind, _identity_text(raw_value))
        group = unresolved_reference_groups.setdefault(
            key,
            {
                "raw_value": str(raw_value).strip(),
                "count": 0,
                "sheets": set(),
                "examples": [],
            },
        )
        group["count"] += 1
        group["sheets"].add(sheet)
        if len(group["examples"]) < 5:
            group["examples"].append(path)

    def record_unmapped_status(*, raw_value: str, path: str, sheet: str) -> None:
        key = _identity_text(raw_value)
        group = unmapped_status_groups.setdefault(
            key,
            {
                "raw_value": str(raw_value).strip(),
                "count": 0,
                "sheets": set(),
                "examples": [],
            },
        )
        group["count"] += 1
        group["sheets"].add(sheet)
        if len(group["examples"]) < 5:
            group["examples"].append(path)

    def record_missing_exact_time(*, path: str, sheet: str) -> None:
        group = missing_exact_time_groups.setdefault(sheet, {"count": 0, "examples": []})
        group["count"] += 1
        if len(group["examples"]) < 5:
            group["examples"].append(path)

    def record_row_data_error(*, exc: ClinicImportError, path: str, sheet: str, row_kind: str) -> None:
        message = str(exc)
        lowered = message.casefold()
        # Missing mapped columns are source/header configuration errors, not bad row data.
        # Let the outer mapping guard surface them once at source level.
        if "mapped column" in lowered and "does not exist" in lowered:
            raise exc
        if "default_timezone is required for local times" in lowered:
            code = f"{row_kind}_timezone_missing"
            summary = f"One or more {row_kind} rows use local times without a timezone default."
        elif "unsupported datetime value" in lowered or "unsupported date value" in lowered:
            code = f"{row_kind}_datetime_invalid"
            summary = f"One or more {row_kind} rows contain an invalid date/datetime value."
        elif "unsupported time value" in lowered or "appointment end must be after start" in lowered:
            code = f"{row_kind}_time_invalid"
            summary = f"One or more {row_kind} rows contain an invalid time range."
        elif "invalid price" in lowered or "amount" in lowered or "price cannot be negative" in lowered:
            code = "invalid_amount"
            summary = "One or more rows contain an invalid financial amount."
        elif "required value" in lowered and "is empty" in lowered:
            code = "required_value_missing"
            summary = "One or more rows are missing a required value."
        elif "phone" in lowered:
            code = "invalid_phone"
            summary = "One or more rows contain an invalid patient phone value."
        else:
            code = f"{row_kind}_row_invalid"
            summary = f"One or more {row_kind} rows contain invalid source data and were skipped."
        key = (code, sheet)
        group = row_data_error_groups.setdefault(
            key,
            {
                "count": 0,
                "examples": [],
                "messages": [],
                "summary": summary,
            },
        )
        group["count"] += 1
        if len(group["examples"]) < 5:
            group["examples"].append(path)
        if len(group["messages"]) < 3 and message not in group["messages"]:
            group["messages"].append(message)

    try:
        if mapping.transformations:
            transformed_sheets, transformed_sheet_summary = apply_structural_transforms(
                workbook.sheets, mapping.transformations
            )
            workbook = TabularWorkbook(
                sheets=transformed_sheets,
                document_names=workbook.document_names,
            )

        for index, row in enumerate(_rows(workbook, mapping.services.sheet), start=2):
            path = f"{mapping.services.sheet}!row{index}"
            try:
                name = _required(row, mapping.services.name, path)
                external_id, generated = _entity_source_key(
                    row,
                    external_id_column=mapping.services.external_id,
                    fallback_parts=(name,),
                    kind="service",
                    path=path,
                )
                service_extra_aliases[external_id] = _catalog_name_aliases(
                    row, mapping.services.name
                )
                if mapping.services.duration_minutes:
                    raw_duration = _optional(row, mapping.services.duration_minutes)
                    duration = (
                        _positive_int(
                            raw_duration,
                            f"{path}.{mapping.services.duration_minutes}",
                        )
                        if raw_duration is not None
                        else mapping.services.default_duration_minutes
                    )
                else:
                    duration = mapping.services.default_duration_minutes
                price_value: str | float = mapping.services.default_price
                if mapping.services.price:
                    raw_price = _optional(row, mapping.services.price)
                    if raw_price is not None:
                        price_value = raw_price
                services.append(
                    NormalizedServiceImport(
                        external_id=external_id,
                        external_id_generated=generated,
                        name=name,
                        duration_minutes=duration,
                        price_minor=_price_minor(price_value, f"{path}.price"),
                        currency="EGP",
                        category=_optional(row, mapping.services.category),
                    )
                )

            except ClinicImportError as exc:
                record_row_data_error(
                    exc=exc,
                    path=path,
                    sheet=mapping.services.sheet,
                    row_kind="service",
                )
                continue
        service_aliases = _alias_map(services, extra_aliases=service_extra_aliases)

        for index, row in enumerate(_rows(workbook, mapping.branches.sheet), start=2):
            path = f"{mapping.branches.sheet}!row{index}"
            try:
                name = _required(row, mapping.branches.name, path)
                city = _optional(row, mapping.branches.city)
                external_id, generated = _entity_source_key(
                    row,
                    external_id_column=mapping.branches.external_id,
                    fallback_parts=(name, city or ""),
                    kind="branch",
                    path=path,
                )
                branch_extra_aliases[external_id] = _catalog_name_aliases(
                    row, mapping.branches.name
                )
                timezone_name = _optional(row, mapping.branches.timezone) or mapping.branches.default_timezone
                if timezone_name:
                    _timezone(timezone_name, path=path)
                branches.append(
                    NormalizedBranchImport(
                        external_id=external_id,
                        external_id_generated=generated,
                        name=name,
                        city=city,
                        address=_optional(row, mapping.branches.address),
                        timezone=timezone_name,
                    )
                )

            except ClinicImportError as exc:
                record_row_data_error(
                    exc=exc,
                    path=path,
                    sheet=mapping.branches.sheet,
                    row_kind="branch",
                )
                continue
        branch_aliases = _alias_map(branches, extra_aliases=branch_extra_aliases)

        # A dedicated doctor master sheet is optional. When it is absent, use one
        # appointment source that explicitly records a doctor as a conservative discovery
        # source. This does not fuzzy-merge doctors across independent files: it only builds
        # a catalog from repeated exact source labels in the selected sheet. The AI/admin may
        # also map ``doctors`` to any other relationship/reference sheet that contains doctor
        # identity plus service/branch evidence; it does not need to be a doctor-only table.
        doctor_mapping = mapping.doctors
        doctor_mapping_discovered_from_appointments = False
        if doctor_mapping is None:
            candidates: list[tuple[int, int, int, AppointmentSheetMapping]] = []
            for order, appointment_mapping in enumerate(_appointment_mappings(mapping)):
                if not appointment_mapping.doctor_external_id:
                    continue
                rows = _rows(workbook, appointment_mapping.sheet)
                nonempty = sum(
                    1 for row in rows if _optional(row, appointment_mapping.doctor_external_id)
                )
                if nonempty:
                    header = _identity_text(appointment_mapping.doctor_external_id)
                    header_tokens = set(header.split())
                    stable_identity_hint = int(
                        bool(header_tokens.intersection({"id", "code", "ref", "key", "uuid"}))
                    )
                    candidates.append(
                        (stable_identity_hint, nonempty, -order, appointment_mapping)
                    )
            if candidates:
                _stable_hint, _count, _order, appointment_source = max(
                    candidates, key=lambda item: (item[0], item[1], item[2])
                )
                doctor_mapping = DoctorSheetMapping(
                    sheet=appointment_source.sheet,
                    name=appointment_source.doctor_external_id or "",
                    service_external_ids=appointment_source.service_external_id,
                    branch_external_ids=appointment_source.branch_external_id,
                )
                doctor_mapping_discovered_from_appointments = True

        doctor_source_is_appointment = bool(
            doctor_mapping is not None
            and doctor_mapping.sheet in {item.sheet for item in _appointment_mappings(mapping)}
        )

        # First pass doctors so exact source labels in schedules/appointments can resolve
        # without requiring a dedicated source key. Repeated rows in a relationship source
        # are aggregated by stable source ID or by the exact selected source label.
        doctor_rows: list[tuple[int, dict[str, Any], str]] = []
        doctor_by_key: dict[str, NormalizedDoctorImport] = {}
        skipped_doctor_identity_rows = 0
        skipped_doctor_examples: list[str] = []
        deferred_doctor_alias_rows = 0
        deferred_doctor_alias_examples: list[str] = []
        doctor_name_header_tokens = (
            set(_identity_text(doctor_mapping.name).split()) if doctor_mapping is not None else set()
        )
        doctor_name_column_is_stable_key = bool(
            doctor_name_header_tokens.intersection({"id", "code", "ref", "key", "uuid"})
        )
        if doctor_mapping is not None:
            for index, row in enumerate(_rows(workbook, doctor_mapping.sheet), start=2):
                path = f"{doctor_mapping.sheet}!row{index}"
                try:
                    name = _optional(row, doctor_mapping.name)
                    if not name:
                        skipped_doctor_identity_rows += 1
                        if len(skipped_doctor_examples) < 5:
                            skipped_doctor_examples.append(path)
                        continue
                    specialization = _optional(row, doctor_mapping.specialization) or ""
                    raw_external_id = _optional(row, doctor_mapping.external_id)
                    if (
                        doctor_source_is_appointment
                        and raw_external_id is None
                        and doctor_name_column_is_stable_key
                    ):
                        source_label = unicodedata.normalize("NFKC", str(name)).strip()
                        key_like = bool(
                            source_label
                            and not any(char.isspace() for char in source_label)
                            and re.fullmatch(r"[\w.:-]+", source_label, flags=re.UNICODE)
                        )
                        if not key_like:
                            deferred_doctor_alias_rows += 1
                            if len(deferred_doctor_alias_examples) < 5:
                                deferred_doctor_alias_examples.append(path)
                            continue
                    if raw_external_id:
                        external_id = raw_external_id
                        generated = False
                    else:
                        # This is exact selected-source identity, not fuzzy name matching. Different
                        # spellings remain distinct candidates until explicit/evidence-backed mapping.
                        exact_label = " ".join(unicodedata.normalize("NFKC", str(name)).casefold().split())
                        exact_specialization = " ".join(
                            unicodedata.normalize("NFKC", str(specialization)).casefold().split()
                        )
                        material = "\x1f".join(part for part in (exact_label, exact_specialization) if part)
                        digest = hashlib.sha256(material.encode("utf-8")).hexdigest()[:24]
                        external_id = f"auto:doctor:{digest}"
                        generated = True

                    doctor_name_aliases = (
                        [str(name).strip()]
                        if doctor_source_is_appointment
                        else _catalog_name_aliases(row, doctor_mapping.name)
                    )
                    aliases = doctor_extra_aliases.setdefault(external_id, [])
                    for alias in [
                        *doctor_name_aliases,
                        *(normalize_doctor_display_name(value) for value in doctor_name_aliases),
                        str(name).strip(),
                        normalize_doctor_display_name(name),
                    ]:
                        if alias and alias not in aliases:
                            aliases.append(alias)

                    doctor = doctor_by_key.get(external_id)
                    if doctor is None:
                        doctor = NormalizedDoctorImport(
                            external_id=external_id,
                            external_id_generated=generated,
                            name=str(name).strip(),
                            specialization=specialization or None,
                            service_external_ids=[],
                            branch_external_ids=[],
                        )
                        doctor_by_key[external_id] = doctor
                        doctors.append(doctor)
                    elif (doctor.name != str(name).strip() or (doctor.specialization or "") != specialization):
                        raise ClinicImportError(
                            f"{path}: doctor source identity {external_id!r} refers to conflicting doctor labels; "
                            "add a stable provider key or split the source identities before import."
                        )
                    doctor_rows.append((index, row, external_id))

                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=doctor_mapping.sheet,
                        row_kind="doctor",
                    )
                    continue
            if deferred_doctor_alias_rows:
                issues.append(
                    ClinicImportIssue(
                        severity="warning",
                        code="doctor_source_aliases_deferred_for_resolution",
                        path=f"doctor_source:{doctor_mapping.sheet}",
                        message=(
                            f"{deferred_doctor_alias_rows} row(s) in {doctor_mapping.sheet!r} contain "
                            "human-readable values inside a doctor ID/code column. Tia did not promote those "
                            "values into new canonical doctors; they remain appointment aliases that must be "
                            "matched to the discovered doctor IDs using evidence or administrator review."
                        ),
                        occurrence_count=deferred_doctor_alias_rows,
                        source_sheets=[doctor_mapping.sheet],
                        example_paths=deferred_doctor_alias_examples,
                    )
                )
            if skipped_doctor_identity_rows:
                issues.append(
                    ClinicImportIssue(
                        severity="warning",
                        code="doctor_source_identity_missing",
                        path=f"doctor_source:{doctor_mapping.sheet}",
                        message=(
                            f"{skipped_doctor_identity_rows} row(s) in {doctor_mapping.sheet!r} do not "
                            "record a doctor identity. Those rows were ignored for doctor discovery rather "
                            "than creating a fake provider."
                        ),
                        occurrence_count=skipped_doctor_identity_rows,
                        source_sheets=[doctor_mapping.sheet],
                        example_paths=skipped_doctor_examples,
                    )
                )
            if doctor_mapping_discovered_from_appointments and doctors:
                issues.append(
                    ClinicImportIssue(
                        severity="warning",
                        code="doctor_catalog_discovered_from_appointment_source",
                        path=f"doctor_source:{doctor_mapping.sheet}",
                        message=(
                            f"No separate doctor catalog was mapped. Tia discovered {len(doctors)} doctor "
                            f"candidate(s) from exact doctor labels in appointment source {doctor_mapping.sheet!r} "
                            "and will use appointment history as service/branch evidence. Review the catalog "
                            "before import; no fuzzy-name merge was performed."
                        ),
                        occurrence_count=len(doctors),
                        source_sheets=[doctor_mapping.sheet],
                    )
                )

        doctor_aliases = _alias_map(doctors, extra_aliases=doctor_extra_aliases)
        for index, row, doctor_key in doctor_rows:
            assert doctor_mapping is not None
            path = f"{doctor_mapping.sheet}!row{index}"
            doctor = doctor_by_key[doctor_key]
            service_refs = (
                _split_ids(
                    _required(row, doctor_mapping.service_external_ids, path),
                    doctor_mapping.delimiter,
                )
                if doctor_mapping.service_external_ids
                and _optional(row, doctor_mapping.service_external_ids)
                else []
            )
            branch_refs = (
                _split_ids(
                    _required(row, doctor_mapping.branch_external_ids, path),
                    doctor_mapping.delimiter,
                )
                if doctor_mapping.branch_external_ids
                and _optional(row, doctor_mapping.branch_external_ids)
                else []
            )
            resolved_services = [
                _resolve_reference(
                    value, service_aliases, path=path, kind="service",
                    reference_overrides=reference_overrides,
                )
                for value in service_refs
            ]
            resolved_branches = [
                _resolve_reference(
                    value, branch_aliases, path=path, kind="branch",
                    reference_overrides=reference_overrides,
                )
                for value in branch_refs
            ]
            doctor.service_external_ids = list(
                dict.fromkeys([*doctor.service_external_ids, *resolved_services])
            )
            doctor.branch_external_ids = list(
                dict.fromkeys([*doctor.branch_external_ids, *resolved_branches])
            )

        if mapping.branch_hours is not None:
            for index, row in enumerate(_rows(workbook, mapping.branch_hours.sheet), start=2):
                path = f"{mapping.branch_hours.sheet}!row{index}"
                try:
                    branch_ref = _required(row, mapping.branch_hours.branch_external_id, path)
                    branch_hours.append(
                        NormalizedWorkingHourImport(
                            owner_external_id=_resolve_reference(
                                branch_ref, branch_aliases, path=path, kind="branch",
                                reference_overrides=reference_overrides,
                            ),
                            weekday=_weekday(_required(row, mapping.branch_hours.weekday, path), f"{path}.weekday"),
                            start_time=_time_text(_required(row, mapping.branch_hours.start_time, path), f"{path}.start"),
                            end_time=_time_text(_required(row, mapping.branch_hours.end_time, path), f"{path}.end"),
                        )
                    )

                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=mapping.branch_hours.sheet,
                        row_kind="working_hour",
                    )
                    continue
        if mapping.doctor_hours is not None:
            for index, row in enumerate(_rows(workbook, mapping.doctor_hours.sheet), start=2):
                path = f"{mapping.doctor_hours.sheet}!row{index}"
                try:
                    doctor_ref = _required(row, mapping.doctor_hours.doctor_external_id, path)
                    branch_ref = _required(row, mapping.doctor_hours.branch_external_id, path)
                    doctor_hours.append(
                        NormalizedWorkingHourImport(
                            owner_external_id=_resolve_reference(
                                doctor_ref, doctor_aliases, path=path, kind="doctor",
                                reference_overrides=reference_overrides,
                            ),
                            branch_external_id=_resolve_reference(
                                branch_ref, branch_aliases, path=path, kind="branch",
                                reference_overrides=reference_overrides,
                            ),
                            weekday=_weekday(_required(row, mapping.doctor_hours.weekday, path), f"{path}.weekday"),
                            start_time=_time_text(_required(row, mapping.doctor_hours.start_time, path), f"{path}.start"),
                            end_time=_time_text(_required(row, mapping.doctor_hours.end_time, path), f"{path}.end"),
                        )
                    )

                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=mapping.doctor_hours.sheet,
                        row_kind="working_hour",
                    )
                    continue
        service_by_key = {item.external_id: item for item in services}
        branch_by_key = {item.external_id: item for item in branches}
        doctor_by_key = {item.external_id: item for item in doctors}

        def ensure_unknown_branch() -> None:
            if UNKNOWN_BRANCH_EXTERNAL_ID in branch_by_key:
                return
            placeholder = NormalizedBranchImport(
                external_id=UNKNOWN_BRANCH_EXTERNAL_ID,
                name="Unknown / Not recorded branch",
                is_placeholder=True,
            )
            branches.append(placeholder)
            branch_by_key[placeholder.external_id] = placeholder

        def ensure_unknown_doctor() -> None:
            if UNKNOWN_DOCTOR_EXTERNAL_ID in doctor_by_key:
                return
            placeholder = NormalizedDoctorImport(
                external_id=UNKNOWN_DOCTOR_EXTERNAL_ID,
                name="Unknown / Unassigned doctor",
                is_placeholder=True,
                service_external_ids=[],
                branch_external_ids=[],
            )
            doctors.append(placeholder)
            doctor_by_key[placeholder.external_id] = placeholder

        for appointment_mapping in _appointment_mappings(mapping):
            appointment_rows = _rows(workbook, appointment_mapping.sheet)
            implicit_branch_id: str | None = None
            implicit_branch_known = True
            if not appointment_mapping.branch_external_id:
                source_default = str(appointment_mapping.default_branch_external_id or "").strip()
                inferred_from_name = _infer_branch_from_source_name(
                    sheet_name=appointment_mapping.sheet,
                    branches=[item for item in branches if not item.is_placeholder],
                    extra_aliases=branch_extra_aliases,
                )
                if source_default:
                    implicit_branch_id = _resolve_reference(
                        source_default,
                        branch_aliases,
                        path=f"appointment_source:{appointment_mapping.sheet}/branch",
                        kind="branch",
                        reference_overrides=reference_overrides,
                    )
                    if appointment_rows:
                        issues.append(
                            ClinicImportIssue(
                                severity="warning",
                                code="appointment_branch_defaulted_source_context",
                                path=f"appointment_source:{appointment_mapping.sheet}/branch",
                                message=(
                                    f"Appointment source {appointment_mapping.sheet!r} has no branch column. "
                                    f"A source-level branch reference {source_default!r} is being used for all "
                                    f"{len(appointment_rows)} row(s); review it before final import."
                                ),
                                occurrence_count=len(appointment_rows),
                                source_value=source_default,
                                reference_kind="branch",
                                source_sheets=[appointment_mapping.sheet],
                            )
                        )
                elif inferred_from_name is not None:
                    implicit_branch_id = inferred_from_name
                    if appointment_rows:
                        branch_name = branch_by_key[inferred_from_name].name
                        issues.append(
                            ClinicImportIssue(
                                severity="warning",
                                code="appointment_branch_inferred_from_source_name",
                                path=f"appointment_source:{appointment_mapping.sheet}/branch",
                                message=(
                                    f"Appointment source {appointment_mapping.sheet!r} has no branch column. "
                                    f"Its source name exactly identifies branch {branch_name!r}, so all "
                                    f"{len(appointment_rows)} row(s) use that branch; review before import."
                                ),
                                occurrence_count=len(appointment_rows),
                                source_sheets=[appointment_mapping.sheet],
                            )
                        )
                elif len([item for item in branches if not item.is_placeholder]) == 1:
                    only_branch = next(item for item in branches if not item.is_placeholder)
                    implicit_branch_id = only_branch.external_id
                    if appointment_rows:
                        issues.append(
                            ClinicImportIssue(
                                severity="warning",
                                code="appointment_branch_defaulted_single_branch",
                                path=f"appointment_source:{appointment_mapping.sheet}/branch",
                                message=(
                                    f"Appointment source {appointment_mapping.sheet!r} has no branch column. "
                                    f"All {len(appointment_rows)} row(s) were deterministically assigned to "
                                    f"the only imported branch {only_branch.name!r}; review the normalized "
                                    "appointments before final import."
                                ),
                                occurrence_count=len(appointment_rows),
                                source_sheets=[appointment_mapping.sheet],
                            )
                        )
                else:
                    ensure_unknown_branch()
                    implicit_branch_id = UNKNOWN_BRANCH_EXTERNAL_ID
                    implicit_branch_known = False
                    if appointment_rows:
                        issues.append(
                            ClinicImportIssue(
                                severity="warning",
                                code="appointment_branch_unknown_unrecorded",
                                path=f"appointment_source:{appointment_mapping.sheet}/branch",
                                message=(
                                    f"Appointment source {appointment_mapping.sheet!r} has no branch data and "
                                    "its source name does not uniquely identify a catalog branch. Tia will "
                                    "preserve these appointments under an inactive Unknown branch instead of guessing."
                                ),
                                occurrence_count=len(appointment_rows),
                                source_sheets=[appointment_mapping.sheet],
                            )
                        )
            unknown_doctor_count = 0
            unknown_doctor_examples: list[str] = []
            historical_completed_inferred_count = 0
            historical_completed_examples: list[str] = []
            for index, row in enumerate(appointment_rows, start=2):
                path = f"{appointment_mapping.sheet}!row{index}"
                try:
                    patient_name = _required(row, appointment_mapping.patient_name, path)
                    patient_phone = _optional(row, appointment_mapping.patient_phone)
                    raw_patient_external_id = _optional(row, appointment_mapping.patient_external_id)
                    if raw_patient_external_id:
                        patient_external_id = raw_patient_external_id
                        patient_generated = False
                    elif patient_phone:
                        # When the source has no stable patient ID, the normalized phone is the
                        # patient identity. Names may vary across exports (Arabic/English,
                        # abbreviations, typos) and must not split one phone into multiple people.
                        try:
                            _display_phone, identity_phone = normalize_patient_identity_phone(patient_phone)
                        except ValueError:
                            # Preserve the row for preview/apply diagnostics; apply will reject the
                            # invalid phone instead of silently merging it with another patient.
                            identity_phone = patient_phone.strip()
                        patient_external_id = _generated_source_key(
                            "patient_phone",
                            identity_phone,
                        )
                        patient_generated = True
                    else:
                        # Name-only appointment exports cannot prove that two equal names refer
                        # to the same person. Preserve every appointment without making a
                        # name-only merge by generating a row-scoped patient source identity.
                        appointment_identity_hint = _optional(row, appointment_mapping.external_id) or path
                        patient_external_id = _generated_source_key(
                            "patient_unlinked",
                            appointment_mapping.sheet,
                            appointment_identity_hint,
                        )
                        patient_generated = True
                        unlinked_patient_counts[appointment_mapping.sheet] = (
                            unlinked_patient_counts.get(appointment_mapping.sheet, 0) + 1
                        )
                    service_ref = _required(row, appointment_mapping.service_external_id, path)
                    service_id = _resolve_reference(
                        service_ref, service_aliases, path=path, kind="service",
                        reference_overrides=reference_overrides,
                    )
                    raw_doctor_ref = _optional(row, appointment_mapping.doctor_external_id)
                    doctor_assignment_known = raw_doctor_ref is not None
                    if doctor_assignment_known:
                        doctor_ref = raw_doctor_ref
                        doctor_id = _resolve_reference(
                            doctor_ref, doctor_aliases, path=path, kind="doctor",
                            reference_overrides=reference_overrides,
                        )
                    else:
                        ensure_unknown_doctor()
                        doctor_ref = UNKNOWN_DOCTOR_EXTERNAL_ID
                        doctor_id = UNKNOWN_DOCTOR_EXTERNAL_ID
                        unknown_doctor_count += 1
                        if len(unknown_doctor_examples) < 5:
                            unknown_doctor_examples.append(path)
                    if appointment_mapping.branch_external_id:
                        branch_ref = _required(row, appointment_mapping.branch_external_id, path)
                        branch_id = _resolve_reference(
                            branch_ref, branch_aliases, path=path, kind="branch",
                            reference_overrides=reference_overrides,
                        )
                    else:
                        assert implicit_branch_id is not None
                        branch_ref = implicit_branch_id
                        branch_id = implicit_branch_id
                    branch_assignment_known = (
                        True if appointment_mapping.branch_external_id else implicit_branch_known
                    )
                    if service_id not in service_by_key:
                        record_unresolved_reference(
                            kind="service", raw_value=service_ref, path=path, sheet=appointment_mapping.sheet
                        )
                    if branch_id not in branch_by_key:
                        record_unresolved_reference(
                            kind="branch", raw_value=branch_ref, path=path, sheet=appointment_mapping.sheet
                        )
                    if doctor_assignment_known and doctor_id not in doctor_by_key:
                        record_unresolved_reference(
                            kind="doctor", raw_value=doctor_ref, path=path, sheet=appointment_mapping.sheet
                        )
                    appointment_date, start_at, end_at, time_precision = _appointment_temporal(
                        row, appointment_mapping, path=path
                    )
                    if time_precision != "exact" or start_at is None:
                        record_missing_exact_time(path=path, sheet=appointment_mapping.sheet)

                    raw_status = _optional(row, appointment_mapping.status)
                    status_is_unresolved = False
                    if raw_status is not None:
                        mapped = _mapped_value(raw_status, appointment_mapping.status_map)
                        if mapped is not None:
                            lifecycle = _normalize_lifecycle(str(mapped))
                        else:
                            automatic = _auto_appointment_lifecycle(raw_status)
                            if automatic is not None:
                                lifecycle = automatic
                            else:
                                lifecycle = "unknown"
                                status_is_unresolved = True
                                record_unmapped_status(
                                    raw_value=raw_status,
                                    path=path,
                                    sheet=appointment_mapping.sheet,
                                )
                    else:
                        lifecycle = _normalize_lifecycle(appointment_mapping.default_status)

                    historical = _is_historical_appointment(
                        appointment_date=appointment_date,
                        start_at=start_at,
                        timezone_name=appointment_mapping.default_timezone,
                    )
                    # Historical clinic data frequently carries stale active lifecycle values
                    # (pending/confirmed/booked) or no lifecycle at all. Once the appointment is
                    # clearly in the past, treat it as completed unless the source provides
                    # contrary terminal evidence (cancelled/no_show) or an explicit but unresolved
                    # source status that still needs administrator mapping.
                    if (
                        historical
                        and not status_is_unresolved
                        and lifecycle in {"scheduled", "unknown"}
                    ):
                        lifecycle = "completed"
                        historical_completed_inferred_count += 1
                        if len(historical_completed_examples) < 5:
                            historical_completed_examples.append(path)

                    raw_amount = _optional(row, appointment_mapping.amount_paid)
                    amount_paid_minor = (
                        _price_minor(raw_amount, f"{path}.amount_paid") if raw_amount is not None else None
                    )
                    raw_payment_status = _optional(row, appointment_mapping.payment_status)
                    if raw_payment_status:
                        mapped_payment = _mapped_value(
                            raw_payment_status, appointment_mapping.payment_status_map
                        )
                        payment_status: PaymentStatus = (
                            mapped_payment
                            if mapped_payment is not None
                            else (_auto_payment_status(raw_payment_status) or "unknown")
                        )
                    else:
                        payment_status = appointment_mapping.default_payment_status

                    service_item = service_by_key.get(service_id)
                    # Amount-based inference is only safe when the source did not provide a
                    # payment-status fact. An explicit but unknown status must not be silently
                    # overwritten by price arithmetic.
                    if raw_payment_status is None and payment_status == "unknown" and amount_paid_minor is not None:
                        if amount_paid_minor <= 0:
                            payment_status = "unpaid"
                        elif service_item is not None and service_item.price_minor > 0:
                            payment_status = (
                                "paid" if amount_paid_minor >= service_item.price_minor else "partial"
                            )
                        # If the source service reference is unresolved, do not guess whether a
                        # positive payment is partial or paid. Semantic validation below will
                        # surface the unknown service reference while preserving the row.

                    raw_method = _optional(row, appointment_mapping.payment_method)
                    payment_method = _normalize_payment_method(
                        raw_method,
                        appointment_mapping.payment_method_map,
                        appointment_mapping.default_payment_method,
                    )
                    raw_payment_context = _optional(row, appointment_mapping.payment_context)
                    billing_context = _normalize_billing_context(
                        raw_payment_context,
                        appointment_mapping.payment_context_map,
                        appointment_mapping.default_payment_context,
                    )
                    package_external_id = _optional(row, appointment_mapping.package_external_id)
                    payment_external_reference = _optional(row, appointment_mapping.payment_reference)
                    raw_refund_amount = _optional(row, appointment_mapping.refund_amount)
                    refund_amount_minor = (
                        _price_minor(raw_refund_amount, f"{path}.refund_amount")
                        if raw_refund_amount is not None
                        else None
                    )
                    refund_reason = _optional(row, appointment_mapping.refund_reason)
                    raw_refunded_at = _optional(row, appointment_mapping.refunded_at)
                    refunded_at = _optional_datetime(
                        raw_refunded_at,
                        timezone_name=appointment_mapping.default_timezone,
                        path=f"{path}.refunded_at",
                    )
                    if billing_context == "package_prepaid":
                        # A package-covered session is settled without a new appointment-level
                        # financial transaction. Keep amount absent so Apply cannot double-count
                        # package revenue while still exposing the session as covered.
                        payment_status = "paid"
                        amount_paid_minor = None
                        payment_method = "unknown"
                        payment_external_reference = None
                        refund_amount_minor = None
                        refund_reason = None
                        refunded_at = None

                    if refund_amount_minor is not None and refund_amount_minor <= 0:
                        refund_amount_minor = None
                    if refund_amount_minor is not None and amount_paid_minor is None:
                        raise ClinicImportError(
                            f"{path}: refund amount is present but the appointment has no direct payment amount."
                        )
                    if (
                        refund_amount_minor is not None
                        and amount_paid_minor is not None
                        and refund_amount_minor > amount_paid_minor
                    ):
                        raise ClinicImportError(
                            f"{path}: refund amount cannot exceed the direct payment amount."
                        )

                    raw_source = _optional(row, appointment_mapping.source)
                    appointment_source = _normalize_appointment_source(
                        raw_source,
                        appointment_mapping.source_map,
                        appointment_mapping.default_source,
                    )

                    appointment_external = _optional(row, appointment_mapping.external_id)
                    if appointment_external:
                        appointment_id = appointment_external
                        appointment_generated = False
                    else:
                        temporal_key = start_at.isoformat() if start_at else appointment_date.isoformat()
                        appointment_id = _generated_source_key(
                            "appointment",
                            patient_external_id,
                            service_id,
                            branch_id,
                            doctor_id,
                            temporal_key,
                        )
                        appointment_generated = True

                    appointments.append(
                        NormalizedAppointmentImport(
                            external_id=appointment_id,
                            external_id_generated=appointment_generated,
                            patient_external_id=patient_external_id,
                            patient_external_id_generated=patient_generated,
                            patient_name=patient_name,
                            patient_phone=patient_phone,
                            service_external_id=service_id,
                            branch_external_id=branch_id,
                            doctor_external_id=doctor_id,
                            branch_assignment_known=branch_assignment_known,
                            doctor_assignment_known=doctor_assignment_known,
                            appointment_date=appointment_date,
                            start_at=start_at,
                            end_at=end_at,
                            time_precision=time_precision,
                            status=lifecycle,
                            payment_status=payment_status,
                            amount_paid_minor=amount_paid_minor,
                            payment_method=payment_method,
                            billing_context=billing_context,
                            package_external_id=package_external_id,
                            payment_external_reference=payment_external_reference,
                            refund_amount_minor=refund_amount_minor,
                            refund_reason=refund_reason,
                            refunded_at=refunded_at,
                            source=appointment_source,
                        )
                    )
                    appointment_origin_sheets.append(appointment_mapping.sheet)
                    appointment_origin_paths.append(path)
                    # Appointment history is factual evidence that a recorded doctor performed
                    # this service and worked at this recorded branch. Use that direct evidence to
                    # enrich a discovered/minimal doctor catalog deterministically. Unknown/unresolved
                    # references never create assignments.
                    if doctor_assignment_known:
                        doctor_item = doctor_by_key.get(doctor_id)
                        if doctor_item is not None and not doctor_item.is_placeholder:
                            if service_id in service_by_key and service_id not in doctor_item.service_external_ids:
                                doctor_item.service_external_ids.append(service_id)
                            if (
                                branch_assignment_known
                                and branch_id in branch_by_key
                                and not branch_by_key[branch_id].is_placeholder
                                and branch_id not in doctor_item.branch_external_ids
                            ):
                                doctor_item.branch_external_ids.append(branch_id)
                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=appointment_mapping.sheet,
                        row_kind="appointment",
                    )
                    continue
            if unknown_doctor_count:
                issues.append(
                    ClinicImportIssue(
                        severity="warning",
                        code="appointment_doctor_unknown_unrecorded",
                        path=f"appointment_source:{appointment_mapping.sheet}/doctor",
                        message=(
                            f"{unknown_doctor_count} appointment row(s) in {appointment_mapping.sheet!r} "
                            "do not record a doctor. Tia will preserve them under an inactive Unknown / "
                            "Unassigned doctor and will not treat that placeholder as a booking resource."
                        ),
                        occurrence_count=unknown_doctor_count,
                        source_sheets=[appointment_mapping.sheet],
                        example_paths=unknown_doctor_examples,
                    )
                )
            if historical_completed_inferred_count:
                issues.append(
                    ClinicImportIssue(
                        severity="warning",
                        code="appointment_status_inferred_historical_completed",
                        path=f"appointment_source:{appointment_mapping.sheet}/status",
                        message=(
                            f"{historical_completed_inferred_count} past appointment row(s) in "
                            f"{appointment_mapping.sheet!r} have no contrary terminal status evidence. "
                            "Tia will treat them as completed even when the old source left status blank "
                            "or still marked them scheduled/confirmed; cancelled/no-show evidence is preserved. "
                            "Review this historical default before final import."
                        ),
                        occurrence_count=historical_completed_inferred_count,
                        source_sheets=[appointment_mapping.sheet],
                        example_paths=historical_completed_examples,
                    )
                )

        # Standalone financial files are optional. Most clinics can keep payment facts
        # embedded in appointment exports; these mappings are only used when a separate
        # source actually exists.
        if mapping.payments is not None:
            payment_mapping = mapping.payments
            payment_rows = _rows(workbook, payment_mapping.sheet)
            for row_index, row in enumerate(payment_rows, start=2):
                path = f"{payment_mapping.sheet}[{row_index}]"
                try:
                    external_id = _required(row, payment_mapping.external_id, path)
                    raw_amount = _required(row, payment_mapping.amount, path)
                    amount_minor = _price_minor(raw_amount, f"{path}.amount")
                    if amount_minor <= 0:
                        raise ClinicImportError(f"{path}.amount: payment amount must be positive.")
                    raw_method = _optional(row, payment_mapping.payment_method)
                    payment_method = _normalize_payment_method(
                        raw_method,
                        payment_mapping.payment_method_map,
                        payment_mapping.default_payment_method,
                    )
                    paid_at = _optional_datetime(
                        _required(row, payment_mapping.paid_at, path),
                        timezone_name=payment_mapping.default_timezone,
                        path=f"{path}.paid_at",
                    )
                    assert paid_at is not None
                    payments.append(
                        NormalizedPaymentImport(
                            external_id=external_id,
                            patient_external_id=_optional(row, payment_mapping.patient_external_id),
                            patient_phone=_optional(row, payment_mapping.patient_phone),
                            appointment_external_id=_optional(row, payment_mapping.appointment_external_id),
                            package_external_id=_optional(row, payment_mapping.package_external_id),
                            amount_minor=amount_minor,
                            payment_method=payment_method,
                            external_reference=_optional(row, payment_mapping.payment_reference),
                            paid_at=paid_at,
                        )
                    )

                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=payment_mapping.sheet,
                        row_kind="payment",
                    )
                    continue
        if mapping.payment_allocations is not None:
            allocation_mapping = mapping.payment_allocations
            allocation_rows = _rows(workbook, allocation_mapping.sheet)
            for row_index, row in enumerate(allocation_rows, start=2):
                path = f"{allocation_mapping.sheet}[{row_index}]"
                try:
                    amount_minor = _price_minor(
                        _required(row, allocation_mapping.amount, path),
                        f"{path}.amount",
                    )
                    if amount_minor <= 0:
                        raise ClinicImportError(f"{path}.amount: allocation amount must be positive.")
                    payment_allocations.append(
                        NormalizedPaymentAllocationImport(
                            payment_external_id=_required(
                                row, allocation_mapping.payment_external_id, path
                            ),
                            appointment_external_id=_required(
                                row, allocation_mapping.appointment_external_id, path
                            ),
                            amount_minor=amount_minor,
                        )
                    )
                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=allocation_mapping.sheet,
                        row_kind="payment_allocation",
                    )
                    continue

        if mapping.packages is not None:
            package_mapping = mapping.packages
            package_rows = _rows(workbook, package_mapping.sheet)
            for row_index, row in enumerate(package_rows, start=2):
                path = f"{package_mapping.sheet}[{row_index}]"
                try:
                    purchased_at = _datetime_value(
                        _required(row, package_mapping.sold_at, path),
                        timezone_name=package_mapping.default_timezone,
                        path=f"{path}.sold_at",
                    )
                    expires_at = None
                    raw_expiry = _raw_optional(row, package_mapping.expires_at)
                    if raw_expiry is not None and str(raw_expiry).strip():
                        expires_at = _date_value(raw_expiry, f"{path}.expires_at")
                    raw_price = _raw_optional(row, package_mapping.sale_price)
                    raw_standalone_price = _raw_optional(
                        row, package_mapping.standalone_session_price_at_purchase
                    )
                    packages.append(
                        NormalizedPackageImport(
                            external_id=_required(row, package_mapping.external_id, path),
                            patient_external_id=_optional(row, package_mapping.patient_external_id),
                            patient_phone=_optional(row, package_mapping.patient_phone),
                            service_external_id=_required(row, package_mapping.service_external_id, path),
                            name=_required(row, package_mapping.name, path),
                            sessions_purchased=_positive_count(
                                _required(row, package_mapping.sessions_purchased, path),
                                f"{path}.sessions_purchased",
                                max_value=1000,
                            ),
                            sale_price_minor=(
                                _price_minor(raw_price, f"{path}.sale_price")
                                if raw_price is not None and str(raw_price).strip()
                                else 0
                            ),
                            standalone_session_price_minor_at_purchase=(
                                _price_minor(
                                    raw_standalone_price,
                                    f"{path}.standalone_session_price_at_purchase",
                                )
                                if raw_standalone_price is not None
                                and str(raw_standalone_price).strip()
                                else None
                            ),
                            purchased_at=purchased_at,
                            expires_at=expires_at,
                            status=_normalize_package_status(
                                _optional(row, package_mapping.status),
                                package_mapping.status_map,
                                package_mapping.default_status,
                            ),
                        )
                    )
                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=package_mapping.sheet,
                        row_kind="package",
                    )
                    continue

        if mapping.package_usages is not None:
            usage_mapping = mapping.package_usages
            usage_rows = _rows(workbook, usage_mapping.sheet)
            for row_index, row in enumerate(usage_rows, start=2):
                path = f"{usage_mapping.sheet}[{row_index}]"
                try:
                    raw_sessions = _optional(row, usage_mapping.sessions_used)
                    used_at = _optional_datetime(
                        _optional(row, usage_mapping.used_at),
                        timezone_name=usage_mapping.default_timezone,
                        path=f"{path}.used_at",
                    )
                    package_usages.append(
                        NormalizedPackageUsageImport(
                            external_id=_optional(row, usage_mapping.external_id),
                            package_external_id=_required(row, usage_mapping.package_external_id, path),
                            appointment_external_id=_required(row, usage_mapping.appointment_external_id, path),
                            sessions_used=(
                                _positive_count(raw_sessions, f"{path}.sessions_used", max_value=100)
                                if raw_sessions is not None
                                else usage_mapping.default_sessions_used
                            ),
                            used_at=used_at,
                        )
                    )
                except ClinicImportError as exc:
                    record_row_data_error(
                        exc=exc,
                        path=path,
                        sheet=usage_mapping.sheet,
                        row_kind="package_usage",
                    )
                    continue
    except (ClinicImportError, StructuralTransformError) as exc:
        issues.append(
            ClinicImportIssue(
                severity="error",
                code=(
                    "structural_transform_error"
                    if isinstance(exc, StructuralTransformError)
                    else "mapping_error"
                ),
                path="mapping",
                message=str(exc),
            )
        )

    for (code, sheet), group in sorted(row_data_error_groups.items()):
        details = " " + " | ".join(group["messages"]) if group["messages"] else ""
        issues.append(
            ClinicImportIssue(
                severity="error",
                code=code,
                path=f"source:{sheet}/rows",
                message=f"{group['summary']}{details}",
                occurrence_count=group["count"],
                source_sheets=[sheet],
                example_paths=group["examples"],
            )
        )

    for sheet_name, count in sorted(unlinked_patient_counts.items()):
        issues.append(
            ClinicImportIssue(
                severity="warning",
                code="appointment_patient_identity_unlinked",
                path=f"appointment_source:{sheet_name}",
                message=(
                    f"{count} appointment rows in {sheet_name!r} have patient names but no stable "
                    "patient ID or phone. Tia preserved every appointment without merging patients "
                    "by name; cross-visit patient history for those rows remains unlinked."
                ),
                occurrence_count=count,
                source_sheets=[sheet_name],
            )
        )

    # Phase 3 deterministic revalidation overrides are applied to normalized
    # canonical records before the same semantic validator below runs. The import
    # engine never invents these values; they come from administrator-confirmed
    # structured fixes.
    if overrides.service_durations:
        by_service = {item.external_id: item for item in services}
        wildcard = overrides.service_durations.get("*")
        if wildcard is not None:
            for item in services:
                item.duration_minutes = wildcard
        for external_id, minutes in overrides.service_durations.items():
            if external_id == "*":
                continue
            item = by_service.get(external_id)
            if item is None:
                raise ClinicImportError(
                    f"Override references unknown service {external_id!r}."
                )
            item.duration_minutes = minutes

    if overrides.doctor_service_assignments or overrides.doctor_branch_assignments or overrides.doctor_active_status:
        by_doctor = {item.external_id: item for item in doctors}
        for external_id, service_ids in overrides.doctor_service_assignments.items():
            item = by_doctor.get(external_id)
            if item is None:
                raise ClinicImportError(
                    f"Override references unknown doctor {external_id!r}."
                )
            item.service_external_ids = list(dict.fromkeys(service_ids))
        for external_id, branch_ids in overrides.doctor_branch_assignments.items():
            item = by_doctor.get(external_id)
            if item is None:
                raise ClinicImportError(
                    f"Override references unknown doctor {external_id!r}."
                )
            item.branch_external_ids = list(dict.fromkeys(branch_ids))
        for external_id, is_active in overrides.doctor_active_status.items():
            item = by_doctor.get(external_id)
            if item is None:
                raise ClinicImportError(
                    f"Override references unknown doctor {external_id!r}."
                )
            item.is_active = bool(is_active)

        if overrides.doctor_branch_assignments or overrides.doctor_active_status:
            # A doctor's current branch assignment is authoritative for current
            # scheduling. Historical source exports may still contain working-hour
            # rows for branches the doctor no longer works at; keep the historical
            # appointments, but do not turn those old schedules into current booking
            # availability.
            current_branches_by_doctor = {
                item.external_id: set(item.branch_external_ids)
                for item in doctors
                if item.is_active
            }
            doctor_hours = [
                item
                for item in doctor_hours
                if item.owner_external_id in current_branches_by_doctor
                and item.branch_external_id in current_branches_by_doctor[item.owner_external_id]
            ]

    if overrides.branch_hours:
        branch_hours = [
            item for item in branch_hours if item.owner_external_id not in overrides.branch_hours
        ]
        for branch_id, schedule in overrides.branch_hours.items():
            for interval in schedule:
                branch_hours.append(
                    NormalizedWorkingHourImport(
                        owner_external_id=branch_id,
                        weekday=interval.weekday,
                        start_time=interval.start_time,
                        end_time=interval.end_time,
                    )
                )

    if overrides.doctor_hours:
        replacement_pairs = {
            (item.doctor_external_id, item.branch_external_id)
            for item in overrides.doctor_hours
        }
        doctor_hours = [
            item
            for item in doctor_hours
            if (item.owner_external_id, item.branch_external_id) not in replacement_pairs
        ]
        for override in overrides.doctor_hours:
            for interval in override.schedule:
                doctor_hours.append(
                    NormalizedWorkingHourImport(
                        owner_external_id=override.doctor_external_id,
                        branch_external_id=override.branch_external_id,
                        weekday=interval.weekday,
                        start_time=interval.start_time,
                        end_time=interval.end_time,
                    )
                )

    # Genuine row-level catalog ambiguity must not block the whole clinic setup.
    # The onboarding service persists the omitted normalized appointment facts in
    # ClinicDataIssue records so they can be materialized after the administrator
    # confirms the alias. Keep any independently attributable payment as a patient-
    # level fact; only defer payments whose sole identity is the omitted appointment.
    deferred_appointment_ids = set(overrides.deferred_appointment_external_ids)
    deferred_payments_from_appointments: set[str] = set()
    if deferred_appointment_ids:
        known_before_defer = {item.external_id for item in appointments}
        unknown_deferred = deferred_appointment_ids - known_before_defer
        if unknown_deferred:
            raise ClinicImportError(
                f"Deferred appointment override references unknown appointment {sorted(unknown_deferred)[0]!r}."
            )
        for payment in payments:
            if payment.appointment_external_id not in deferred_appointment_ids:
                continue
            if payment.patient_external_id or payment.patient_phone:
                payment.appointment_external_id = None
            # If the appointment is the payment's only patient identity, leave the
            # row intact for semantic validation. The payment quarantine layer will
            # persist it as its own critical data issue instead of silently dropping
            # a financial fact.
        payment_allocations = [
            item for item in payment_allocations
            if item.appointment_external_id not in deferred_appointment_ids
            and item.payment_external_id not in deferred_payments_from_appointments
        ]
        package_usages = [
            item for item in package_usages
            if item.appointment_external_id not in deferred_appointment_ids
        ]
        appointments = [
            item for item in appointments
            if item.external_id not in deferred_appointment_ids
        ]

    if (
        overrides.package_patient_assignments
        or overrides.package_service_assignments
        or overrides.package_usage_assignments
        or overrides.excluded_package_usage_appointments
        or overrides.deferred_package_external_ids
        or overrides.deferred_payment_external_ids
        or overrides.deferred_appointment_external_ids
    ):
        known_patient_ids = {item.patient_external_id for item in appointments if item.patient_external_id}
        known_service_ids = {item.external_id for item in services}
        known_appointment_ids = {item.external_id for item in appointments}

        # Quarantine unresolved package facts instead of blocking the whole clinic import.
        # Administrator-confirmed repairs always win over an older deferred decision.
        deferred_package_ids = set(overrides.deferred_package_external_ids)
        deferred_package_ids -= set(overrides.package_patient_assignments)
        deferred_package_ids -= set(overrides.package_service_assignments)
        if deferred_package_ids:
            # Preserve financial facts whenever the payment has an independent patient
            # identity. Only the unverified package relation is removed. Payments whose
            # *only* identity is the deferred package stay quarantined and are persisted
            # in the data-issue context so they can be materialized after repair.
            deferred_payment_ids: set[str] = set()
            for payment in payments:
                if payment.package_external_id not in deferred_package_ids:
                    continue
                has_independent_patient_identity = bool(
                    payment.patient_external_id or payment.patient_phone
                )
                if has_independent_patient_identity:
                    payment.package_external_id = None
                else:
                    deferred_payment_ids.add(payment.external_id)

            packages = [item for item in packages if item.external_id not in deferred_package_ids]
            package_usages = [
                item for item in package_usages if item.package_external_id not in deferred_package_ids
            ]
            payments = [item for item in payments if item.external_id not in deferred_payment_ids]
            if deferred_payment_ids:
                payment_allocations = [
                    item
                    for item in payment_allocations
                    if item.payment_external_id not in deferred_payment_ids
                ]
            for appointment in appointments:
                if appointment.package_external_id not in deferred_package_ids:
                    continue
                appointment.package_external_id = None
                if appointment.billing_context == "package_prepaid":
                    # Preserve the appointment itself, but do not materialize an unverified
                    # entitlement/payment relationship as financial truth.
                    appointment.billing_context = "standard"
                    appointment.payment_status = "unknown"
                    appointment.amount_paid_minor = None
                    appointment.payment_method = "unknown"

        deferred_payment_ids = set(overrides.deferred_payment_external_ids) | deferred_payments_from_appointments
        if deferred_payment_ids:
            payments = [item for item in payments if item.external_id not in deferred_payment_ids]
            payment_allocations = [
                item
                for item in payment_allocations
                if item.payment_external_id not in deferred_payment_ids
            ]

        by_package = {item.external_id: item for item in packages}

        for package_id, patient_id in overrides.package_patient_assignments.items():
            package = by_package.get(package_id)
            if package is None:
                raise ClinicImportError(f"Repair references unknown package {package_id!r}.")
            if patient_id not in known_patient_ids:
                raise ClinicImportError(f"Repair references unknown patient {patient_id!r}.")
            package.patient_external_id = patient_id

        for package_id, service_id in overrides.package_service_assignments.items():
            package = by_package.get(package_id)
            if package is None:
                raise ClinicImportError(f"Repair references unknown package {package_id!r}.")
            if service_id not in known_service_ids:
                raise ClinicImportError(f"Repair references unknown service {service_id!r}.")
            package.service_external_id = service_id

        excluded_usage_appointments = set(overrides.excluded_package_usage_appointments)
        if excluded_usage_appointments:
            unknown = excluded_usage_appointments - known_appointment_ids
            if unknown:
                raise ClinicImportError(
                    f"Repair excludes unknown package-usage appointment {sorted(unknown)[0]!r}."
                )
            package_usages = [
                item
                for item in package_usages
                if item.appointment_external_id not in excluded_usage_appointments
            ]
            for appointment in appointments:
                if appointment.external_id not in excluded_usage_appointments:
                    continue
                appointment.package_external_id = None
                if appointment.billing_context == "package_prepaid":
                    appointment.billing_context = "standard"
                    appointment.payment_status = "unknown"
                    appointment.amount_paid_minor = None
                    appointment.payment_method = "unknown"

        if overrides.package_usage_assignments:
            for appointment_id, package_id in overrides.package_usage_assignments.items():
                if appointment_id not in known_appointment_ids:
                    raise ClinicImportError(
                        f"Repair references unknown appointment {appointment_id!r}."
                    )
                if package_id not in by_package:
                    raise ClinicImportError(f"Repair references unknown package {package_id!r}.")
                matched = False
                for item in package_usages:
                    if item.appointment_external_id == appointment_id:
                        item.package_external_id = package_id
                        matched = True
                if not matched:
                    raise ClinicImportError(
                        f"Repair cannot find package usage for appointment {appointment_id!r}."
                    )

    for item in services:
        if item.duration_minutes is None:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="service_duration_missing",
                    path=f"service:{item.external_id}/duration_minutes",
                    message=f"Duration is missing for service {item.name!r}.",
                )
            )

    def duplicates(values: list[str]) -> set[str]:
        seen: set[str] = set()
        dup: set[str] = set()
        for value in values:
            if value in seen:
                dup.add(value)
            seen.add(value)
        return dup

    for entity, items in (("service", services), ("branch", branches), ("doctor", doctors)):
        by_id = {item.external_id: item for item in items}
        for source_id in sorted(duplicates([item.external_id for item in items])):
            generated = bool(getattr(by_id[source_id], "external_id_generated", False))
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="ambiguous_generated_identity" if generated else "duplicate_external_id",
                    path=entity,
                    message=(
                        f"Multiple {entity} rows resolve to the same generated identity. Add "
                        "disambiguating data or a stable source key."
                        if generated
                        else f"Duplicate {entity} external id {source_id!r}."
                    ),
                )
            )

    def _same_patient_identity(
        left: NormalizedAppointmentImport, right: NormalizedAppointmentImport
    ) -> bool:
        if left.patient_external_id == right.patient_external_id:
            return True
        if left.patient_phone and right.patient_phone:
            try:
                _, left_phone = normalize_patient_identity_phone(left.patient_phone)
                _, right_phone = normalize_patient_identity_phone(right.patient_phone)
            except ValueError:
                return False
            return bool(left_phone and left_phone == right_phone)
        return False

    def _appointment_core_compatible(
        left: NormalizedAppointmentImport, right: NormalizedAppointmentImport
    ) -> bool:
        if not _same_patient_identity(left, right):
            return False
        if left.service_external_id != right.service_external_id:
            return False
        if left.appointment_date != right.appointment_date:
            return False
        if left.start_at is not None and right.start_at is not None and left.start_at != right.start_at:
            return False
        if left.end_at is not None and right.end_at is not None and left.end_at != right.end_at:
            return False
        if (
            left.branch_assignment_known
            and right.branch_assignment_known
            and left.branch_external_id != right.branch_external_id
        ):
            return False
        if (
            left.doctor_assignment_known
            and right.doctor_assignment_known
            and left.doctor_external_id != right.doctor_external_id
        ):
            return False
        return True

    def _merge_duplicate_appointment(
        base: NormalizedAppointmentImport, candidate: NormalizedAppointmentImport
    ) -> bool:
        """Merge richer facts from one duplicate source row without inventing facts.

        Returns False when two rows carrying the same identity contain contradictory terminal
        lifecycle or financial facts that need administrator review.
        """
        terminal = {"completed", "cancelled", "no_show"}
        if base.status in terminal and candidate.status in terminal and base.status != candidate.status:
            return False
        if base.status in {"unknown", "scheduled"} and candidate.status not in {"unknown", "scheduled"}:
            base.status = candidate.status
        elif base.status == "unknown" and candidate.status == "scheduled":
            base.status = "scheduled"

        if base.amount_paid_minor is not None and candidate.amount_paid_minor is not None:
            if base.amount_paid_minor != candidate.amount_paid_minor:
                return False
        elif base.amount_paid_minor is None and candidate.amount_paid_minor is not None:
            base.amount_paid_minor = candidate.amount_paid_minor
            base.payment_status = candidate.payment_status
            base.payment_method = candidate.payment_method
            base.payment_external_reference = candidate.payment_external_reference

        if (
            base.refund_amount_minor is not None
            and candidate.refund_amount_minor is not None
            and base.refund_amount_minor != candidate.refund_amount_minor
        ):
            return False
        if base.refund_amount_minor is None and candidate.refund_amount_minor is not None:
            base.refund_amount_minor = candidate.refund_amount_minor
            base.refund_reason = candidate.refund_reason
            base.refunded_at = candidate.refunded_at

        if base.billing_context != candidate.billing_context:
            if base.billing_context == "standard" and candidate.billing_context == "package_prepaid":
                # Conflicting billing semantics are material: do not silently change a direct-pay
                # appointment into package coverage (or vice versa).
                return False
            if base.billing_context == "package_prepaid" and candidate.billing_context == "standard":
                return False
        if (
            base.package_external_id
            and candidate.package_external_id
            and base.package_external_id != candidate.package_external_id
        ):
            return False
        if not base.package_external_id and candidate.package_external_id:
            base.package_external_id = candidate.package_external_id

        if not base.patient_phone and candidate.patient_phone:
            base.patient_phone = candidate.patient_phone
        if not base.branch_assignment_known and candidate.branch_assignment_known:
            base.branch_external_id = candidate.branch_external_id
            base.branch_assignment_known = True
        if not base.doctor_assignment_known and candidate.doctor_assignment_known:
            base.doctor_external_id = candidate.doctor_external_id
            base.doctor_assignment_known = True
        if base.start_at is None and candidate.start_at is not None:
            base.start_at = candidate.start_at
            base.time_precision = candidate.time_precision
        if base.end_at is None and candidate.end_at is not None:
            base.end_at = candidate.end_at
        if base.source == "other" and candidate.source != "other":
            base.source = candidate.source
        if base.payment_status == "unknown" and candidate.payment_status != "unknown":
            base.payment_status = candidate.payment_status
        if base.payment_method == "unknown" and candidate.payment_method != "unknown":
            base.payment_method = candidate.payment_method
        if not base.payment_external_reference and candidate.payment_external_reference:
            base.payment_external_reference = candidate.payment_external_reference
        return True

    # The same appointment is commonly repeated across current, backup and monthly exports.
    # Collapse compatible repeats before duplicate validation so source redundancy is not a blocker.
    deduped_appointments: list[NormalizedAppointmentImport] = []
    deduped_sheets: list[set[str]] = []
    deduped_paths: list[list[str]] = []
    appointment_index: dict[str, int] = {}
    unresolved_duplicate_ids: set[str] = set()
    collapsed_counts: dict[str, int] = {}

    for item, sheet, origin_path in zip(
        appointments, appointment_origin_sheets, appointment_origin_paths, strict=True
    ):
        existing_index = appointment_index.get(item.external_id)
        if existing_index is None:
            appointment_index[item.external_id] = len(deduped_appointments)
            deduped_appointments.append(item)
            deduped_sheets.append({sheet})
            deduped_paths.append([origin_path])
            continue

        existing = deduped_appointments[existing_index]
        source_is_cross_export = sheet not in deduped_sheets[existing_index]
        can_auto_collapse = (
            _appointment_core_compatible(existing, item)
            and (not item.external_id_generated or source_is_cross_export)
        )
        if can_auto_collapse and _merge_duplicate_appointment(existing, item):
            deduped_sheets[existing_index].add(sheet)
            if len(deduped_paths[existing_index]) < 5:
                deduped_paths[existing_index].append(origin_path)
            collapsed_counts[item.external_id] = collapsed_counts.get(item.external_id, 1) + 1
            continue

        unresolved_duplicate_ids.add(item.external_id)
        # Preserve the conflicting row so downstream preview makes the conflict visible.
        deduped_appointments.append(item)
        deduped_sheets.append({sheet})
        deduped_paths.append([origin_path])

    appointments = deduped_appointments

    for source_id in sorted(collapsed_counts):
        index = appointment_index[source_id]
        issues.append(
            ClinicImportIssue(
                severity="warning",
                code="appointment_duplicate_collapsed",
                path=f"appointment:{source_id}",
                message=(
                    f"{collapsed_counts[source_id]} compatible source rows for appointment "
                    f"{source_id!r} were collapsed into one canonical appointment."
                ),
                occurrence_count=collapsed_counts[source_id],
                source_sheets=sorted(deduped_sheets[index]),
                example_paths=deduped_paths[index],
            )
        )

    for source_id in sorted(unresolved_duplicate_ids):
        item = next(value for value in appointments if value.external_id == source_id)
        issues.append(
            ClinicImportIssue(
                severity="error",
                code="ambiguous_generated_identity" if item.external_id_generated else "duplicate_external_id",
                path="appointment",
                message=(
                    "Multiple appointment rows resolve to the same generated identity but contain "
                    "conflicting facts. Add a booking reference or disambiguate the rows."
                    if item.external_id_generated
                    else (
                        f"Appointment external id {source_id!r} appears more than once with "
                        "conflicting factual data. Review the source rows before import."
                    )
                ),
            )
        )

    service_ids = {item.external_id for item in services}
    branch_ids = {item.external_id for item in branches}
    doctor_ids = {item.external_id for item in doctors}

    deferred_doctor_service_ids = set(overrides.deferred_doctor_service_external_ids)
    deferred_doctor_branch_ids = set(overrides.deferred_doctor_branch_external_ids)
    for doctor in doctors:
        if doctor.is_placeholder or not doctor.is_active:
            continue
        if not doctor.service_external_ids and doctor.external_id not in deferred_doctor_service_ids:
            issues.append(ClinicImportIssue(severity="error", code="doctor_services_missing", path=f"doctor:{doctor.external_id}", message="Doctor has no service assignment."))
        if not doctor.branch_external_ids and doctor.external_id not in deferred_doctor_branch_ids:
            issues.append(ClinicImportIssue(severity="error", code="doctor_branches_missing", path=f"doctor:{doctor.external_id}", message="Doctor has no branch assignment."))
        for external_id in doctor.service_external_ids:
            if external_id not in service_ids:
                issues.append(ClinicImportIssue(severity="error", code="unknown_service_reference", path=f"doctor:{doctor.external_id}", message=f"Doctor references unknown service {external_id!r}."))
        for external_id in doctor.branch_external_ids:
            if external_id not in branch_ids:
                issues.append(ClinicImportIssue(severity="error", code="unknown_branch_reference", path=f"doctor:{doctor.external_id}", message=f"Doctor references unknown branch {external_id!r}."))

    branch_hours_by_branch = {item.owner_external_id for item in branch_hours}
    deferred_branch_hour_ids = set(overrides.deferred_branch_hour_external_ids)
    doctor_hours_by_pair = {(item.owner_external_id, item.branch_external_id) for item in doctor_hours}
    deferred_doctor_hour_pairs = {
        (item.doctor_external_id, item.branch_external_id)
        for item in overrides.deferred_doctor_hour_pairs
    }
    for branch in branches:
        if branch.is_placeholder:
            continue
        if branch.external_id not in branch_hours_by_branch and branch.external_id not in deferred_branch_hour_ids:
            issues.append(ClinicImportIssue(severity="error", code="branch_hours_missing", path=f"branch:{branch.external_id}", message="Branch has no structured working hours."))
    for doctor in doctors:
        if doctor.is_placeholder or not doctor.is_active:
            continue
        for branch_id in doctor.branch_external_ids:
            pair = (doctor.external_id, branch_id)
            if pair not in doctor_hours_by_pair and pair not in deferred_doctor_hour_pairs:
                issues.append(ClinicImportIssue(severity="error", code="doctor_hours_missing", path=f"doctor:{doctor.external_id}/branch:{branch_id}", message=f"Doctor has no working hours for branch {branch_id!r}."))

    for item in branch_hours:
        if item.owner_external_id not in branch_ids:
            issues.append(ClinicImportIssue(severity="error", code="unknown_branch_hours_reference", path="branch_hours", message=f"Working hours reference unknown branch {item.owner_external_id!r}."))
    for item in doctor_hours:
        if item.owner_external_id not in doctor_ids:
            issues.append(ClinicImportIssue(severity="error", code="unknown_doctor_hours_reference", path="doctor_hours", message=f"Working hours reference unknown doctor {item.owner_external_id!r}."))
        if item.branch_external_id not in branch_ids:
            issues.append(ClinicImportIssue(severity="error", code="unknown_doctor_branch_hours_reference", path="doctor_hours", message=f"Working hours reference unknown branch {item.branch_external_id!r}."))
        if item.start_time >= item.end_time:
            issues.append(ClinicImportIssue(severity="error", code="invalid_working_hour_interval", path="doctor_hours", message="Working-hour end time must be after start time."))
    for item in branch_hours:
        if item.start_time >= item.end_time:
            issues.append(ClinicImportIssue(severity="error", code="invalid_working_hour_interval", path="branch_hours", message="Working-hour end time must be after start time."))

    reference_issue_codes = {
        "service": "unknown_appointment_service_reference",
        "branch": "unknown_appointment_branch_reference",
        "doctor": "unknown_appointment_doctor_reference",
    }
    reference_labels = {"service": "service", "branch": "branch", "doctor": "doctor"}
    # Recompute unresolved vocabulary from the appointments that are still part of
    # this preview. Rows quarantined by onboarding must not keep their grouped alias
    # error alive after the unsafe appointment fact itself has been deferred.
    remaining_unresolved_reference_counts: dict[tuple[str, str], int] = defaultdict(int)
    for appointment in appointments:
        if appointment.service_external_id not in service_ids:
            remaining_unresolved_reference_counts[("service", _identity_text(appointment.service_external_id))] += 1
        if appointment.branch_external_id not in branch_ids:
            remaining_unresolved_reference_counts[("branch", _identity_text(appointment.branch_external_id))] += 1
        if appointment.doctor_assignment_known and appointment.doctor_external_id not in doctor_ids:
            remaining_unresolved_reference_counts[("doctor", _identity_text(appointment.doctor_external_id))] += 1
    for (kind, _normalized_value), group in sorted(
        unresolved_reference_groups.items(), key=lambda item: (item[0][0], item[0][1])
    ):
        source_value = str(group["raw_value"])
        count = int(remaining_unresolved_reference_counts.get((kind, _normalized_value), 0))
        if count <= 0:
            continue
        source_sheets = sorted(str(value) for value in group["sheets"])
        examples = [str(value) for value in group["examples"]]
        issues.append(
            ClinicImportIssue(
                severity="error",
                code=reference_issue_codes[kind],
                path=f"appointment_reference:{kind}:{_identity_text(source_value)}",
                message=(
                    f"Appointment {reference_labels[kind]} reference {source_value!r} is unresolved "
                    f"in {count} appointment row(s). Confirm one mapping for this source value."
                ),
                occurrence_count=count,
                source_value=source_value,
                reference_kind=kind,
                source_sheets=source_sheets,
                example_paths=examples,
            )
        )

    for _normalized_status, group in sorted(unmapped_status_groups.items()):
        raw_value = str(group["raw_value"])
        count = int(group["count"])
        issues.append(
            ClinicImportIssue(
                severity="error",
                code="appointment_status_mapping_missing",
                path=f"appointment_status:{_normalized_status}",
                message=(
                    f"Appointment status {raw_value!r} is not mapped in {count} appointment row(s); "
                    "map this source value before import so Tia does not invent an operational status."
                ),
                occurrence_count=count,
                source_value=raw_value,
                source_sheets=sorted(str(value) for value in group["sheets"]),
                example_paths=[str(value) for value in group["examples"]],
            )
        )

    for sheet_name, group in sorted(missing_exact_time_groups.items()):
        count = int(group["count"])
        issues.append(
            ClinicImportIssue(
                severity="error",
                code="appointment_exact_time_missing",
                path=f"appointment_source:{sheet_name}/exact_time",
                message=(
                    f"{count} appointment row(s) in {sheet_name!r} have a date but no exact start time. "
                    "They can be preserved as history, but availability cannot be activated safely "
                    "until exact times are supplied."
                ),
                occurrence_count=count,
                source_sheets=[sheet_name],
                example_paths=[str(value) for value in group["examples"]],
            )
        )


    payment_by_external = {item.external_id: item for item in payments}
    duplicate_payment_ids = duplicates([item.external_id for item in payments])
    for source_id in sorted(duplicate_payment_ids):
        issues.append(
            ClinicImportIssue(
                severity="error",
                code="duplicate_payment_external_id",
                path="payment",
                message=f"Duplicate payment external id {source_id!r}.",
            )
        )

    appointment_by_external = {item.external_id: item for item in appointments}
    patient_ids_from_appointments = {item.patient_external_id for item in appointments}
    patient_phones_from_appointments: set[str] = set()
    for appointment in appointments:
        if not appointment.patient_phone:
            continue
        try:
            _display_phone, normalized_phone = normalize_patient_identity_phone(appointment.patient_phone)
        except ValueError:
            normalized_phone = re.sub(r"\D+", "", appointment.patient_phone)
        if normalized_phone:
            patient_phones_from_appointments.add(normalized_phone)

    # Explicit allocation rows are authoritative. A payment may remain entirely
    # patient-level/unallocated; Tia never chooses the "nearest" appointment.
    explicit_allocations_by_payment: dict[str, list[NormalizedPaymentAllocationImport]] = {}
    seen_allocation_pairs: set[tuple[str, str]] = set()
    for allocation in payment_allocations:
        pair = (allocation.payment_external_id, allocation.appointment_external_id)
        if pair in seen_allocation_pairs:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="duplicate_payment_allocation",
                    path="payment_allocations",
                    message=(
                        f"Payment {allocation.payment_external_id!r} contains a duplicate allocation "
                        f"for appointment {allocation.appointment_external_id!r}."
                    ),
                )
            )
        seen_allocation_pairs.add(pair)
        payment = payment_by_external.get(allocation.payment_external_id)
        appointment = appointment_by_external.get(allocation.appointment_external_id)
        if payment is None:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_payment_allocation_payment",
                    path="payment_allocations",
                    message=f"Allocation references unknown payment {allocation.payment_external_id!r}.",
                )
            )
            continue
        if appointment is None:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_payment_allocation_appointment",
                    path="payment_allocations",
                    message=(
                        f"Allocation references unknown appointment "
                        f"{allocation.appointment_external_id!r}."
                    ),
                )
            )
            continue
        explicit_allocations_by_payment.setdefault(allocation.payment_external_id, []).append(allocation)

    package_by_external = {item.external_id: item for item in packages}

    # Validate payment patient identity only against facts imported in the same batch.
    # Existing-workspace links are revalidated again during Apply.
    for payment in payments:
        linked_appointment = (
            appointment_by_external.get(payment.appointment_external_id)
            if payment.appointment_external_id
            else None
        )
        linked_package = (
            package_by_external.get(payment.package_external_id)
            if payment.package_external_id
            else None
        )
        if payment.appointment_external_id and linked_appointment is None:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_payment_appointment_reference",
                    path=f"payment:{payment.external_id}",
                    message=(
                        f"Payment {payment.external_id!r} references unknown appointment "
                        f"{payment.appointment_external_id!r}."
                    ),
                )
            )
        if payment.package_external_id and linked_package is None:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_payment_package_reference",
                    path=f"payment:{payment.external_id}",
                    message=(
                        f"Payment {payment.external_id!r} references unknown package "
                        f"{payment.package_external_id!r}."
                    ),
                )
            )
        if (
            payment.patient_external_id is None
            and payment.patient_phone is None
            and linked_appointment is None
            and linked_package is None
        ):
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="payment_patient_identity_missing",
                    path=f"payment:{payment.external_id}",
                    message="Standalone payment has no patient, appointment, or package identity.",
                )
            )
        payment_phone_matches = False
        if payment.patient_phone and linked_appointment is None:
            try:
                _display_phone, phone_key = normalize_patient_identity_phone(payment.patient_phone)
            except ValueError:
                phone_key = re.sub(r"\D+", "", payment.patient_phone)
            payment_phone_matches = bool(phone_key and phone_key in patient_phones_from_appointments)
            if not payment_phone_matches:
                issues.append(
                    ClinicImportIssue(
                        severity="error",
                        code="unknown_payment_patient_phone",
                        path=f"payment:{payment.external_id}",
                        message=(
                            f"Payment {payment.external_id!r} has a patient phone that does not match "
                            "any patient in the imported appointment history."
                        ),
                    )
                )
        if (
            payment.patient_external_id is not None
            and payment.patient_external_id not in patient_ids_from_appointments
            and linked_appointment is None
            and not payment_phone_matches
        ):
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_payment_patient_reference",
                    path=f"payment:{payment.external_id}",
                    message=(
                        f"Payment {payment.external_id!r} references patient "
                        f"{payment.patient_external_id!r}, which is not present in imported appointments "
                        "and cannot be reconciled by normalized phone."
                    ),
                )
            )

        allocations = explicit_allocations_by_payment.get(payment.external_id, [])
        if payment.package_external_id and (allocations or payment.appointment_external_id):
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="package_payment_has_appointment_allocation",
                    path=f"payment:{payment.external_id}",
                    message=(
                        f"Payment {payment.external_id!r} is linked to package "
                        f"{payment.package_external_id!r} and cannot also be allocated to an appointment."
                    ),
                )
            )
        if allocations:
            allocated_total = sum(item.amount_minor for item in allocations)
            if allocated_total > payment.amount_minor:
                issues.append(
                    ClinicImportIssue(
                        severity="error",
                        code="payment_allocation_exceeds_payment",
                        path=f"payment:{payment.external_id}",
                        message=(
                            f"Allocations for payment {payment.external_id!r} total {allocated_total} "
                            f"minor units, exceeding payment amount {payment.amount_minor}."
                        ),
                    )
                )
            # Allocation patient must agree with the payment patient when both are known.
            if payment.patient_external_id:
                for allocation in allocations:
                    appointment = appointment_by_external.get(allocation.appointment_external_id)
                    if appointment and appointment.patient_external_id != payment.patient_external_id:
                        issues.append(
                            ClinicImportIssue(
                                severity="error",
                                code="payment_patient_mismatch",
                                path=f"payment:{payment.external_id}",
                                message=(
                                    f"Payment {payment.external_id!r} is allocated to an appointment "
                                    "belonging to a different patient."
                                ),
                            )
                        )
        elif payment.appointment_external_id:
            # A direct appointment reference is itself an explicit full-allocation fact.
            payment_allocations.append(
                NormalizedPaymentAllocationImport(
                    payment_external_id=payment.external_id,
                    appointment_external_id=payment.appointment_external_id,
                    amount_minor=payment.amount_minor,
                )
            )
        elif payment.external_reference and not payment.package_external_id:
            # Merge a standalone transaction with appointment-embedded references when
            # both sources describe the same receipt. This avoids double counting while
            # preserving explicit appointment linkage from the booking export.
            embedded = [
                item
                for item in appointments
                if item.payment_external_reference == payment.external_reference
                and item.amount_paid_minor is not None
                and item.billing_context == "standard"
            ]
            if embedded:
                embedded_total = sum(int(item.amount_paid_minor or 0) for item in embedded)
                if embedded_total <= payment.amount_minor:
                    for item in embedded:
                        payment_allocations.append(
                            NormalizedPaymentAllocationImport(
                                payment_external_id=payment.external_id,
                                appointment_external_id=item.external_id,
                                amount_minor=int(item.amount_paid_minor or 0),
                            )
                        )

    # Optional prepaid-package sources are first-class entitlement facts. They are
    # validated against imported patients/services/appointments without inventing links.
    for source_id in sorted(duplicates([item.external_id for item in packages])):
        issues.append(
            ClinicImportIssue(
                severity="error",
                code="duplicate_package_external_id",
                path="packages",
                message=f"Duplicate package external id {source_id!r}.",
            )
        )

    patient_label_by_id: dict[str, dict[str, str | None]] = {}
    for appointment in appointments:
        patient_id = appointment.patient_external_id
        current = patient_label_by_id.get(patient_id)
        if current is None:
            patient_label_by_id[patient_id] = {
                "patient_external_id": patient_id,
                "patient_name": appointment.patient_name,
                "patient_phone": appointment.patient_phone,
            }
        elif not current.get("patient_phone") and appointment.patient_phone:
            current["patient_phone"] = appointment.patient_phone

    usage_rows_by_package: dict[str, list[NormalizedPackageUsageImport]] = {}
    for usage in package_usages:
        usage_rows_by_package.setdefault(usage.package_external_id, []).append(usage)

    def package_matches_appointment_patient(
        package: NormalizedPackageImport,
        appointment: NormalizedAppointmentImport,
    ) -> bool:
        if package.patient_external_id and package.patient_external_id == appointment.patient_external_id:
            return True
        if package.patient_phone and appointment.patient_phone:
            try:
                return (
                    normalize_patient_identity_phone(package.patient_phone)[1]
                    == normalize_patient_identity_phone(appointment.patient_phone)[1]
                )
            except ValueError:
                return False
        return False

    def package_candidate_options_for_appointment(
        appointment: NormalizedAppointmentImport,
    ) -> list[dict[str, Any]]:
        candidates: list[dict[str, Any]] = []
        for candidate in packages:
            if candidate.service_external_id != appointment.service_external_id:
                continue
            if not package_matches_appointment_patient(candidate, appointment):
                continue
            if candidate.purchased_at.date() > appointment.appointment_date:
                continue
            if candidate.expires_at is not None and candidate.expires_at < appointment.appointment_date:
                continue
            candidates.append(
                {
                    "label": candidate.name or candidate.external_id,
                    "detail": f"الباقة {candidate.external_id}",
                    "recommended": False,
                    "fix": {
                        "kind": "package_usage_assignment",
                        "appointment_external_id": appointment.external_id,
                        "package_external_id": candidate.external_id,
                    },
                }
            )
        if len(candidates) == 1:
            candidates[0]["recommended"] = True
        return candidates[:10]

    normalized_package_phone: dict[str, str] = {}
    for package in packages:
        if package.service_external_id not in service_by_key:
            package_service_key = _identity_text(package.service_external_id)
            matching_service_ids = {
                service.external_id
                for service in services
                if package_service_key
                and package_service_key
                in {_identity_text(service.external_id), _identity_text(service.name)}
            }
            service_repair_options = [
                {
                    "label": service.name,
                    "detail": f"الخدمة {service.external_id}",
                    "recommended": service.external_id in matching_service_ids and len(matching_service_ids) == 1,
                    "auto_apply": service.external_id in matching_service_ids and len(matching_service_ids) == 1,
                    "fix": {
                        "kind": "package_service_assignment",
                        "package_external_id": package.external_id,
                        "service_external_id": service.external_id,
                    },
                }
                for service in services[:30]
            ]
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_package_service",
                    path=f"package:{package.external_id}",
                    message=(
                        f"Package {package.external_id!r} references unknown service "
                        f"{package.service_external_id!r}."
                    ),
                    repair_group=f"package:{package.external_id}:service",
                    repair_title=f"تحديد خدمة الباقة «{package.name or package.external_id}»",
                    repair_detail=(
                        "اختار الخدمة المقابلة من كتالوج العيادة مرة واحدة. Tia هتستخدم الاختيار "
                        "داخل الاستيراد من غير تعديل الملف الأصلي."
                    ),
                    repair_options=service_repair_options,
                    package_external_id=package.external_id,
                )
            )
        if package.expires_at is not None and package.expires_at < package.purchased_at.date():
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="package_expiry_before_purchase",
                    path=f"package:{package.external_id}",
                    message=f"Package {package.external_id!r} expires before it was purchased.",
                )
            )
        phone_match = False
        if package.patient_phone:
            try:
                _display, phone_key = normalize_patient_identity_phone(package.patient_phone)
            except ValueError:
                phone_key = re.sub(r"\D+", "", package.patient_phone)
            if phone_key:
                normalized_package_phone[package.external_id] = phone_key
                phone_match = phone_key in patient_phones_from_appointments
        if (
            package.patient_external_id not in patient_ids_from_appointments
            if package.patient_external_id is not None
            else True
        ) and not phone_match:
            patient_counts: dict[str, int] = {}
            usage_patient_ids: set[str] = set()
            for usage in usage_rows_by_package.get(package.external_id, []):
                appointment = appointment_by_external.get(usage.appointment_external_id)
                if appointment is None:
                    continue
                patient_counts[appointment.patient_external_id] = (
                    patient_counts.get(appointment.patient_external_id, 0) + 1
                )
                usage_patient_ids.add(appointment.patient_external_id)

            # A package-linked payment may carry a stable patient ID/phone even when
            # the package export itself does not. Treat that as strong deterministic
            # evidence instead of asking the administrator to repair the source file.
            payment_patient_ids: set[str] = set()
            patient_ids_by_phone: dict[str, set[str]] = {}
            for appointment in appointments:
                if not appointment.patient_phone:
                    continue
                try:
                    _display, appointment_phone_key = normalize_patient_identity_phone(appointment.patient_phone)
                except ValueError:
                    appointment_phone_key = re.sub(r"\D+", "", appointment.patient_phone)
                if appointment_phone_key:
                    patient_ids_by_phone.setdefault(appointment_phone_key, set()).add(appointment.patient_external_id)
            for payment in payments:
                if payment.package_external_id != package.external_id:
                    continue
                candidate_patient_id = None
                if payment.patient_external_id in patient_ids_from_appointments:
                    candidate_patient_id = payment.patient_external_id
                elif payment.patient_phone:
                    try:
                        _display, payment_phone_key = normalize_patient_identity_phone(payment.patient_phone)
                    except ValueError:
                        payment_phone_key = re.sub(r"\D+", "", payment.patient_phone)
                    matching_ids = patient_ids_by_phone.get(payment_phone_key, set())
                    if len(matching_ids) == 1:
                        candidate_patient_id = next(iter(matching_ids))
                if candidate_patient_id:
                    payment_patient_ids.add(candidate_patient_id)
                    patient_counts[candidate_patient_id] = patient_counts.get(candidate_patient_id, 0) + 2

            evidence_source = "usage" if usage_patient_ids else ("payment" if payment_patient_ids else "service_history")
            if not patient_counts:
                # No explicit package-usage export: use same-service appointments
                # inside the package lifetime only as a shortlist, never as an
                # automatic identity decision.
                evidence_source = "service_history"
                for appointment in appointments:
                    if appointment.service_external_id != package.service_external_id:
                        continue
                    if appointment.appointment_date < package.purchased_at.date():
                        continue
                    if package.expires_at is not None and appointment.appointment_date > package.expires_at:
                        continue
                    patient_counts[appointment.patient_external_id] = (
                        patient_counts.get(appointment.patient_external_id, 0) + 1
                    )
            repair_options: list[dict[str, Any]] = []
            ranked_patients = sorted(
                patient_counts.items(), key=lambda item: (-item[1], item[0])
            )
            total_usage_evidence = sum(patient_counts.values())
            for patient_id, count in ranked_patients[:8]:
                label = patient_label_by_id.get(patient_id) or {}
                patient_name = str(label.get("patient_name") or patient_id)
                repair_options.append(
                    {
                        "label": f"ربط الباقة بـ {patient_name}",
                        "detail": (
                            "كل الأدلة المتاحة من جلسات/دفعات الباقة تشير لهذا العميل."
                            if evidence_source in {"usage", "payment"} and len(ranked_patients) == 1
                            else (
                                f"{count} نقاط دليل من استخدامات/دفعات الباقة مرتبطة بهذا العميل."
                                if evidence_source in {"usage", "payment"}
                                else f"عنده {count} موعد لنفس خدمة الباقة خلال فترة صلاحيتها."
                            )
                        ),
                        "recommended": evidence_source in {"usage", "payment"} and len(ranked_patients) == 1,
                        "auto_apply": (
                            evidence_source in {"usage", "payment"}
                            and len(ranked_patients) == 1
                            and (total_usage_evidence >= 2 or bool(payment_patient_ids))
                        ),
                        "fix": {
                            "kind": "package_patient_assignment",
                            "package_external_id": package.external_id,
                            "patient_external_id": patient_id,
                        },
                    }
                )
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_package_patient",
                    path=f"package:{package.external_id}",
                    message=(
                        f"Package {package.external_id!r} cannot be linked to an imported patient "
                        "by stable patient ID or normalized phone."
                    ),
                    repair_group=f"package:{package.external_id}",
                    repair_title=f"تأكيد صاحب الباقة «{package.name or package.external_id}»",
                    repair_detail=(
                        "Tia راجعت الجلسات المرتبطة بالباقة وطلعت العملاء الأكثر احتمالًا. "
                        "اختر العميل الصحيح مرة واحدة؛ مش محتاج تعدّل ملف المصدر."
                    ),
                    repair_options=repair_options,
                    package_external_id=package.external_id,
                )
            )

    usage_totals: dict[str, int] = {}
    seen_usage_appointments: set[str] = set()
    seen_usage_external_ids: set[str] = set()
    for usage in package_usages:
        if usage.external_id:
            if usage.external_id in seen_usage_external_ids:
                issues.append(
                    ClinicImportIssue(
                        severity="error",
                        code="duplicate_package_usage_external_id",
                        path="package_usages",
                        message=f"Duplicate package usage external id {usage.external_id!r}.",
                    )
                )
            seen_usage_external_ids.add(usage.external_id)
        if usage.appointment_external_id in seen_usage_appointments:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="duplicate_package_usage_appointment",
                    path="package_usages",
                    message=(
                        f"Appointment {usage.appointment_external_id!r} is linked to more than one "
                        "package usage row."
                    ),
                )
            )
        seen_usage_appointments.add(usage.appointment_external_id)
        package = package_by_external.get(usage.package_external_id)
        appointment = appointment_by_external.get(usage.appointment_external_id)
        if package is None:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_package_usage_package",
                    path="package_usages",
                    message=f"Usage references unknown package {usage.package_external_id!r}.",
                )
            )
            continue
        if appointment is None:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="unknown_package_usage_appointment",
                    path="package_usages",
                    message=f"Usage references unknown appointment {usage.appointment_external_id!r}.",
                )
            )
            continue
        if appointment.service_external_id != package.service_external_id:
            service_candidates = package_candidate_options_for_appointment(appointment)
            service_candidates.append(
                {
                    "label": "الجلسة ليست استخدام باقة",
                    "detail": "احتفظ بالموعد، لكن لا تخصم جلسة من أي باقة.",
                    "recommended": False,
                    "fix": {
                        "kind": "package_usage_exclusion",
                        "appointment_external_id": appointment.external_id,
                    },
                }
            )
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="package_usage_service_mismatch",
                    path=f"package_usage:{usage.appointment_external_id}",
                    message="Package usage appointment belongs to a different service.",
                    repair_group=f"package_usage:{usage.appointment_external_id}",
                    repair_title=f"مراجعة استخدام الباقة في الموعد {usage.appointment_external_id}",
                    repair_detail=(
                        "الخدمة في الموعد مختلفة عن خدمة الباقة المسجلة. اختر الباقة الصحيحة "
                        "أو اعتبر الجلسة خارج الباقات."
                    ),
                    repair_options=service_candidates,
                    package_external_id=usage.package_external_id,
                    appointment_external_id=usage.appointment_external_id,
                    patient_external_id=appointment.patient_external_id,
                    service_external_id=appointment.service_external_id,
                )
            )
        patient_matches = bool(
            package.patient_external_id
            and appointment.patient_external_id == package.patient_external_id
        )
        if not patient_matches and package.patient_phone and appointment.patient_phone:
            try:
                patient_matches = (
                    normalize_patient_identity_phone(package.patient_phone)[1]
                    == normalize_patient_identity_phone(appointment.patient_phone)[1]
                )
            except ValueError:
                patient_matches = False
        if not patient_matches:
            patient_candidates = package_candidate_options_for_appointment(appointment)
            patient_candidates.append(
                {
                    "label": "الجلسة ليست استخدام باقة",
                    "detail": "احتفظ بالموعد، لكن لا تخصم جلسة من أي باقة.",
                    "recommended": False,
                    "fix": {
                        "kind": "package_usage_exclusion",
                        "appointment_external_id": appointment.external_id,
                    },
                }
            )
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="package_usage_patient_mismatch",
                    path=f"package_usage:{usage.appointment_external_id}",
                    message="Package usage appointment belongs to a different patient.",
                    repair_group=f"package_usage:{usage.appointment_external_id}",
                    repair_title=f"مراجعة استخدام الباقة في الموعد {usage.appointment_external_id}",
                    repair_detail=(
                        "الموعد مسجل لعميل مختلف عن صاحب الباقة الحالية. Tia جهزت الباقات "
                        "المتوافقة مع العميل والخدمة عشان تختار من غير تعديل الملف."
                    ),
                    repair_options=patient_candidates,
                    package_external_id=usage.package_external_id,
                    appointment_external_id=usage.appointment_external_id,
                    patient_external_id=appointment.patient_external_id,
                    service_external_id=appointment.service_external_id,
                )
            )
        usage_totals[usage.package_external_id] = (
            usage_totals.get(usage.package_external_id, 0) + usage.sessions_used
        )

    # Appointment-level package_prepaid markers are explicit entitlement usage when no
    # dedicated usage row exists. Count them too so an overused 6-session package is
    # caught even when the clinic has no package_usage export.
    for appointment in appointments:
        if (
            appointment.billing_context == "package_prepaid"
            and appointment.package_external_id
            and appointment.external_id not in seen_usage_appointments
            and appointment.status not in {"cancelled", "no_show"}
        ):
            usage_totals[appointment.package_external_id] = (
                usage_totals.get(appointment.package_external_id, 0) + 1
            )

    for package_external_id, total in usage_totals.items():
        package = package_by_external.get(package_external_id)
        if package is not None and total > package.sessions_purchased:
            issues.append(
                ClinicImportIssue(
                    severity="error",
                    code="package_usage_exceeds_entitlement",
                    path=f"package:{package_external_id}",
                    message=(
                        f"Package {package_external_id!r} has {total} used sessions but only "
                        f"{package.sessions_purchased} purchased sessions."
                    ),
                )
            )

    if packages:
        for appointment in appointments:
            if appointment.billing_context != "package_prepaid":
                continue
            if not appointment.package_external_id:
                issues.append(
                    ClinicImportIssue(
                        severity="error",
                        code="package_prepaid_reference_missing",
                        path=f"appointment:{appointment.external_id}",
                        message="Package-prepaid appointment is missing its package reference.",
                    )
                )
            elif appointment.package_external_id not in package_by_external:
                issues.append(
                    ClinicImportIssue(
                        severity="error",
                        code="unknown_appointment_package_reference",
                        path=f"appointment:{appointment.external_id}",
                        message=(
                            f"Appointment {appointment.external_id!r} references unknown package "
                            f"{appointment.package_external_id!r}."
                        ),
                    )
                )

    if not _appointment_mappings(mapping) and not confirm_no_existing_appointments:
        issues.append(
            ClinicImportIssue(
                severity="error",
                code="existing_appointments_unaccounted_for",
                path="mapping.appointments",
                message=(
                    "Existing appointments must be imported so availability cannot double-book, "
                    "or confirm_no_existing_appointments must be true."
                ),
            )
        )

    no_errors = not any(issue.severity == "error" for issue in issues)
    real_doctors = [item for item in doctors if not item.is_placeholder]
    catalog_ready = no_errors and bool(services and branches)
    # Historical data can be imported without a recorded doctor catalog. Operational booking
    # stays disabled until at least one real doctor exists with structured availability.
    booking_ready = no_errors and bool(real_doctors and doctor_hours and branch_hours)
    appointments_read_ready = no_errors and bool(appointments)
    capabilities = ClinicImportCapabilities(
        catalog_read=catalog_ready,
        availability_read=booking_ready,
        appointments_read=appointments_read_ready,
        appointments_create=booking_ready,
        appointments_confirm=booking_ready,
        appointments_cancel=booking_ready,
        appointments_reschedule=booking_ready,
    )
    return ClinicImportPreviewResponse(
        services=services,
        branches=branches,
        doctors=doctors,
        branch_hours=branch_hours,
        doctor_hours=doctor_hours,
        appointments=appointments,
        payments=payments,
        payment_allocations=payment_allocations,
        packages=packages,
        package_usages=package_usages,
        issues=issues,
        capabilities=capabilities,
        # Import safety and booking readiness are separate concerns. A clinic with
        # real doctors but no trustworthy current schedule can still finish setup;
        # availability/create capabilities remain disabled until hours exist.
        can_apply=no_errors and capabilities.catalog_read,
        source_summary={
            "documents": list(workbook.document_names),
            "sheets": raw_sheet_summary,
            "rows": sum(raw_sheet_summary.values()),
            "transformed_sheets": transformed_sheet_summary,
            "currency": "EGP",
        },
    )
