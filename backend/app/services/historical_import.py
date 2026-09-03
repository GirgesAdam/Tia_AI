from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import re
from collections import Counter, defaultdict
from collections.abc import Iterable
from datetime import UTC, date, datetime, time, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from uuid import UUID
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select, update
from sqlalchemy.orm import Session

from app.models.agent_action import AgentAction
from app.models.appointment import Appointment
from app.models.automation_job import AutomationJob
from app.models.crm_campaign_conversion import CRMCampaignConversion
from app.models.doctor import Doctor
from app.models.doctor_branch import DoctorBranch
from app.models.doctor_service import DoctorService
from app.models.historical_import import (
    ClinicHistoricalImportBatch,
    ClinicHistoricalImportLink,
    ClinicHistoricalImportRow,
)
from app.models.patient import Patient
from app.models.patient_package import PackageUsage, PatientPackage
from app.models.payment_transaction import PaymentAllocation, PaymentTransaction
from app.models.service import Service
from app.models.staff import Staff
from app.models.workspace import Workspace
from app.schemas.crm import normalize_patient_identity_phone
from app.schemas.historical_import import (
    HistoricalImportBatchRead,
    HistoricalImportDocument,
    HistoricalImportIssueGroup,
    HistoricalImportPreviewResponse,
)
from app.services.activity import record_activity_event

SCHEMA_VERSION = "tia_history_v1"
MAX_DOCUMENT_BYTES = 25 * 1024 * 1024
MAX_UPLOAD_BYTES = 60 * 1024 * 1024
MAX_IMPORT_ROWS = 250_000
RECOGNIZED_SHEETS = {
    "patients": "patient",
    "appointments": "appointment",
    "payments": "payment",
    "payment_allocations": "payment_allocation",
    "packages": "package",
}
EGYPT_TZ = ZoneInfo("Africa/Cairo")
VALID_APPOINTMENT_STATUSES = {
    "pending",
    "confirmed",
    "checked_in",
    "in_progress",
    "completed",
    "cancelled",
    "no_show",
    "rescheduled",
}
VALID_PAYMENT_METHODS = {"unknown", "cash", "card", "bank_transfer", "wallet", "online", "other"}
VALID_PACKAGE_STATUSES = {"active", "expired", "cancelled"}
VALID_PATIENT_SOURCES = {
    "whatsapp", "instagram", "facebook", "website", "referral", "walk_in",
    "campaign", "phone", "other",
}


class HistoricalImportError(ValueError):
    pass


class HistoricalImportConflictError(HistoricalImportError):
    pass


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _header(value: Any) -> str:
    text = _clean(value) or ""
    text = text.casefold().replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def _digest(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def _payload_hash(payload: dict[str, Any]) -> str:
    return _digest(json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str, separators=(",", ":")))


def _decode(document: HistoricalImportDocument) -> bytes:
    try:
        return base64.b64decode(document.content_base64, validate=True)
    except Exception as exc:  # pragma: no cover - defensive boundary
        raise HistoricalImportError(f"{document.name}: invalid file encoding.") from exc




def _validate_documents(documents: list[HistoricalImportDocument]) -> None:
    seen_names: set[str] = set()
    total = 0
    for document in documents:
        key = document.name.casefold()
        if key in seen_names:
            raise HistoricalImportError(f"Duplicate upload filename: {document.name}.")
        seen_names.add(key)
        raw = _decode(document)
        size = len(raw)
        if size > MAX_DOCUMENT_BYTES:
            raise HistoricalImportError(f"{document.name}: file is too large for historical import.")
        total += size
        if total > MAX_UPLOAD_BYTES:
            raise HistoricalImportError("Historical import files are too large in total.")

def _source_fingerprint(documents: list[HistoricalImportDocument]) -> str:
    h = hashlib.sha256()
    for document in sorted(documents, key=lambda item: item.name.casefold()):
        raw = _decode(document)
        h.update(document.name.encode("utf-8"))
        h.update(document.format.encode("ascii"))
        h.update(raw)
    return h.hexdigest()


def _rows_from_xlsx(document: HistoricalImportDocument) -> Iterable[tuple[str, int, dict[str, Any]]]:
    raw = _decode(document)
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise HistoricalImportError(f"{document.name}: unreadable Excel workbook.") from exc
    try:
        for sheet in workbook.worksheets:
            key = _header(sheet.title)
            if key not in RECOGNIZED_SHEETS:
                continue
            iterator = sheet.iter_rows(values_only=True)
            try:
                headers = [_header(cell) for cell in next(iterator)]
            except StopIteration:
                continue
            if not any(headers):
                continue
            for row_number, values in enumerate(iterator, start=2):
                if not any(_clean(value) is not None for value in values):
                    continue
                row = {headers[index]: values[index] for index in range(min(len(headers), len(values))) if headers[index]}
                yield key, row_number, row
    finally:
        workbook.close()


def _rows_from_csv(document: HistoricalImportDocument) -> Iterable[tuple[str, int, dict[str, Any]]]:
    key = _header(Path(document.name).stem)
    if key not in RECOGNIZED_SHEETS:
        raise HistoricalImportError(
            f"{document.name}: CSV filename must be one of {', '.join(sorted(RECOGNIZED_SHEETS))}."
        )
    raw = _decode(document)
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise HistoricalImportError(f"{document.name}: CSV must be UTF-8.") from exc
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return
    headers = {_header(name): name for name in reader.fieldnames if _header(name)}
    for row_number, raw_row in enumerate(reader, start=2):
        row = {key_name: raw_row.get(original) for key_name, original in headers.items()}
        if any(_clean(value) is not None for value in row.values()):
            yield key, row_number, row


def _iter_rows(documents: list[HistoricalImportDocument]) -> Iterable[tuple[str, str, int, dict[str, Any]]]:
    recognized = False
    for document in documents:
        iterator = _rows_from_xlsx(document) if document.format == "xlsx" else _rows_from_csv(document)
        for sheet, row_number, row in iterator:
            recognized = True
            yield document.name, sheet, row_number, row
    if not recognized:
        raise HistoricalImportError(
            "No recognized historical-data sheet was found. Use the Tia Import Template v1."
        )


def _parse_uuid(value: Any) -> UUID | None:
    text = _clean(value)
    if not text:
        return None
    try:
        return UUID(text)
    except ValueError:
        return None


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _parse_time(value: Any) -> time | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.time().replace(tzinfo=None)
    if isinstance(value, time):
        return value.replace(tzinfo=None)
    if isinstance(value, (int, float)):
        fraction = float(value) % 1
        seconds = round(fraction * 24 * 3600)
        return time(hour=(seconds // 3600) % 24, minute=(seconds % 3600) // 60, second=seconds % 60)
    text = (_clean(value) or "").strip()
    if not text:
        return None
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M %p", "%I:%M%p"):
        try:
            return datetime.strptime(text.upper(), fmt).time()
        except ValueError:
            pass
    return None


def _parse_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, time.min)
    else:
        text = _clean(value)
        if not text:
            return None
        try:
            dt = datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            parsed_date = _parse_date(text)
            if parsed_date is None:
                return None
            dt = datetime.combine(parsed_date, time.min)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=EGYPT_TZ)
    return dt.astimezone(UTC)


def _money_minor(value: Any, *, allow_negative: bool = True) -> int | None:
    text = _clean(value)
    if text is None:
        return None
    cleaned = text.replace(",", "").replace("ج.م", "").replace("EGP", "").strip()
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None
    if not allow_negative and amount < 0:
        return None
    return int((amount * Decimal("100")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _int_value(value: Any) -> int | None:
    text = _clean(value)
    if text is None:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if number != number.to_integral_value():
        return None
    return int(number)


def _name_parts(full_name: str | None) -> tuple[str, str | None]:
    text = " ".join((full_name or "عميل مستورد").split()) or "عميل مستورد"
    parts = text.split(" ", 1)
    return parts[0][:120], parts[1][:120] if len(parts) > 1 else None


def _patient_identity(row: dict[str, Any]) -> tuple[str | None, str | None, str | None]:
    patient_id = _clean(row.get("patient_id"))
    phone_display, phone_normalized = normalize_patient_identity_phone(_clean(row.get("phone") or row.get("patient_phone")))
    if patient_id:
        return f"id:{patient_id}", phone_display, phone_normalized
    if phone_normalized:
        return f"phone:{_digest(phone_normalized)[:24]}", phone_display, phone_normalized
    return None, phone_display, phone_normalized


def _source_id(entity_type: str, explicit: str | None, fallback: str) -> str:
    if explicit:
        return f"{entity_type}:id:{explicit}"
    return f"{entity_type}:auto:{_digest(fallback)[:32]}"


def _service_catalog(db: Session, workspace_id: UUID) -> tuple[dict[UUID, Service], dict[str, Service]]:
    rows = list(db.scalars(select(Service).where(Service.workspace_id == workspace_id, Service.is_active.is_(True))))
    return {row.id: row for row in rows}, {row.name.strip().casefold(): row for row in rows}


def _resolve_service(row: dict[str, Any], by_id: dict[UUID, Service], by_name: dict[str, Service]) -> Service | None:
    service_uuid = _parse_uuid(row.get("service_id"))
    if service_uuid and service_uuid in by_id:
        return by_id[service_uuid]
    name = _clean(row.get("service_name"))
    if name:
        return by_name.get(name.casefold())
    return None


def _normalize_patient(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None, str | None]:
    identity, phone_display, phone_normalized = _patient_identity(row)
    if not identity:
        return None, "patient_identity_missing", "Patient requires patient_id or phone."
    birth = _parse_date(row.get("birth_date") or row.get("date_of_birth"))
    source = (_clean(row.get("source")) or "other").casefold()
    if source not in VALID_PATIENT_SOURCES:
        source = "other"
    payload = {
        "identity": identity,
        "patient_id": _clean(row.get("patient_id")),
        "full_name": _clean(row.get("full_name") or row.get("patient_name")),
        "phone": phone_display,
        "phone_normalized": phone_normalized,
        "birth_date": birth.isoformat() if birth else None,
        "source": source,
    }
    return payload, None, None


def _normalize_appointment(
    row: dict[str, Any],
    *,
    services_by_id: dict[UUID, Service],
    services_by_name: dict[str, Service],
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    identity, phone_display, phone_normalized = _patient_identity(row)
    if not identity:
        return None, "appointment_patient_identity_missing", "Appointment requires patient_id or patient_phone."
    service = _resolve_service(row, services_by_id, services_by_name)
    if service is None:
        return None, "appointment_service_unknown", "Appointment service must match a configured Tia service."
    day = _parse_date(row.get("date") or row.get("appointment_date"))
    start_time = _parse_time(row.get("start_time") or row.get("time"))
    if day is None:
        return None, "appointment_date_invalid", "Appointment date is missing or invalid."
    if start_time is None:
        return None, "appointment_time_invalid", "Appointment start_time is missing or invalid."
    status = (_clean(row.get("status")) or "").casefold().replace(" ", "_")
    if not status:
        local_start = datetime.combine(day, start_time, tzinfo=EGYPT_TZ)
        status = "completed" if local_start <= datetime.now(EGYPT_TZ) else "pending"
    if status not in VALID_APPOINTMENT_STATUSES:
        return None, "appointment_status_invalid", "Appointment status is not a supported Tia status."
    explicit = _clean(row.get("appointment_id"))
    fallback = "|".join([
        identity,
        str(service.id),
        day.isoformat(),
        start_time.isoformat(),
        _clean(row.get("doctor_id")) or _clean(row.get("doctor_name")) or "unassigned",
    ])
    payload = {
        "identity": identity,
        "patient_id": _clean(row.get("patient_id")),
        "patient_phone": phone_display,
        "patient_phone_normalized": phone_normalized,
        "patient_name": _clean(row.get("patient_name") or row.get("full_name")),
        "service_id": str(service.id),
        "service_name": service.name,
        "doctor_id": _clean(row.get("doctor_id")),
        "doctor_name": _clean(row.get("doctor_name")),
        "date": day.isoformat(),
        "start_time": start_time.isoformat(),
        "status": status,
        "package_id": _clean(row.get("package_id")),
        "appointment_id": explicit,
    }
    payload["source_record_id"] = _source_id("appointment", explicit, fallback)
    return payload, None, None


def _normalize_payment(row: dict[str, Any], *, source_file: str, sheet: str, row_number: int) -> tuple[dict[str, Any] | None, str | None, str | None]:
    identity, phone_display, phone_normalized = _patient_identity(row)
    if not identity:
        return None, "payment_patient_identity_missing", "Payment requires patient_id or patient_phone."
    amount_minor = _money_minor(row.get("amount"))
    if amount_minor is None or amount_minor == 0:
        return None, "payment_amount_invalid", "Payment amount must be a non-zero number; refunds use a negative amount."
    paid_at = _parse_datetime(row.get("paid_at") or row.get("date"))
    if paid_at is None:
        return None, "payment_date_invalid", "Payment paid_at is missing or invalid."
    method = (_clean(row.get("payment_method")) or "unknown").casefold().replace(" ", "_")
    if method not in VALID_PAYMENT_METHODS:
        method = "other"
    explicit = _clean(row.get("transaction_id"))
    fallback = f"{source_file}|{sheet}|{row_number}"
    payload = {
        "identity": identity,
        "patient_id": _clean(row.get("patient_id")),
        "patient_phone": phone_display,
        "patient_phone_normalized": phone_normalized,
        "transaction_id": explicit,
        "amount_minor": amount_minor,
        "paid_at": paid_at.isoformat(),
        "payment_method": method,
        "appointment_id": _clean(row.get("appointment_id")),
        "package_id": _clean(row.get("package_id")),
        "reference_transaction_id": _clean(row.get("reference_transaction_id")),
    }
    payload["source_record_id"] = _source_id("payment", explicit, fallback)
    return payload, None, None


def _normalize_allocation(row: dict[str, Any]) -> tuple[dict[str, Any] | None, str | None, str | None]:
    transaction_id = _clean(row.get("transaction_id"))
    appointment_id = _clean(row.get("appointment_id"))
    amount_minor = _money_minor(row.get("amount"), allow_negative=False)
    if not transaction_id or not appointment_id:
        return None, "allocation_reference_missing", "Payment allocation requires transaction_id and appointment_id."
    if amount_minor is None or amount_minor <= 0:
        return None, "allocation_amount_invalid", "Payment allocation amount must be positive."
    return {
        "transaction_id": transaction_id,
        "appointment_id": appointment_id,
        "amount_minor": amount_minor,
        "source_record_id": _source_id("payment_allocation", None, f"{transaction_id}|{appointment_id}"),
    }, None, None


def _normalize_package(
    row: dict[str, Any],
    *,
    services_by_id: dict[UUID, Service],
    services_by_name: dict[str, Service],
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    identity, phone_display, phone_normalized = _patient_identity(row)
    if not identity:
        return None, "package_patient_identity_missing", "Package requires patient_id or patient_phone."
    service = _resolve_service(row, services_by_id, services_by_name)
    if service is None:
        return None, "package_service_unknown", "Package service must match a configured Tia service."
    remaining = _int_value(row.get("sessions_remaining"))
    if remaining is None or remaining < 0:
        return None, "package_remaining_invalid", "Package sessions_remaining must be zero or greater."
    total = _int_value(row.get("sessions_total"))
    if total is not None and (total <= 0 or remaining > total):
        return None, "package_total_invalid", "sessions_total must be positive and not less than sessions_remaining."
    price_minor = _money_minor(row.get("price"), allow_negative=False)
    standalone_minor = _money_minor(row.get("standalone_session_price"), allow_negative=False)
    if _clean(row.get("price")) is not None and price_minor is None:
        return None, "package_price_invalid", "Package price must be a valid non-negative number."
    if _clean(row.get("standalone_session_price")) is not None and standalone_minor is None:
        return None, "package_standalone_price_invalid", "standalone_session_price must be a valid non-negative number."
    purchased_at = _parse_datetime(row.get("purchased_at")) or datetime.now(UTC)
    expires_at = _parse_date(row.get("expires_at"))
    status = (_clean(row.get("status")) or "active").casefold().replace(" ", "_")
    if status not in VALID_PACKAGE_STATUSES:
        return None, "package_status_invalid", "Package status is not supported."
    explicit = _clean(row.get("package_id"))
    fallback = f"{identity}|{service.id}|{_clean(row.get('package_name')) or service.name}|{purchased_at.date().isoformat()}"
    payload = {
        "identity": identity,
        "patient_id": _clean(row.get("patient_id")),
        "patient_phone": phone_display,
        "patient_phone_normalized": phone_normalized,
        "package_id": explicit,
        "package_name": _clean(row.get("package_name")) or f"{service.name} Package",
        "service_id": str(service.id),
        "service_name": service.name,
        "sessions_total": total,
        "sessions_remaining": remaining,
        "price_minor": price_minor or 0,
        "standalone_session_price_minor": standalone_minor,
        "purchased_at": purchased_at.isoformat(),
        "expires_at": expires_at.isoformat() if expires_at else None,
        "status": status,
    }
    payload["source_record_id"] = _source_id("package", explicit, fallback)
    return payload, None, None


def _normalize_row(
    entity_type: str,
    row: dict[str, Any],
    *,
    source_file: str,
    sheet: str,
    row_number: int,
    services_by_id: dict[UUID, Service],
    services_by_name: dict[str, Service],
) -> tuple[dict[str, Any] | None, str | None, str | None]:
    if entity_type == "patient":
        payload, code, message = _normalize_patient(row)
        if payload is not None:
            payload["source_record_id"] = _source_id("patient", payload.get("patient_id"), payload["identity"])
        return payload, code, message
    if entity_type == "appointment":
        return _normalize_appointment(row, services_by_id=services_by_id, services_by_name=services_by_name)
    if entity_type == "payment":
        return _normalize_payment(row, source_file=source_file, sheet=sheet, row_number=row_number)
    if entity_type == "payment_allocation":
        return _normalize_allocation(row)
    if entity_type == "package":
        return _normalize_package(row, services_by_id=services_by_id, services_by_name=services_by_name)
    raise HistoricalImportError(f"Unsupported entity type: {entity_type}")


def _batch_read(batch: ClinicHistoricalImportBatch) -> HistoricalImportBatchRead:
    return HistoricalImportBatchRead(
        batch_id=batch.id,
        mode=batch.mode,
        status=batch.status,
        schema_version=batch.schema_version,
        source_name=batch.source_name,
        summary=batch.summary_json or {},
        error_message=batch.error_message,
    )


def preview_historical_import(
    db: Session,
    *,
    workspace: Workspace,
    user_id: UUID,
    documents: list[HistoricalImportDocument],
    mode: str,
) -> HistoricalImportPreviewResponse:
    if workspace.primary_branch_id is None:
        raise HistoricalImportError("Complete Clinic Setup before importing historical data.")
    services_by_id, services_by_name = _service_catalog(db, workspace.id)
    if not services_by_id:
        raise HistoricalImportError("Add at least one service before importing historical data.")

    _validate_documents(documents)
    fingerprint = _source_fingerprint(documents)
    batch = ClinicHistoricalImportBatch(
        workspace_id=workspace.id,
        created_by_user_id=user_id,
        mode=mode,
        status="preview_ready",
        schema_version=SCHEMA_VERSION,
        source_name=", ".join(document.name for document in documents)[:255],
        source_fingerprint=fingerprint,
        summary_json={},
    )
    db.add(batch)
    db.flush()

    ready_counts: Counter[str] = Counter()
    rejected_counts: Counter[str] = Counter()
    issue_counts: dict[tuple[str, str, str], dict[str, Any]] = {}
    seen: dict[tuple[str, str], str] = {}
    rows_to_add: list[ClinicHistoricalImportRow] = []

    processed_rows = 0
    for source_file, sheet, row_number, raw_row in _iter_rows(documents):
        processed_rows += 1
        if processed_rows > MAX_IMPORT_ROWS:
            db.rollback()
            raise HistoricalImportError(f"Historical import exceeds the {MAX_IMPORT_ROWS:,}-row limit.")
        entity_type = RECOGNIZED_SHEETS[sheet]
        payload, issue_code, issue_message = _normalize_row(
            entity_type,
            raw_row,
            source_file=source_file,
            sheet=sheet,
            row_number=row_number,
            services_by_id=services_by_id,
            services_by_name=services_by_name,
        )
        if payload is None:
            source_record_id = f"{entity_type}:rejected:{_digest(f'{source_file}|{sheet}|{row_number}')[:32]}"
            payload = {"raw": {key: _clean(value) for key, value in raw_row.items()}}
            row_status = "rejected"
        else:
            source_record_id = str(payload.pop("source_record_id"))
            content_hash = _payload_hash(payload)
            key = (entity_type, source_record_id)
            if key in seen:
                issue_code = f"duplicate_{entity_type}_record"
                issue_message = f"Duplicate {entity_type} identity in the same import file."
                source_record_id = f"{source_record_id}:duplicate:{row_number}"
                row_status = "rejected"
            else:
                seen[key] = content_hash
                row_status = "ready"
        payload_hash = _payload_hash(payload)
        if row_status == "ready":
            ready_counts[entity_type] += 1
        else:
            rejected_counts[entity_type] += 1
            issue_key = (entity_type, issue_code or "invalid_row", issue_message or "Invalid row")
            group = issue_counts.setdefault(issue_key, {"count": 0, "rows": []})
            group["count"] += 1
            if len(group["rows"]) < 5:
                group["rows"].append(row_number)
        rows_to_add.append(
            ClinicHistoricalImportRow(
                workspace_id=workspace.id,
                batch_id=batch.id,
                entity_type=entity_type,
                source_sheet=sheet,
                row_number=row_number,
                source_record_id=source_record_id,
                payload_hash=payload_hash,
                row_status=row_status,
                normalized_json=payload,
                issue_code=issue_code,
                issue_message=issue_message,
            )
        )

    total_ready = sum(ready_counts.values())
    if total_ready == 0:
        db.rollback()
        raise HistoricalImportError("The uploaded files contain no usable historical records.")

    # Allocation references must point to explicit transaction/appointment IDs in the same
    # upload or to a previously imported source record. The contract does not guess joins.
    explicit_tx = {
        row.normalized_json.get("transaction_id"): int(row.normalized_json.get("amount_minor") or 0)
        for row in rows_to_add
        if row.entity_type == "payment" and row.row_status == "ready" and row.normalized_json.get("transaction_id")
    }
    explicit_appt = {
        row.normalized_json.get("appointment_id")
        for row in rows_to_add
        if row.entity_type == "appointment" and row.row_status == "ready" and row.normalized_json.get("appointment_id")
    }
    allocation_totals: Counter[str] = Counter()
    for staged in rows_to_add:
        if staged.entity_type != "payment_allocation" or staged.row_status != "ready":
            continue
        tx_id = staged.normalized_json.get("transaction_id")
        appt_id = staged.normalized_json.get("appointment_id")
        issue = None
        if tx_id not in explicit_tx or appt_id not in explicit_appt:
            issue = (
                "allocation_reference_unknown",
                "Allocation references must match explicit transaction_id and appointment_id rows in this upload.",
            )
        elif explicit_tx[tx_id] <= 0:
            issue = (
                "allocation_refund_not_supported",
                "Payment allocations can only allocate positive payment transactions, not refunds.",
            )
        else:
            allocation_totals[tx_id] += int(staged.normalized_json.get("amount_minor") or 0)
            if allocation_totals[tx_id] > explicit_tx[tx_id]:
                issue = (
                    "allocation_exceeds_payment",
                    "Payment allocations cannot exceed the referenced payment amount.",
                )
        if issue is None:
            continue
        staged.row_status = "rejected"
        staged.issue_code, staged.issue_message = issue
        ready_counts["payment_allocation"] -= 1
        rejected_counts["payment_allocation"] += 1
        issue_key = ("payment_allocation", staged.issue_code, staged.issue_message)
        group = issue_counts.setdefault(issue_key, {"count": 0, "rows": []})
        group["count"] += 1
        if len(group["rows"]) < 5:
            group["rows"].append(staged.row_number)

    db.add_all(rows_to_add)
    issue_groups = [
        HistoricalImportIssueGroup(
            entity_type=entity_type,
            code=code,
            message=message,
            occurrence_count=meta["count"],
            example_rows=meta["rows"],
        )
        for (entity_type, code, message), meta in sorted(issue_counts.items(), key=lambda item: (-item[1]["count"], item[0]))
    ]
    batch.summary_json = {
        "ready_counts": dict(ready_counts),
        "rejected_counts": dict(rejected_counts),
        "issues": [group.model_dump() for group in issue_groups],
        "source_fingerprint": fingerprint,
    }
    db.commit()
    db.refresh(batch)
    return HistoricalImportPreviewResponse(
        batch=_batch_read(batch),
        ready_counts=dict(ready_counts),
        rejected_counts=dict(rejected_counts),
        issue_groups=issue_groups,
        can_import=sum(ready_counts.values()) > 0,
    )


def get_historical_import_batch(db: Session, *, workspace_id: UUID, batch_id: UUID) -> ClinicHistoricalImportBatch | None:
    return db.scalar(
        select(ClinicHistoricalImportBatch).where(
            ClinicHistoricalImportBatch.workspace_id == workspace_id,
            ClinicHistoricalImportBatch.id == batch_id,
        )
    )


def list_historical_import_batches(db: Session, *, workspace_id: UUID, limit: int = 20) -> list[ClinicHistoricalImportBatch]:
    return list(
        db.scalars(
            select(ClinicHistoricalImportBatch)
            .where(ClinicHistoricalImportBatch.workspace_id == workspace_id)
            .order_by(ClinicHistoricalImportBatch.created_at.desc())
            .limit(limit)
        )
    )


def start_historical_import(db: Session, *, batch: ClinicHistoricalImportBatch) -> ClinicHistoricalImportBatch:
    if batch.status == "imported":
        return batch
    if batch.status == "importing":
        raise HistoricalImportConflictError("This historical import is already running.")
    if batch.status not in {"preview_ready", "failed"}:
        raise HistoricalImportConflictError("This historical import cannot be started from its current state.")
    batch.status = "importing"
    batch.error_message = None
    record_activity_event(
        db, workspace_id=batch.workspace_id, actor_type="staff", actor_user_id=batch.created_by_user_id,
        action="clinic.history_import_started", entity_type="historical_import", entity_id=batch.id,
        summary="Historical clinic import started.", metadata={"mode": batch.mode}, flush=False,
    )
    db.commit()
    db.refresh(batch)
    return batch


def mark_historical_import_failed(db: Session, *, batch_id: UUID, workspace_id: UUID, message: str) -> None:
    batch = get_historical_import_batch(db, workspace_id=workspace_id, batch_id=batch_id)
    if batch is None:
        return
    batch.status = "failed"
    batch.error_message = message[:1200]
    record_activity_event(
        db, workspace_id=batch.workspace_id, actor_type="system", actor_user_id=None,
        action="clinic.history_import_failed", entity_type="historical_import", entity_id=batch.id,
        summary="Historical clinic import failed.", metadata={"mode": batch.mode}, flush=False,
    )
    db.commit()


def _existing_link_map(db: Session, workspace_id: UUID) -> dict[tuple[str, str], ClinicHistoricalImportLink]:
    return {
        (row.entity_type, row.source_record_id): row
        for row in db.scalars(
            select(ClinicHistoricalImportLink).where(ClinicHistoricalImportLink.workspace_id == workspace_id)
        )
    }


def _assign_patient_phone_if_available(
    db: Session,
    *,
    workspace_id: UUID,
    patient: Patient,
    phone: str | None,
    phone_normalized: str | None,
) -> bool:
    """Assign a canonical patient phone only when it is not owned by another patient.

    Historical exports can disagree about which stable patient id owns a phone number.
    The database unique constraint is the final identity safety boundary, so Append/Replace
    must never steal a normalized phone from another canonical patient.
    """
    if not phone_normalized:
        return False
    owner = db.scalar(
        select(Patient).where(
            Patient.workspace_id == workspace_id,
            Patient.phone_normalized == phone_normalized,
        )
    )
    if owner is not None and owner.id != patient.id:
        return False
    if phone:
        patient.phone = phone
    patient.phone_normalized = phone_normalized
    return True


def _patient_for_identity(
    db: Session,
    *,
    workspace: Workspace,
    identity: str,
    payload: dict[str, Any],
    cache: dict[str, Patient],
) -> Patient:
    if identity in cache:
        patient = cache[identity]
        # Fill a missing phone only when it is not already owned by another canonical patient.
        if not patient.phone_normalized and payload.get("patient_phone_normalized"):
            _assign_patient_phone_if_available(
                db,
                workspace_id=workspace.id,
                patient=patient,
                phone=payload.get("patient_phone"),
                phone_normalized=payload.get("patient_phone_normalized"),
            )
        return patient
    phone_normalized = payload.get("phone_normalized") or payload.get("patient_phone_normalized")
    patient = None
    if phone_normalized:
        patient = db.scalar(
            select(Patient).where(
                Patient.workspace_id == workspace.id,
                Patient.phone_normalized == phone_normalized,
            )
        )
    if patient is None:
        full_name = payload.get("full_name") or payload.get("patient_name")
        first_name, last_name = _name_parts(full_name)
        birth_date = _parse_date(payload.get("birth_date"))
        patient = Patient(
            workspace_id=workspace.id,
            first_name=first_name,
            last_name=last_name,
            phone=payload.get("phone") or payload.get("patient_phone"),
            phone_normalized=phone_normalized,
            gender="female",
            birth_date=birth_date,
            preferred_language="ar",
            preferred_branch_id=workspace.primary_branch_id,
            source=payload.get("source") or "other",
            source_detail=None,
            status="active",
            marketing_consent=False,
        )
        db.add(patient)
        db.flush()
    cache[identity] = patient
    return patient


def _fill_missing_patient_facts(
    patient: Patient,
    payload: dict[str, Any],
    *,
    db: Session | None = None,
    workspace_id: UUID | None = None,
) -> None:
    """Merge non-destructive patient presentation facts for Append imports.

    Patient identity is established by patient_id/normalized phone, never by name.
    Append must not erase or silently replace existing canonical facts when a later
    clinic export is partial. It may only fill fields that are currently missing.
    When a database session is supplied, a phone is filled only if no other canonical
    patient in the workspace already owns that normalized number.
    """
    full_name = _clean(payload.get("full_name") or payload.get("patient_name"))
    if full_name and not (patient.first_name or patient.last_name):
        first_name, last_name = _name_parts(full_name)
        patient.first_name = first_name
        patient.last_name = last_name

    phone = payload.get("phone") or payload.get("patient_phone")
    phone_normalized = payload.get("phone_normalized") or payload.get("patient_phone_normalized")
    if phone_normalized and not patient.phone_normalized:
        if db is not None and workspace_id is not None:
            _assign_patient_phone_if_available(
                db,
                workspace_id=workspace_id,
                patient=patient,
                phone=phone,
                phone_normalized=phone_normalized,
            )
        else:
            if phone:
                patient.phone = phone
            patient.phone_normalized = phone_normalized
    elif phone and not patient.phone and (not phone_normalized or phone_normalized == patient.phone_normalized):
        patient.phone = phone

    birth_date = _parse_date(payload.get("birth_date"))
    if birth_date is not None and patient.birth_date is None:
        patient.birth_date = birth_date

    source = _clean(payload.get("source"))
    if source and source != "other" and (not patient.source or patient.source == "other"):
        patient.source = source


def _doctor_catalog(db: Session, workspace_id: UUID) -> tuple[dict[UUID, Doctor], dict[str, Doctor]]:
    rows = list(
        db.execute(
            select(Doctor, Staff)
            .join(Staff, (Staff.workspace_id == Doctor.workspace_id) & (Staff.id == Doctor.staff_id))
            .where(Doctor.workspace_id == workspace_id, Doctor.is_active.is_(True), Staff.is_active.is_(True))
        ).all()
    )
    return {doctor.id: doctor for doctor, _staff in rows}, {
        f"{staff.first_name} {staff.last_name}".strip().casefold(): doctor for doctor, staff in rows
    }


def _create_visiting_doctor(db: Session, *, workspace: Workspace, full_name: str) -> Doctor:
    if workspace.primary_branch_id is None:
        raise HistoricalImportError("Clinic profile is incomplete.")
    first, last = _name_parts(full_name)
    staff = Staff(
        workspace_id=workspace.id,
        user_id=None,
        first_name=first,
        last_name=last or "",
        email=None,
        phone=None,
        job_title="doctor",
        is_active=True,
    )
    db.add(staff)
    db.flush()
    doctor = Doctor(
        workspace_id=workspace.id,
        staff_id=staff.id,
        doctor_type="visiting",
        specialization=None,
        license_number=None,
        bio=None,
        booking_enabled=False,
        is_active=True,
    )
    db.add(doctor)
    db.flush()
    db.add(
        DoctorBranch(
            workspace_id=workspace.id,
            doctor_id=doctor.id,
            branch_id=workspace.primary_branch_id,
            is_primary=True,
            is_active=True,
        )
    )
    db.flush()
    return doctor


def _unassigned_historical_doctor(db: Session, *, workspace: Workspace, by_name: dict[str, Doctor]) -> Doctor:
    key = "غير محدد (تاريخي)".casefold()
    if key in by_name:
        return by_name[key]
    doctor = _create_visiting_doctor(db, workspace=workspace, full_name="غير محدد (تاريخي)")
    by_name[key] = doctor
    return doctor


def _resolve_or_create_doctor(
    db: Session,
    *,
    workspace: Workspace,
    payload: dict[str, Any],
    by_id: dict[UUID, Doctor],
    by_name: dict[str, Doctor],
) -> tuple[Doctor, bool]:
    doctor_uuid = _parse_uuid(payload.get("doctor_id"))
    if doctor_uuid and doctor_uuid in by_id:
        return by_id[doctor_uuid], True
    name = _clean(payload.get("doctor_name"))
    if name:
        key = name.casefold()
        if key in by_name:
            return by_name[key], True
        doctor = _create_visiting_doctor(db, workspace=workspace, full_name=name)
        by_id[doctor.id] = doctor
        by_name[key] = doctor
        return doctor, True
    return _unassigned_historical_doctor(db, workspace=workspace, by_name=by_name), False


def _safe_remove_previous_imports(
    db: Session,
    *,
    workspace_id: UUID,
    incoming_hashes: dict[tuple[str, str], str],
) -> dict[str, int]:
    links = list(db.scalars(select(ClinicHistoricalImportLink).where(ClinicHistoricalImportLink.workspace_id == workspace_id)))
    # Replace means synchronize the previously imported history with this batch.
    # Missing source records and changed source records are removed; identical records
    # remain idempotent and are reused.
    stale = [
        row for row in links
        if row.entity_type != "patient"
        and (
            (row.entity_type, row.source_record_id) not in incoming_hashes
            or incoming_hashes[(row.entity_type, row.source_record_id)] != row.payload_hash
        )
    ]
    ids_by_type: dict[str, set[UUID]] = defaultdict(set)
    for link in stale:
        ids_by_type[link.entity_type].add(link.canonical_id)

    appointment_ids = ids_by_type.get("appointment", set())
    payment_ids = ids_by_type.get("payment", set())
    package_ids = ids_by_type.get("package", set())

    if appointment_ids:
        campaign_ref = db.scalar(
            select(CRMCampaignConversion.id).where(
                CRMCampaignConversion.workspace_id == workspace_id,
                (CRMCampaignConversion.appointment_id.in_(appointment_ids))
                | (CRMCampaignConversion.original_appointment_id.in_(appointment_ids)),
            ).limit(1)
        )
        if campaign_ref is not None:
            raise HistoricalImportConflictError(
                "Some previously imported appointments now have campaign attribution. Use Append or keep those records."
            )
        runtime_payment_ref = db.scalar(
            select(PaymentTransaction.id).where(
                PaymentTransaction.workspace_id == workspace_id,
                PaymentTransaction.id.not_in(payment_ids or {UUID(int=0)}),
                (PaymentTransaction.appointment_id.in_(appointment_ids))
                | (PaymentTransaction.origin_appointment_id.in_(appointment_ids)),
            ).limit(1)
        )
        if runtime_payment_ref is not None:
            raise HistoricalImportConflictError(
                "Some previously imported appointments have newer Tia payment activity. Use Append instead of Replace."
            )

    if package_ids:
        usage_ref = db.scalar(
            select(PackageUsage.id).where(
                PackageUsage.workspace_id == workspace_id,
                PackageUsage.patient_package_id.in_(package_ids),
            ).limit(1)
        )
        if usage_ref is not None:
            raise HistoricalImportConflictError(
                "Some imported packages were used after migration. Use Append so current package balances are preserved."
            )
        runtime_package_payment = db.scalar(
            select(PaymentTransaction.id).where(
                PaymentTransaction.workspace_id == workspace_id,
                PaymentTransaction.id.not_in(payment_ids or {UUID(int=0)}),
                PaymentTransaction.patient_package_id.in_(package_ids),
            ).limit(1)
        )
        if runtime_package_payment is not None:
            raise HistoricalImportConflictError(
                "Some imported packages have newer Tia payment activity. Use Append instead of Replace."
            )

    if payment_ids:
        referenced = db.scalar(
            select(PaymentTransaction.id).where(
                PaymentTransaction.workspace_id == workspace_id,
                PaymentTransaction.id.not_in(payment_ids),
                PaymentTransaction.reference_transaction_id.in_(payment_ids),
            ).limit(1)
        )
        if referenced is not None:
            raise HistoricalImportConflictError(
                "A newer Tia refund references an imported payment. Use Append instead of Replace."
            )

    # Preserve agent audit history. Only clear the nullable appointment pointer.
    if appointment_ids:
        db.execute(
            update(AgentAction)
            .where(AgentAction.workspace_id == workspace_id, AgentAction.appointment_id.in_(appointment_ids))
            .values(appointment_id=None)
        )
        # Appointment-specific queued/derived automation jobs cannot remain valid after the
        # imported appointment is removed. They are scoped to these historical appointments only.
        db.execute(
            delete(AutomationJob).where(
                AutomationJob.workspace_id == workspace_id,
                AutomationJob.appointment_id.in_(appointment_ids),
            )
        )

    if payment_ids or appointment_ids:
        alloc_filter = PaymentAllocation.workspace_id == workspace_id
        clauses = []
        if payment_ids:
            clauses.append(PaymentAllocation.transaction_id.in_(payment_ids))
        if appointment_ids:
            clauses.append(PaymentAllocation.appointment_id.in_(appointment_ids))
        if clauses:
            from sqlalchemy import or_
            db.execute(delete(PaymentAllocation).where(alloc_filter, or_(*clauses)))

    if appointment_ids:
        db.execute(delete(PackageUsage).where(PackageUsage.workspace_id == workspace_id, PackageUsage.appointment_id.in_(appointment_ids)))
    if payment_ids:
        # Clear self-references and package purchase pointers before deleting the imported
        # financial facts; this keeps FK ordering deterministic.
        db.execute(
            update(PaymentTransaction)
            .where(PaymentTransaction.workspace_id == workspace_id, PaymentTransaction.id.in_(payment_ids))
            .values(reference_transaction_id=None)
        )
        db.execute(
            update(PatientPackage)
            .where(PatientPackage.workspace_id == workspace_id, PatientPackage.purchase_transaction_id.in_(payment_ids))
            .values(purchase_transaction_id=None)
        )
        db.execute(delete(PaymentTransaction).where(PaymentTransaction.workspace_id == workspace_id, PaymentTransaction.id.in_(payment_ids)))
    if appointment_ids:
        db.execute(delete(Appointment).where(Appointment.workspace_id == workspace_id, Appointment.id.in_(appointment_ids)))
    if package_ids:
        db.execute(delete(PatientPackage).where(PatientPackage.workspace_id == workspace_id, PatientPackage.id.in_(package_ids)))

    # Patients deliberately remain: patient identity may have conversations/messages/runtime data.
    stale_ids = [row.id for row in stale]
    if stale_ids:
        db.execute(delete(ClinicHistoricalImportLink).where(ClinicHistoricalImportLink.id.in_(stale_ids)))
    return {entity: len(ids) for entity, ids in ids_by_type.items()}


def apply_historical_import(
    db: Session,
    *,
    workspace: Workspace,
    batch: ClinicHistoricalImportBatch,
) -> dict[str, Any]:
    if batch.status not in {"importing", "preview_ready", "failed"}:
        if batch.status == "imported":
            return batch.summary_json or {}
        raise HistoricalImportConflictError("Historical import is not ready to apply.")
    rows = list(
        db.scalars(
            select(ClinicHistoricalImportRow)
            .where(
                ClinicHistoricalImportRow.workspace_id == workspace.id,
                ClinicHistoricalImportRow.batch_id == batch.id,
                ClinicHistoricalImportRow.row_status == "ready",
            )
            .order_by(ClinicHistoricalImportRow.entity_type, ClinicHistoricalImportRow.row_number)
        )
    )
    if not rows:
        raise HistoricalImportError("Historical import contains no ready rows.")

    existing_links = _existing_link_map(db, workspace.id)
    incoming_hashes = {
        (row.entity_type, row.source_record_id): row.payload_hash
        for row in rows if row.entity_type != "payment_allocation"
    }
    if batch.mode == "append":
        for row in rows:
            if row.entity_type in {"payment_allocation", "patient"}:
                # Patients are identity anchors rather than immutable historical facts.
                # A later export may repeat the same patient with partial presentation
                # fields (or a changed CRM source). Reuse the canonical patient instead
                # of forcing Replace. Appointment/payment/package facts remain strict.
                continue
            link = existing_links.get((row.entity_type, row.source_record_id))
            if link and link.payload_hash != row.payload_hash:
                raise HistoricalImportConflictError(
                    f"{row.entity_type} {row.source_record_id} was imported before with different data. Choose Replace previous imports."
                )
    else:
        _safe_remove_previous_imports(db, workspace_id=workspace.id, incoming_hashes=incoming_hashes)
        existing_links = _existing_link_map(db, workspace.id)

    services_by_id, _services_by_name = _service_catalog(db, workspace.id)
    doctors_by_id, doctors_by_name = _doctor_catalog(db, workspace.id)
    patient_cache: dict[str, Patient] = {}
    imported: Counter[str] = Counter()
    skipped: Counter[str] = Counter()

    # Seed patient facts first, then implicit identities from all other sheets. Patient
    # canonical rows deliberately survive Replace because conversations/runtime CRM may
    # already reference them. A stable historical link is reused instead of creating a
    # second patient when a later export changes presentation fields.
    for row in rows:
        payload = row.normalized_json
        identity = payload.get("identity")
        if not identity:
            continue
        patient_source_id = row.source_record_id if row.entity_type == "patient" else _source_id("patient", payload.get("patient_id"), identity)
        patient_hash = row.payload_hash if row.entity_type == "patient" else _payload_hash({"identity": identity})
        key = ("patient", patient_source_id)
        link = existing_links.get(key)
        patient = db.get(Patient, link.canonical_id) if link is not None else None
        if patient is None:
            patient = _patient_for_identity(db, workspace=workspace, identity=identity, payload=payload, cache=patient_cache)
        else:
            patient_cache[identity] = patient

        if row.entity_type == "patient" and batch.mode == "append":
            _fill_missing_patient_facts(
                patient,
                payload,
                db=db,
                workspace_id=workspace.id,
            )

        if row.entity_type == "patient" and batch.mode == "replace_previous_imports":
            first_name, last_name = _name_parts(payload.get("full_name"))
            if payload.get("full_name"):
                patient.first_name = first_name
                patient.last_name = last_name
            if payload.get("phone_normalized"):
                _assign_patient_phone_if_available(
                    db,
                    workspace_id=workspace.id,
                    patient=patient,
                    phone=payload.get("phone"),
                    phone_normalized=payload.get("phone_normalized"),
                )
            if payload.get("birth_date"):
                patient.birth_date = _parse_date(payload.get("birth_date"))
            if payload.get("source"):
                patient.source = payload.get("source")

        if link is None:
            link = ClinicHistoricalImportLink(
                workspace_id=workspace.id,
                batch_id=batch.id,
                entity_type="patient",
                canonical_id=patient.id,
                source_record_id=patient_source_id,
                payload_hash=patient_hash,
            )
            db.add(link)
            db.flush()
            existing_links[key] = link
        elif batch.mode == "replace_previous_imports" and row.entity_type == "patient":
            link.payload_hash = patient_hash
            link.batch_id = batch.id

    package_by_external: dict[str, PatientPackage] = {}
    for row in [item for item in rows if item.entity_type == "package"]:
        link = existing_links.get(("package", row.source_record_id))
        if link:
            skipped["package"] += 1
            package = db.get(PatientPackage, link.canonical_id) if hasattr(link, "canonical_id") else None
            if package and row.normalized_json.get("package_id"):
                package_by_external[row.normalized_json["package_id"]] = package
            continue
        payload = row.normalized_json
        patient = patient_cache[payload["identity"]]
        service = services_by_id[UUID(payload["service_id"])]
        total_known = payload.get("sessions_total") is not None
        total = int(payload.get("sessions_total") or max(int(payload["sessions_remaining"]), 1))
        package = PatientPackage(
            workspace_id=workspace.id,
            patient_id=patient.id,
            service_id=service.id,
            purchase_transaction_id=None,
            created_by_user_id=batch.created_by_user_id,
            external_id=payload.get("package_id"),
            name=payload["package_name"],
            sessions_purchased=total,
            opening_sessions_remaining=int(payload["sessions_remaining"]),
            sessions_total_known=total_known,
            sale_price_minor=int(payload.get("price_minor") or 0),
            standalone_session_price_minor_at_purchase=payload.get("standalone_session_price_minor"),
            currency="EGP",
            purchased_at=_parse_datetime(payload["purchased_at"]) or datetime.now(UTC),
            expires_at=_parse_date(payload.get("expires_at")),
            status=payload.get("status") or "active",
            source="integration",
            idempotency_key=f"historical:{row.source_record_id}"[:128],
        )
        db.add(package)
        db.flush()
        db.add(ClinicHistoricalImportLink(
            workspace_id=workspace.id,
            batch_id=batch.id,
            entity_type="package",
            canonical_id=package.id,
            source_record_id=row.source_record_id,
            payload_hash=row.payload_hash,
        ))
        if payload.get("package_id"):
            package_by_external[payload["package_id"]] = package
        imported["package"] += 1

    appointment_by_external: dict[str, Appointment] = {}
    for row in [item for item in rows if item.entity_type == "appointment"]:
        link = existing_links.get(("appointment", row.source_record_id))
        if link:
            skipped["appointment"] += 1
            appointment = db.get(Appointment, link.canonical_id) if hasattr(link, "canonical_id") else None
            if appointment and row.normalized_json.get("appointment_id"):
                appointment_by_external[row.normalized_json["appointment_id"]] = appointment
            continue
        payload = row.normalized_json
        patient = patient_cache[payload["identity"]]
        service = services_by_id[UUID(payload["service_id"])]
        doctor, assignment_known = _resolve_or_create_doctor(
            db, workspace=workspace, payload=payload, by_id=doctors_by_id, by_name=doctors_by_name
        )
        # Give newly discovered visiting doctors the historical service association for semantic context.
        assignment = db.scalar(select(DoctorService.id).where(
            DoctorService.workspace_id == workspace.id,
            DoctorService.doctor_id == doctor.id,
            DoctorService.service_id == service.id,
        ))
        if assignment is None:
            db.add(DoctorService(
                workspace_id=workspace.id,
                doctor_id=doctor.id,
                service_id=service.id,
                custom_duration_minutes=None,
                custom_price_minor=None,
                is_active=True,
            ))
        local_start = datetime.combine(date.fromisoformat(payload["date"]), time.fromisoformat(payload["start_time"]), tzinfo=EGYPT_TZ)
        start_at = local_start.astimezone(UTC)
        end_at = start_at + timedelta(minutes=service.duration_minutes)
        package = package_by_external.get(payload.get("package_id") or "")
        status = payload["status"]
        appointment = Appointment(
            workspace_id=workspace.id,
            patient_id=patient.id,
            branch_id=workspace.primary_branch_id,
            doctor_id=doctor.id,
            doctor_assignment_known=assignment_known,
            service_id=service.id,
            patient_package_id=package.id if package else None,
            lead_id=None,
            created_by_user_id=batch.created_by_user_id,
            rescheduled_from_appointment_id=None,
            status=status,
            source="other",
            start_at=start_at,
            end_at=end_at,
            busy_start_at=start_at,
            busy_end_at=end_at,
            duration_minutes=service.duration_minutes,
            price_minor=service.price_minor,
            currency="EGP",
            payment_status="unknown",
            amount_paid_minor=None,
            payment_method="unknown",
            billing_context="package_prepaid" if package else "standard",
            package_external_id=payload.get("package_id"),
            customer_note=None,
            cancellation_reason=None,
            idempotency_key=f"historical:{row.source_record_id}"[:128],
            completed_at=start_at if status == "completed" else None,
            cancelled_at=start_at if status == "cancelled" else None,
            no_show_at=start_at if status == "no_show" else None,
        )
        db.add(appointment)
        db.flush()
        db.add(ClinicHistoricalImportLink(
            workspace_id=workspace.id,
            batch_id=batch.id,
            entity_type="appointment",
            canonical_id=appointment.id,
            source_record_id=row.source_record_id,
            payload_hash=row.payload_hash,
        ))
        if payload.get("appointment_id"):
            appointment_by_external[payload["appointment_id"]] = appointment
        imported["appointment"] += 1

    payment_by_external: dict[str, PaymentTransaction] = {}
    pending_refs: list[tuple[PaymentTransaction, str]] = []
    for row in [item for item in rows if item.entity_type == "payment"]:
        link = existing_links.get(("payment", row.source_record_id))
        if link:
            skipped["payment"] += 1
            transaction = db.get(PaymentTransaction, link.canonical_id) if hasattr(link, "canonical_id") else None
            if transaction and row.normalized_json.get("transaction_id"):
                payment_by_external[row.normalized_json["transaction_id"]] = transaction
            continue
        payload = row.normalized_json
        patient = patient_cache[payload["identity"]]
        signed_amount = int(payload["amount_minor"])
        transaction = PaymentTransaction(
            workspace_id=workspace.id,
            appointment_id=(appointment_by_external.get(payload.get("appointment_id") or "") or None).id if appointment_by_external.get(payload.get("appointment_id") or "") else None,
            origin_appointment_id=None,
            patient_id=patient.id,
            created_by_user_id=batch.created_by_user_id,
            reference_transaction_id=None,
            patient_package_id=(package_by_external.get(payload.get("package_id") or "") or None).id if package_by_external.get(payload.get("package_id") or "") else None,
            transaction_type="refund" if signed_amount < 0 else "payment",
            amount_minor=abs(signed_amount),
            currency="EGP",
            payment_method=payload.get("payment_method") or "unknown",
            source="integration",
            external_reference=payload.get("transaction_id"),
            reason="Historical import" if signed_amount < 0 else None,
            idempotency_key=f"historical:{row.source_record_id}"[:128],
            created_at=_parse_datetime(payload.get("paid_at")) or datetime.now(UTC),
        )
        db.add(transaction)
        db.flush()
        db.add(ClinicHistoricalImportLink(
            workspace_id=workspace.id,
            batch_id=batch.id,
            entity_type="payment",
            canonical_id=transaction.id,
            source_record_id=row.source_record_id,
            payload_hash=row.payload_hash,
        ))
        if payload.get("transaction_id"):
            payment_by_external[payload["transaction_id"]] = transaction
        if payload.get("reference_transaction_id"):
            pending_refs.append((transaction, payload["reference_transaction_id"]))
        imported["payment"] += 1

    for transaction, external_reference in pending_refs:
        referenced = payment_by_external.get(external_reference)
        if referenced is not None and referenced.patient_id == transaction.patient_id:
            transaction.reference_transaction_id = referenced.id

    allocations_by_transaction: Counter[UUID] = Counter()
    for row in [item for item in rows if item.entity_type == "payment_allocation"]:
        payload = row.normalized_json
        transaction = payment_by_external.get(payload["transaction_id"])
        appointment = appointment_by_external.get(payload["appointment_id"])
        if transaction is None or appointment is None:
            skipped["payment_allocation"] += 1
            continue
        amount_minor = int(payload["amount_minor"])
        allocations_by_transaction[transaction.id] += amount_minor
        if allocations_by_transaction[transaction.id] > transaction.amount_minor:
            raise HistoricalImportError(
                f"Payment allocations exceed transaction amount for {payload['transaction_id']}."
            )
        existing = db.scalar(select(PaymentAllocation.id).where(
            PaymentAllocation.workspace_id == workspace.id,
            PaymentAllocation.transaction_id == transaction.id,
            PaymentAllocation.appointment_id == appointment.id,
        ))
        if existing is None:
            db.add(PaymentAllocation(
                workspace_id=workspace.id,
                transaction_id=transaction.id,
                appointment_id=appointment.id,
                amount_minor=amount_minor,
            ))
            imported["payment_allocation"] += 1
        else:
            skipped["payment_allocation"] += 1

    # Count explicit patient sheet facts separately; implicit patients are still retained.
    imported["patient"] = len({patient.id for patient in patient_cache.values()})
    db.flush()
    batch.status = "imported"
    batch.completed_at = datetime.now(UTC)
    summary = dict(batch.summary_json or {})
    summary.update({
        "imported_counts": dict(imported),
        "skipped_counts": dict(skipped),
        "mode": batch.mode,
    })
    batch.summary_json = summary
    batch.error_message = None
    record_activity_event(
        db, workspace_id=workspace.id, actor_type="system", actor_user_id=batch.created_by_user_id,
        action="clinic.history_imported", entity_type="historical_import", entity_id=batch.id,
        summary="Historical clinic import completed.",
        metadata={"mode": batch.mode, "imported_counts": dict(imported), "skipped_counts": dict(skipped)},
        flush=False,
    )
    db.commit()
    return summary


def build_historical_import_template() -> bytes:
    """Return the stable Tia History v1 workbook contract.

    The workbook intentionally mirrors clinic-facing facts rather than Tia's
    internal database schema. Currency is always EGP, patient email is not part
    of the contract, appointment end time is derived from the configured service
    duration, every patient is treated as female without a gender import column,
    and refunds are represented by a negative payment amount.
    """
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "README"
    instructions.append(["Tia Historical Import v1"])
    instructions.append(["All sheets are optional. Keep only the sheets you have data for."])
    instructions.append(["Patients are identified by patient_id or phone. If patient_id is missing, Tia creates a stable identity from the normalized phone number."])
    instructions.append(["Appointments need patient identity, service identity, date and start_time. end_at is calculated from the service duration configured in Tia."])
    instructions.append(["Payments use EGP. Positive amount = payment; negative amount = refund. transaction_id is optional."])
    instructions.append(["payment_allocations is optional and only used when both transaction_id and appointment_id exist explicitly."])
    instructions.append(["Active packages can be migrated using sessions_remaining. sessions_total may be left blank if the old system does not know it."])
    instructions.append(["Before import, Tia shows a preview and asks whether to append or replace previous historical-import records. Tia-created runtime data is never deleted by this flow."])

    sheets: dict[str, list[str]] = {
        "patients": [
            "patient_id", "full_name", "phone", "date_of_birth", "source"
        ],
        "appointments": [
            "appointment_id", "patient_id", "patient_phone", "patient_name",
            "service_id", "service_name", "doctor_id", "doctor_name", "date", "start_time",
            "status", "package_id",
        ],
        "payments": [
            "transaction_id", "patient_id", "patient_phone", "amount", "paid_at",
            "payment_method", "appointment_id", "package_id", "reference_transaction_id",
        ],
        "payment_allocations": [
            "transaction_id", "appointment_id", "amount",
        ],
        "packages": [
            "package_id", "patient_id", "patient_phone", "package_name",
            "service_id", "service_name", "sessions_total", "sessions_remaining",
            "price", "standalone_session_price", "purchased_at", "expires_at", "status",
        ],
    }
    for name, headers in sheets.items():
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = max(14, min(28, len(header) + 4))

    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


__all__ = [
    "HistoricalImportConflictError",
    "HistoricalImportError",
    "apply_historical_import",
    "build_historical_import_template",
    "get_historical_import_batch",
    "list_historical_import_batches",
    "mark_historical_import_failed",
    "preview_historical_import",
    "start_historical_import",
    "_batch_read",
]
