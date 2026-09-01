from __future__ import annotations

import base64
import io
import re
from collections import Counter, defaultdict
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from typing import Any
from uuid import UUID
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.doctor import Doctor
from app.models.doctor_service import DoctorService
from app.models.service import Service
from app.models.staff import Staff
from app.models.workspace import Workspace
from app.schemas.clinic_setup_v2 import (
    BookingPolicyUpdateV2,
    ClinicDoctorCreateV2,
    ClinicDoctorUpdateV2,
    ClinicProfileUpsert,
    ClinicServiceCreateV2,
    ClinicServiceUpdateV2,
    ClinicSetupDraft,
    ClinicSetupImportIssue,
    ClinicSetupImportResponse,
    ClinicSetupPreviewResponse,
    VisitingWindowInputV2,
    VisitingWindowsUpdateV2,
    WorkingHourInputV2,
    WorkingHoursUpdateV2,
)
from app.services.clinic_setup_v2 import (
    ClinicSetupV2Error,
    build_setup_v2_snapshot,
    create_doctor_v2,
    create_service_v2,
    replace_clinic_hours_v2,
    replace_doctor_services_v2,
    replace_regular_doctor_hours_v2,
    replace_visiting_windows_v2,
    update_booking_policy_v2,
    update_doctor_v2,
    update_service_v2,
    upsert_clinic_profile,
)

MAX_SETUP_WORKBOOK_BYTES = 10 * 1024 * 1024
SETUP_SHEETS = (
    "clinic_profile",
    "services",
    "doctors",
    "doctor_services",
    "clinic_hours",
    "doctor_hours",
    "visiting_windows",
    "booking_policy",
)


class ClinicSetupImportError(ValueError):
    pass


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (date, datetime, time)):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _key(value: Any) -> str:
    text = (_clean(value) or "").casefold().replace("-", "_").replace(" ", "_")
    return re.sub(r"_+", "_", text).strip("_")


def _decode(content_base64: str) -> bytes:
    try:
        raw = base64.b64decode(content_base64, validate=True)
    except Exception as exc:
        raise ClinicSetupImportError("ملف إعدادات العيادة غير صالح.") from exc
    if len(raw) > MAX_SETUP_WORKBOOK_BYTES:
        raise ClinicSetupImportError("ملف إعدادات العيادة أكبر من 10MB.")
    return raw


def _rows(sheet) -> list[tuple[int, dict[str, Any]]]:
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [_key(value) for value in next(iterator)]
    except StopIteration:
        return []
    result: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(iterator, start=2):
        if not any(_clean(value) for value in values):
            continue
        row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        result.append((row_number, row))
    return result


def _vertical(sheet, *, key_column: str, value_column: str) -> dict[str, tuple[int, Any]]:
    rows = _rows(sheet)
    result: dict[str, tuple[int, Any]] = {}
    for row_number, row in rows:
        raw_key = row.get(key_column)
        if raw_key is None:
            # Allow human-friendly first-column labels from the template.
            values = list(row.values())
            if not values:
                continue
            raw_key = values[0]
            value = values[1] if len(values) > 1 else None
        else:
            value = row.get(value_column)
        key = _key(raw_key)
        if key:
            result[key] = (row_number, value)
    return result


def _parse_decimal(value: Any) -> Decimal | None:
    text = _clean(value)
    if text is None:
        return None
    try:
        return Decimal(text.replace(",", ""))
    except (InvalidOperation, ValueError):
        return None


def _parse_int(value: Any) -> int | None:
    text = _clean(value)
    if text is None:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if number != number.to_integral_value():
        return None
    return int(number)


def _parse_bool(value: Any) -> bool | None:
    text = (_clean(value) or "").casefold()
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off"}:
        return False
    return None


def _parse_time(value: Any) -> time | None:
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    text = _clean(value)
    if text is None:
        return None
    try:
        return time.fromisoformat(text).replace(second=0, microsecond=0)
    except ValueError:
        return None


def _parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    if text is None:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


_WEEKDAY_NAMES = {
    "monday": 0, "mon": 0, "الاثنين": 0, "الإثنين": 0,
    "tuesday": 1, "tue": 1, "tues": 1, "الثلاثاء": 1,
    "wednesday": 2, "wed": 2, "الأربعاء": 2, "الاربعاء": 2,
    "thursday": 3, "thu": 3, "thur": 3, "thurs": 3, "الخميس": 3,
    "friday": 4, "fri": 4, "الجمعة": 4,
    "saturday": 5, "sat": 5, "السبت": 5,
    "sunday": 6, "sun": 6, "الأحد": 6, "الاحد": 6,
}


def _parse_weekday(row: dict[str, Any]) -> int | None:
    numeric = _parse_int(row.get("weekday"))
    if numeric is not None and 0 <= numeric <= 6:
        return numeric
    raw_day = _clean(row.get("day"))
    if raw_day is None:
        return None
    day_number = _parse_int(raw_day)
    if day_number is not None and 0 <= day_number <= 6:
        return day_number
    return _WEEKDAY_NAMES.get(raw_day.strip().casefold())


def _doctor_rows(db: Session, workspace_id: UUID) -> dict[str, tuple[Doctor, Staff]]:
    rows = db.execute(
        select(Doctor, Staff)
        .join(Staff, (Staff.workspace_id == Doctor.workspace_id) & (Staff.id == Doctor.staff_id))
        .where(Doctor.workspace_id == workspace_id, Doctor.is_active.is_(True), Staff.is_active.is_(True))
    ).all()
    return {f"{staff.first_name} {staff.last_name}".strip().casefold(): (doctor, staff) for doctor, staff in rows}


def _services(db: Session, workspace_id: UUID) -> dict[str, Service]:
    return {
        row.name.strip().casefold(): row
        for row in db.scalars(
            select(Service).where(Service.workspace_id == workspace_id, Service.is_active.is_(True))
        )
    }


def _current_doctor_service_ids(db: Session, workspace_id: UUID, doctor_id: UUID) -> list[UUID]:
    return list(
        db.scalars(
            select(DoctorService.service_id).where(
                DoctorService.workspace_id == workspace_id,
                DoctorService.doctor_id == doctor_id,
                DoctorService.is_active.is_(True),
            )
        )
    )


def _issue(issues: list[ClinicSetupImportIssue], *, sheet: str, row: int, message: str) -> None:
    issues.append(ClinicSetupImportIssue(sheet=sheet, row=row, message=message))




def _draft_value(value: Any) -> str | int | float | bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0).isoformat(timespec="minutes")
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    return text or None


def _draft_rows(sheet) -> list[dict[str, str | int | float | bool | None]]:
    return [
        {key: _draft_value(value) for key, value in row.items()}
        for _row_number, row in _rows(sheet)
    ]


def preview_clinic_setup_workbook(
    *,
    filename: str,
    content_base64: str,
) -> ClinicSetupPreviewResponse:
    """Read an XLSX into an editable draft without writing clinic data.

    Missing cells stay missing/blank.  This is intentionally a staging step so
    the admin can review and complete the workbook in the UI before persistence.
    """
    if not filename.lower().endswith(".xlsx"):
        raise ClinicSetupImportError("استخدم ملف Excel بصيغة .xlsx لإعدادات العيادة.")
    raw = _decode(content_base64)
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ClinicSetupImportError("تعذر قراءة ملف Excel.") from exc

    try:
        sheets = {_key(sheet.title): sheet for sheet in workbook.worksheets}
        recognized = [name for name in SETUP_SHEETS if name in sheets]
        if not recognized:
            raise ClinicSetupImportError("الملف لا يحتوي على أي sheet من Tia Clinic Setup Template.")

        profile: dict[str, str | int | float | bool | None] = {}
        if "clinic_profile" in sheets:
            values = _vertical(sheets["clinic_profile"], key_column="field", value_column="value")
            aliases = {
                "clinic_name": "name", "name": "name",
                "phone": "phone", "clinic_phone": "phone",
                "address": "address", "city": "city",
            }
            for raw_key, (_row, value) in values.items():
                target = aliases.get(raw_key)
                if target:
                    profile[target] = _draft_value(value)

        policy: dict[str, str | int | float | bool | None] = {}
        if "booking_policy" in sheets:
            values = _vertical(sheets["booking_policy"], key_column="setting", value_column="value")
            for key, (_row, value) in values.items():
                policy[key] = _draft_value(value)

        draft = ClinicSetupDraft(
            clinic_profile=profile,
            services=_draft_rows(sheets["services"]) if "services" in sheets else [],
            doctors=_draft_rows(sheets["doctors"]) if "doctors" in sheets else [],
            doctor_services=_draft_rows(sheets["doctor_services"]) if "doctor_services" in sheets else [],
            clinic_hours=_draft_rows(sheets["clinic_hours"]) if "clinic_hours" in sheets else [],
            doctor_hours=_draft_rows(sheets["doctor_hours"]) if "doctor_hours" in sheets else [],
            visiting_windows=_draft_rows(sheets["visiting_windows"]) if "visiting_windows" in sheets else [],
            booking_policy=policy,
        )
        return ClinicSetupPreviewResponse(draft=draft, recognized_sheets=recognized, issues=[])
    finally:
        workbook.close()


def clinic_setup_draft_to_workbook_base64(draft: ClinicSetupDraft) -> str:
    """Serialize the edited UI draft back to the fixed setup workbook contract."""
    wb = Workbook()
    wb.remove(wb.active)

    profile = wb.create_sheet("clinic_profile")
    profile.append(["field", "value"])
    for key in ("name", "phone", "address", "city"):
        value = draft.clinic_profile.get(key)
        if value not in (None, ""):
            profile.append(["clinic_name" if key == "name" else key, value])

    table_sheets: dict[str, tuple[list[str], list[dict[str, Any]]]] = {
        "services": (["name", "category", "duration_minutes", "price"], draft.services),
        "doctors": (["full_name", "doctor_type", "specialization"], draft.doctors),
        "doctor_services": (["doctor_name", "service_name"], draft.doctor_services),
        "clinic_hours": (["day", "weekday", "start_time", "end_time"], draft.clinic_hours),
        "doctor_hours": (["doctor_name", "day", "weekday", "start_time", "end_time"], draft.doctor_hours),
        "visiting_windows": (["doctor_name", "date", "start_time", "end_time"], draft.visiting_windows),
    }
    for sheet_name, (headers, rows) in table_sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            if not any(value not in (None, "") for value in row.values()):
                continue
            ws.append([row.get(header) for header in headers])

    policy = wb.create_sheet("booking_policy")
    policy.append(["setting", "value"])
    for key, value in draft.booking_policy.items():
        if value not in (None, ""):
            policy.append([key, value])

    output = io.BytesIO()
    wb.save(output)
    return base64.b64encode(output.getvalue()).decode("ascii")


def import_clinic_setup_workbook(
    db: Session,
    *,
    workspace: Workspace,
    filename: str,
    content_base64: str,
) -> ClinicSetupImportResponse:
    if not filename.lower().endswith(".xlsx"):
        raise ClinicSetupImportError("استخدم ملف Excel بصيغة .xlsx لإعدادات العيادة.")
    raw = _decode(content_base64)
    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        raise ClinicSetupImportError("تعذر قراءة ملف Excel.") from exc

    issues: list[ClinicSetupImportIssue] = []
    imported = Counter({name: 0 for name in SETUP_SHEETS})
    skipped = Counter({name: 0 for name in SETUP_SHEETS})
    sheets = {_key(sheet.title): sheet for sheet in workbook.worksheets}
    recognized = set(sheets).intersection(SETUP_SHEETS)
    if not recognized:
        workbook.close()
        raise ClinicSetupImportError("الملف لا يحتوي على أي sheet من Tia Clinic Setup Template.")

    try:
        # 1) Clinic profile. This may create the single primary branch needed by doctor/hour imports.
        profile_sheet = sheets.get("clinic_profile")
        if profile_sheet is not None:
            values = _vertical(profile_sheet, key_column="field", value_column="value")
            aliases = {
                "clinic_name": "name", "name": "name",
                "phone": "phone", "clinic_phone": "phone",
                "address": "address", "city": "city",
            }
            normalized: dict[str, Any] = {}
            first_row = 2
            for raw_key, (row_number, value) in values.items():
                target = aliases.get(raw_key)
                if target:
                    normalized[target] = value
                    first_row = min(first_row, row_number)
            current_profile = build_setup_v2_snapshot(db, workspace=workspace).clinic
            name = _clean(normalized.get("name"))
            if name is None and current_profile.branch_id is not None:
                name = _clean(current_profile.name)
            if values and name:
                try:
                    upsert_clinic_profile(
                        db,
                        workspace=workspace,
                        payload=ClinicProfileUpsert(
                            name=name,
                            phone=_clean(normalized.get("phone")) if "phone" in normalized else current_profile.phone,
                            address=_clean(normalized.get("address")) if "address" in normalized else current_profile.address,
                            city=_clean(normalized.get("city")) if "city" in normalized else current_profile.city,
                        ),
                    )
                    imported["clinic_profile"] = 1
                except (ClinicSetupV2Error, ValidationError) as exc:
                    skipped["clinic_profile"] = 1
                    _issue(issues, sheet="clinic_profile", row=first_row, message=str(exc))
            elif values:
                skipped["clinic_profile"] = 1
                _issue(issues, sheet="clinic_profile", row=first_row, message="تعذر تحديد اسم للعيادة؛ كمّله يدويًا من بيانات العيادة.")

        # 2) Services: name identifies the row. Optional values preserve existing data.
        service_sheet = sheets.get("services")
        if service_sheet is not None:
            seen: set[str] = set()
            for row_number, row in _rows(service_sheet):
                name = _clean(row.get("name"))
                if not name:
                    skipped["services"] += 1
                    _issue(issues, sheet="services", row=row_number, message="اسم الخدمة مطلوب لتحديد الصف؛ باقي البيانات يمكن استكمالها يدويًا.")
                    continue
                name_key = name.casefold()
                if name_key in seen:
                    skipped["services"] += 1
                    _issue(issues, sheet="services", row=row_number, message="اسم الخدمة مكرر داخل الملف.")
                    continue
                seen.add(name_key)
                current = _services(db, workspace.id).get(name_key)

                raw_duration = _clean(row.get("duration_minutes"))
                duration = _parse_int(row.get("duration_minutes"))
                if raw_duration is not None and (duration is None or duration <= 0):
                    skipped["services"] += 1
                    _issue(issues, sheet="services", row=row_number, message="duration_minutes غير صالح؛ اتركه فارغًا أو اكتب عدد دقائق صحيح.")
                    continue
                if duration is None:
                    duration = current.duration_minutes if current is not None else None
                if duration is None:
                    skipped["services"] += 1
                    _issue(issues, sheet="services", row=row_number, message="الخدمة الجديدة تحتاج duration_minutes لأن Tia تستخدمه لحساب نهاية الموعد.")
                    continue

                raw_price = _clean(row.get("price"))
                price = _parse_decimal(row.get("price"))
                if raw_price is not None and (price is None or price < 0):
                    skipped["services"] += 1
                    _issue(issues, sheet="services", row=row_number, message="price غير صالح؛ اتركه فارغًا أو اكتب سعرًا غير سالب.")
                    continue
                if price is None:
                    if current is None:
                        skipped["services"] += 1
                        _issue(issues, sheet="services", row=row_number, message="الخدمة الجديدة تحتاج price؛ اترك الخانة فاضية في الـdraft وكملها يدويًا قبل الحفظ.")
                        continue
                    price = Decimal(current.price_minor) / Decimal(100)

                category = _clean(row.get("category"))
                if category is None and current is not None:
                    category = current.category
                try:
                    payload = ClinicServiceUpdateV2(name=name, category=category, duration_minutes=duration, price=price)
                    if current is None:
                        create_service_v2(db, workspace=workspace, payload=ClinicServiceCreateV2(**payload.model_dump()))
                    else:
                        update_service_v2(db, workspace=workspace, service_id=current.id, payload=payload)
                    imported["services"] += 1
                except (ClinicSetupV2Error, ValidationError) as exc:
                    skipped["services"] += 1
                    _issue(issues, sheet="services", row=row_number, message=str(exc))

        # 3) Doctors: name identifies the row. doctor_type is optional and defaults safely.
        doctor_sheet = sheets.get("doctors")
        if doctor_sheet is not None:
            seen: set[str] = set()
            for row_number, row in _rows(doctor_sheet):
                name = _clean(row.get("full_name") or row.get("name"))
                if not name:
                    skipped["doctors"] += 1
                    _issue(issues, sheet="doctors", row=row_number, message="اسم الدكتور مطلوب لتحديد الصف؛ النوع والتخصص يمكن استكمالهما يدويًا.")
                    continue
                name_key = " ".join(name.split()).casefold()
                if name_key in seen:
                    skipped["doctors"] += 1
                    _issue(issues, sheet="doctors", row=row_number, message="اسم الدكتور مكرر داخل الملف.")
                    continue
                seen.add(name_key)
                existing = _doctor_rows(db, workspace.id).get(name_key)
                raw_type = (_clean(row.get("doctor_type")) or "").casefold()
                if raw_type and raw_type not in {"regular", "visiting"}:
                    skipped["doctors"] += 1
                    _issue(issues, sheet="doctors", row=row_number, message="doctor_type لو موجود لازم يكون regular أو visiting؛ ويمكن تركه فارغًا.")
                    continue
                if not raw_type and existing is None:
                    skipped["doctors"] += 1
                    _issue(issues, sheet="doctors", row=row_number, message="الدكتور الجديد يحتاج doctor_type؛ سيبها فاضية في الـdraft واختار ثابت أو زائر قبل الحفظ.")
                    continue
                doctor_type = raw_type or existing[0].doctor_type
                specialization = _clean(row.get("specialization"))
                if specialization is None and existing is not None:
                    specialization = existing[0].specialization
                try:
                    if existing is None:
                        create_doctor_v2(
                            db,
                            workspace=workspace,
                            payload=ClinicDoctorCreateV2(full_name=name, doctor_type=doctor_type, specialization=specialization, service_ids=[]),
                        )
                    else:
                        doctor, _staff = existing
                        update_doctor_v2(
                            db,
                            workspace=workspace,
                            doctor_id=doctor.id,
                            payload=ClinicDoctorUpdateV2(
                                full_name=name,
                                doctor_type=doctor_type,
                                specialization=specialization,
                                service_ids=_current_doctor_service_ids(db, workspace.id, doctor.id),
                            ),
                        )
                    imported["doctors"] += 1
                except (ClinicSetupV2Error, ValidationError) as exc:
                    skipped["doctors"] += 1
                    _issue(issues, sheet="doctors", row=row_number, message=str(exc))

        # Refresh lookup maps after the two entity sheets.
        service_map = _services(db, workspace.id)
        doctor_map = _doctor_rows(db, workspace.id)

        # 4) Doctor ↔ service links. Merge uploaded links with existing links, never silently remove others.
        doctor_service_sheet = sheets.get("doctor_services")
        if doctor_service_sheet is not None:
            grouped: dict[UUID, set[UUID]] = defaultdict(set)
            rows_by_doctor: dict[UUID, list[int]] = defaultdict(list)
            for row_number, row in _rows(doctor_service_sheet):
                doctor_name = _clean(row.get("doctor_name") or row.get("full_name"))
                service_name = _clean(row.get("service_name") or row.get("name"))
                doctor_entry = doctor_map.get((doctor_name or "").casefold())
                service = service_map.get((service_name or "").casefold())
                if doctor_entry is None or service is None:
                    skipped["doctor_services"] += 1
                    _issue(issues, sheet="doctor_services", row=row_number, message="الدكتور أو الخدمة غير موجودين؛ كمّلهم أولًا من نفس الصفحة.")
                    continue
                doctor = doctor_entry[0]
                grouped[doctor.id].add(service.id)
                rows_by_doctor[doctor.id].append(row_number)
                imported["doctor_services"] += 1
            for doctor_id, new_ids in grouped.items():
                try:
                    merged = set(_current_doctor_service_ids(db, workspace.id, doctor_id)) | new_ids
                    replace_doctor_services_v2(db, workspace=workspace, doctor_id=doctor_id, service_ids=list(merged))
                except (ClinicSetupV2Error, ValidationError) as exc:
                    imported["doctor_services"] -= len(rows_by_doctor[doctor_id])
                    skipped["doctor_services"] += len(rows_by_doctor[doctor_id])
                    _issue(issues, sheet="doctor_services", row=rows_by_doctor[doctor_id][0], message=str(exc))

        # 5) Clinic weekly hours. A present sheet replaces the weekly schedule with valid rows.
        clinic_hours_sheet = sheets.get("clinic_hours")
        if clinic_hours_sheet is not None:
            intervals: list[WorkingHourInputV2] = []
            for row_number, row in _rows(clinic_hours_sheet):
                weekday = _parse_weekday(row)
                start = _parse_time(row.get("start_time"))
                end = _parse_time(row.get("end_time"))
                try:
                    if weekday is None or start is None or end is None:
                        raise ValueError
                    intervals.append(WorkingHourInputV2(weekday=weekday, start_time=start, end_time=end))
                    imported["clinic_hours"] += 1
                except Exception:
                    skipped["clinic_hours"] += 1
                    _issue(issues, sheet="clinic_hours", row=row_number, message="اكتب day (مثل Monday أو الثلاثاء) أو weekday من 0 إلى 6، مع start_time/end_time صحيحين.")
            if intervals:
                try:
                    replace_clinic_hours_v2(db, workspace=workspace, payload=WorkingHoursUpdateV2(intervals=intervals))
                except ClinicSetupV2Error as exc:
                    imported["clinic_hours"] = 0
                    skipped["clinic_hours"] += len(intervals)
                    _issue(issues, sheet="clinic_hours", row=2, message=str(exc))

        # 6) Regular doctor weekly hours.
        doctor_hours_sheet = sheets.get("doctor_hours")
        if doctor_hours_sheet is not None:
            grouped_hours: dict[UUID, list[WorkingHourInputV2]] = defaultdict(list)
            for row_number, row in _rows(doctor_hours_sheet):
                doctor_name = _clean(row.get("doctor_name"))
                entry = doctor_map.get((doctor_name or "").casefold())
                weekday = _parse_weekday(row)
                start = _parse_time(row.get("start_time"))
                end = _parse_time(row.get("end_time"))
                if entry is None or entry[0].doctor_type != "regular" or weekday is None or start is None or end is None:
                    skipped["doctor_hours"] += 1
                    _issue(issues, sheet="doctor_hours", row=row_number, message="الصف يحتاج دكتور regular موجود وday أو weekday صحيح، مع start_time/end_time صحيحين.")
                    continue
                try:
                    grouped_hours[entry[0].id].append(WorkingHourInputV2(weekday=weekday, start_time=start, end_time=end))
                    imported["doctor_hours"] += 1
                except Exception:
                    skipped["doctor_hours"] += 1
                    _issue(issues, sheet="doctor_hours", row=row_number, message="وقت نهاية الدوام لازم يكون بعد البداية.")
            for doctor_id, intervals in grouped_hours.items():
                try:
                    replace_regular_doctor_hours_v2(db, workspace=workspace, doctor_id=doctor_id, payload=WorkingHoursUpdateV2(intervals=intervals))
                except (ClinicSetupV2Error, ValidationError) as exc:
                    imported["doctor_hours"] -= len(intervals)
                    skipped["doctor_hours"] += len(intervals)
                    _issue(issues, sheet="doctor_hours", row=2, message=str(exc))

        # 7) Visiting doctor dated windows.
        visiting_sheet = sheets.get("visiting_windows")
        if visiting_sheet is not None:
            grouped_windows: dict[UUID, list[VisitingWindowInputV2]] = defaultdict(list)
            now_utc = datetime.now(UTC)
            cairo = ZoneInfo("Africa/Cairo")
            for row_number, row in _rows(visiting_sheet):
                doctor_name = _clean(row.get("doctor_name"))
                entry = doctor_map.get((doctor_name or "").casefold())
                visit_date = _parse_date(row.get("date"))
                start = _parse_time(row.get("start_time"))
                end = _parse_time(row.get("end_time"))
                if entry is None or entry[0].doctor_type != "visiting" or visit_date is None or start is None or end is None:
                    skipped["visiting_windows"] += 1
                    _issue(issues, sheet="visiting_windows", row=row_number, message="الصف يحتاج دكتور visiting موجود وdate/start_time/end_time صحيحة.")
                    continue
                try:
                    item = VisitingWindowInputV2(date=visit_date, start_time=start, end_time=end)
                    if datetime.combine(visit_date, end, tzinfo=cairo).astimezone(UTC) <= now_utc:
                        raise ValueError
                    grouped_windows[entry[0].id].append(item)
                    imported["visiting_windows"] += 1
                except Exception:
                    skipped["visiting_windows"] += 1
                    _issue(issues, sheet="visiting_windows", row=row_number, message="زيارة الدكتور لازم تنتهي في وقت مستقبلي وبعد وقت البداية.")
            for doctor_id, windows in grouped_windows.items():
                try:
                    replace_visiting_windows_v2(db, workspace=workspace, doctor_id=doctor_id, payload=VisitingWindowsUpdateV2(windows=windows))
                except (ClinicSetupV2Error, ValidationError) as exc:
                    imported["visiting_windows"] -= len(windows)
                    skipped["visiting_windows"] += len(windows)
                    _issue(issues, sheet="visiting_windows", row=2, message=str(exc))

        # 8) Booking policy. Missing settings preserve current values and can be edited manually.
        policy_sheet = sheets.get("booking_policy")
        if policy_sheet is not None:
            values = _vertical(policy_sheet, key_column="setting", value_column="value")
            snapshot = build_setup_v2_snapshot(db, workspace=workspace)
            current = snapshot.booking_policy
            kwargs = dict(current)
            parsers = {
                "slot_interval_minutes": _parse_int,
                "minimum_notice_minutes": _parse_int,
                "booking_horizon_days": _parse_int,
                "cancellation_notice_minutes": _parse_int,
                "allow_same_day_booking": _parse_bool,
                "require_confirmation": _parse_bool,
            }
            for setting, (row_number, value) in values.items():
                parser = parsers.get(setting)
                if parser is None:
                    skipped["booking_policy"] += 1
                    _issue(issues, sheet="booking_policy", row=row_number, message=f"إعداد غير معروف: {setting}.")
                    continue
                parsed = parser(value)
                if parsed is None:
                    skipped["booking_policy"] += 1
                    _issue(issues, sheet="booking_policy", row=row_number, message=f"قيمة {setting} غير صالحة.")
                    continue
                kwargs[setting] = parsed
                imported["booking_policy"] += 1
            try:
                update_booking_policy_v2(db, workspace=workspace, payload=BookingPolicyUpdateV2(**kwargs))
            except Exception as exc:
                _issue(issues, sheet="booking_policy", row=2, message=f"تعذر حفظ سياسة الحجز: {exc}")

        db.flush()
        snapshot = build_setup_v2_snapshot(db, workspace=workspace)
        return ClinicSetupImportResponse(
            imported_counts=dict(imported),
            skipped_counts=dict(skipped),
            issues=issues,
            snapshot=snapshot,
        )
    finally:
        workbook.close()


def build_clinic_setup_template() -> bytes:
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    readme.append(["Tia Clinic Setup v1"])
    readme.append(["ارفع الملف من صفحة إعدادات العيادة. أسماء الـsheets ثابتة، لكن الأعمدة الاختيارية يمكن حذفها. Tia تقرأ المتاح وتترك غير الضروري للاستكمال اليدوي."])
    readme.append(["العملة ثابتة EGP. لا يوجد branch_id لأن العيادة فرع واحد."])

    sheets: dict[str, list[str]] = {
        "clinic_profile": ["field", "value"],
        "services": ["name", "category", "duration_minutes", "price"],
        "doctors": ["full_name", "doctor_type", "specialization"],
        "doctor_services": ["doctor_name", "service_name"],
        "clinic_hours": ["day", "start_time", "end_time"],
        "doctor_hours": ["doctor_name", "day", "start_time", "end_time"],
        "visiting_windows": ["doctor_name", "date", "start_time", "end_time"],
        "booking_policy": ["setting", "value"],
    }
    for name, headers in sheets.items():
        ws = wb.create_sheet(name)
        ws.append(headers)

    profile = wb["clinic_profile"]
    profile.append(["clinic_name", ""])
    profile.append(["phone", ""])
    profile.append(["address", ""])
    profile.append(["city", ""])

    wb["booking_policy"].append(["slot_interval_minutes", ""])
    wb["booking_policy"].append(["minimum_notice_minutes", ""])
    wb["booking_policy"].append(["booking_horizon_days", ""])
    wb["booking_policy"].append(["cancellation_notice_minutes", ""])
    wb["booking_policy"].append(["allow_same_day_booking", ""])
    wb["booking_policy"].append(["require_confirmation", ""])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
